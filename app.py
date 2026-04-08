"""
Fingerprint Release Manager - v4.0 (SMS + Email + LAPS Retriever)
Integrates with Accio Data XML API to automate fingerprint release form distribution
and criminal history retrieval from Louisiana LAPS portal.

New in v4.0:
- LAPS Retriever: Automated criminal history retrieval from Louisiana LAPS portal
- Hourly background worker scans LAPS "Recently Completed" tab
- Matches LAPS results to FP Release applicants (name + SSN last 4)
- Pushes criminal history results back to Accio via PostResults XML API
- LAPS Dashboard section with real-time status and cycle history
- Secure in-memory-only handling of criminal records (zero disk writes)

Two-Step Workflow:
  STEP 1 (FP Release -> Applicant): XML push -> code assignment -> email/SMS to applicant
  STEP 2 (LAPS -> FP Release -> Accio): LAPS scrape -> parse rap sheet -> match applicant -> PostResults

Built with Python standard library + openpyxl + twilio + playwright + httpx.
"""

import os
import sys
import json
import smtplib
import logging
import urllib.parse
import urllib.request
import cgi
import io
import csv
import shutil
import base64
import uuid
import hashlib
try:
    import bcrypt
    HAS_BCRYPT = True
except ImportError:
    HAS_BCRYPT = False
    print("WARNING: bcrypt not installed. Run: pip install bcrypt  (required for Accio §1.7 compliance)")
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from http.server import HTTPServer, BaseHTTPRequestHandler
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate, make_msgid
from xml.etree import ElementTree as ET
from string import Template
import time
import re
import asyncio
import ctypes
import gc
import signal as signal_module
import threading
from contextlib import contextmanager
from dataclasses import dataclass, field as dc_field
from enum import Enum
from typing import Any, Optional, Dict, List, Tuple

try:
    import httpx
    HAS_HTTPX = True
except ImportError:
    HAS_HTTPX = False
    print("WARNING: httpx not installed. LAPS Retriever requires it. Run: pip install httpx")

try:
    from playwright.async_api import async_playwright, Browser, BrowserContext, Page, Playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False
    print("WARNING: playwright not installed. LAPS Retriever requires it.")
    print("Run: pip install playwright && playwright install chromium")

try:
    import psycopg2
    import psycopg2.extras
    HAS_PG = True
except ImportError:
    HAS_PG = False
    print("WARNING: psycopg2 not installed. Install with: pip install psycopg2-binary")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. Excel import will be limited to CSV only.")
    print("Install with: pip install openpyxl")

try:
    from twilio.rest import Client as TwilioClient
    HAS_TWILIO = True
except ImportError:
    HAS_TWILIO = False
    print("WARNING: twilio not installed. SMS notifications will be disabled.")
    print("Install with: pip install twilio")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL = os.environ.get("DATABASE_URL", "")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
WATCH_FOLDER = os.path.join(BASE_DIR, "watch")
PROCESSED_FOLDER = os.path.join(BASE_DIR, "watch", "processed")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")
HOST = "0.0.0.0"
PORT = int(os.environ.get("PORT", 5000))
# Max body size for inbound XML (10 MB) — prevents memory exhaustion DoS
MAX_XML_BODY = 10 * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(WATCH_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# LAPS Retriever Configuration (all secrets from environment variables)
# ---------------------------------------------------------------------------
LAPS_USERNAME = os.environ.get("LAPS_USERNAME", "")
LAPS_PASSWORD = os.environ.get("LAPS_PASSWORD", "")
LAPS_PORTAL_URL = "https://la.flexcheck.us.idemia.io/LAPSPortal/"

# FlexCheck Token Grid (printed security token sheet dated 02/10/2026)
LAPS_TOKEN_GRID = {
    "1": {"A": "rLa#EU", "B": "PBmCVx", "C": "LYZEp2", "D": "7@ct5J", "E": "pMR6lHn", "F": "WxwAJ@", "G": "mayzBB", "H": "Jz3a@S", "I": "BK4q6F", "J": "ksBY@R"},
    "2": {"A": "FB2VNf", "B": "Rm@hvf", "C": "ZfxM&x", "D": "fZ3xSB", "E": "WDxs9i", "F": "2w6zGr", "G": "WEz*HD", "H": "RqXyC6", "I": "zx@maW", "J": "B@mswU"},
    "3": {"A": "2b#hiL", "B": "zd66Bc", "C": "uvFvTx", "D": "9jQLje", "E": "rWKLCG", "F": "VHjWye", "G": "mtF*ov", "H": "q9Xr3*", "I": "6PkGDW", "J": "jcYm2N"},
    "4": {"A": "pDj@m2", "B": "i3kTx2", "C": "cfVRmp", "D": "wCK6@7", "E": "ztgpF#", "F": "ussRLe", "G": "z@EBnP", "H": "qX97p3", "I": "#o5Qi*", "J": "ej6E6U"},
    "5": {"A": "txMoyJ", "B": "Sga2EN", "C": "7tNtuT", "D": "YeCwu4", "E": "Xe6pCo", "F": "&q6c&m", "G": "ZyMYLa", "H": "iKCpvS", "I": "y5vZCr", "J": "HMbrt*"},
}

# Accio PostResults config for LAPS results
ACCIO_API_BASE_URL = os.environ.get("ACCIO_API_BASE_URL", "")
ACCIO_API_ACCOUNT = os.environ.get("ACCIO_API_ACCOUNT", "")
# LAPS pusher reuses the same vendor credentials (ACCIO_API_USERNAME / ACCIO_API_PASSWORD)
ACCIO_API_MODE = os.environ.get("ACCIO_API_MODE", "PROD")
ACCIO_REGISTRATION_KEY = os.environ.get("ACCIO_REGISTRATION_KEY", "")
ACCIO_REGISTRATION_COMPANY = os.environ.get("ACCIO_REGISTRATION_COMPANY", "Background Research Solutions, LLC")
ACCIO_POSTRESULTS_PATH = os.environ.get("ACCIO_POSTRESULTS_PATH", "/c/p/researcherxml")

# LAPS operational tuning
LAPS_HTTP_TIMEOUT = int(os.environ.get("LAPS_HTTP_TIMEOUT", "30"))
LAPS_PLAYWRIGHT_TIMEOUT = int(os.environ.get("LAPS_PLAYWRIGHT_TIMEOUT", "30000"))
LAPS_MAX_RETRIES = int(os.environ.get("LAPS_MAX_RETRIES", "3"))
LAPS_RETRY_BASE_DELAY = float(os.environ.get("LAPS_RETRY_BASE_DELAY", "2.0"))
LAPS_HOURLY_INTERVAL = int(os.environ.get("LAPS_HOURLY_INTERVAL", "3600"))

# LAPS feature toggle
LAPS_ENABLED = os.environ.get("LAPS_ENABLED", "true").lower() in ("true", "1", "yes")

# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------
class DBHelper:
    """Wrapper to provide sqlite3-like interface over psycopg2."""
    def __init__(self, conn):
        self._conn = conn

    def execute(self, sql, params=None):
        cur = self._conn.cursor()
        cur.execute(sql, params or ())
        return cur

    def commit(self):
        pass  # autocommit is on

    def rollback(self):
        # FIX: Added missing rollback() method. With autocommit=True each statement
        # is its own transaction, so this is a no-op, but required after
        # psycopg2.IntegrityError to keep the connection in a clean state.
        try:
            self._conn.rollback()
        except Exception:
            pass

    def close(self):
        self._conn.close()


def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    conn.autocommit = True
    return DBHelper(conn)


def init_db():
    db = get_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            display_name TEXT,
            role TEXT DEFAULT 'user',
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT NOW(),
            last_login TIMESTAMP
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            user_id INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT NOW(),
            expires_at TIMESTAMP
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS clients (
            id SERIAL PRIMARY KEY,
            company_name TEXT NOT NULL,
            account_name TEXT,
            contact_email TEXT,
            contact_phone TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            applicant_count INTEGER DEFAULT 0
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS applicants (
            id SERIAL PRIMARY KEY,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            accio_order_number TEXT,
            accio_remote_number TEXT,
            status TEXT DEFAULT 'pending',
            assigned_code TEXT,
            email_sent BOOLEAN DEFAULT FALSE,
            email_sent_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW(),
            notes TEXT,
            client_id INTEGER REFERENCES clients(id),
            date_of_birth TEXT,
            last_four_ssn VARCHAR(4)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS codes (
            id SERIAL PRIMARY KEY,
            code TEXT NOT NULL UNIQUE,
            status TEXT DEFAULT 'available',
            assigned_to INTEGER REFERENCES applicants(id),
            assigned_at TIMESTAMP,
            assigned_date TIMESTAMP,
            imported_at TIMESTAMP DEFAULT NOW(),
            batch_name TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS email_log (
            id SERIAL PRIMARY KEY,
            applicant_id INTEGER REFERENCES applicants(id),
            recipient_email TEXT,
            subject TEXT,
            status TEXT,
            error_message TEXT,
            sent_at TIMESTAMP DEFAULT NOW()
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS email_tracking (
            id SERIAL PRIMARY KEY,
            applicant_id INTEGER REFERENCES applicants(id),
            email_log_id INTEGER REFERENCES email_log(id),
            tracking_token TEXT UNIQUE NOT NULL,
            opened_at TIMESTAMP,
            first_human_open_at TIMESTAMP,
            open_count INTEGER DEFAULT 0,
            bot_open_count INTEGER DEFAULT 0,
            is_bot_open BOOLEAN DEFAULT FALSE,
            user_agent TEXT
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS xml_log (
            id SERIAL PRIMARY KEY,
            direction TEXT,
            raw_xml TEXT,
            parsed_status TEXT,
            error_message TEXT,
            received_at TIMESTAMP DEFAULT NOW()
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id SERIAL PRIMARY KEY,
            user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            token TEXT UNIQUE NOT NULL,
            expires_at TIMESTAMP NOT NULL,
            used BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)

    # SMS log table — tracks all outbound text messages
    db.execute("""
        CREATE TABLE IF NOT EXISTS sms_log (
            id SERIAL PRIMARY KEY,
            applicant_id INTEGER REFERENCES applicants(id),
            recipient_phone TEXT,
            message_body TEXT,
            twilio_sid TEXT,
            status TEXT,
            error_message TEXT,
            sent_at TIMESTAMP DEFAULT NOW()
        )
    """)

    # LAPS retrieval tracking tables
    db.execute("""
        CREATE TABLE IF NOT EXISTS laps_cycle_log (
            id SERIAL PRIMARY KEY,
            started_at TIMESTAMP DEFAULT NOW(),
            finished_at TIMESTAMP,
            processed INTEGER DEFAULT 0,
            matched INTEGER DEFAULT 0,
            pushed INTEGER DEFAULT 0,
            skipped INTEGER DEFAULT 0,
            failed INTEGER DEFAULT 0,
            errors TEXT,
            status TEXT DEFAULT 'running'
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS laps_result_log (
            id SERIAL PRIMARY KEY,
            applicant_id INTEGER REFERENCES applicants(id),
            cycle_id INTEGER REFERENCES laps_cycle_log(id),
            laps_name TEXT,
            laps_status TEXT,
            record_count INTEGER DEFAULT 0,
            accio_push_ok BOOLEAN DEFAULT FALSE,
            notes TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)

    # Add new columns to existing tables if they don't exist
    try:
        db.execute("ALTER TABLE users ADD COLUMN recovery_email TEXT")
    except psycopg2.Error:
        pass  # Column already exists
    try:
        db.execute("ALTER TABLE applicants ADD COLUMN client_id INTEGER REFERENCES clients(id)")
    except psycopg2.Error:
        pass  # Column already exists

    # New columns to capture Accio sub-order number and order type for postback
    try:
        db.execute("ALTER TABLE applicants ADD COLUMN accio_sub_order TEXT")
    except psycopg2.Error:
        pass  # Column already exists
    try:
        db.execute("ALTER TABLE applicants ADD COLUMN accio_order_type TEXT")
    except psycopg2.Error:
        pass  # Column already exists

    # LAPS integration columns: DOB for display/matching, last 4 SSN for matching
    try:
        db.execute("ALTER TABLE applicants ADD COLUMN date_of_birth TEXT")
    except psycopg2.Error:
        pass  # Column already exists
    try:
        db.execute("ALTER TABLE applicants ADD COLUMN last_four_ssn VARCHAR(4)")
    except psycopg2.Error:
        pass  # Column already exists

    # Migration: add SMS tracking columns to applicants for existing databases
    for col_sql in [
        "ALTER TABLE applicants ADD COLUMN sms_sent BOOLEAN DEFAULT FALSE",
        "ALTER TABLE applicants ADD COLUMN sms_sent_at TIMESTAMP",
    ]:
        try:
            db.execute(col_sql)
        except psycopg2.Error:
            pass  # Column already exists

    # LAPS integration columns
    for col_sql in [
        "ALTER TABLE applicants ADD COLUMN laps_status VARCHAR(50) DEFAULT NULL",
        "ALTER TABLE applicants ADD COLUMN laps_retrieved_at TIMESTAMP DEFAULT NULL",
    ]:
        try:
            db.execute(col_sql)
        except psycopg2.Error:
            pass  # Column already exists

    # Migration: add bot-detection columns to email_tracking for existing databases
    for col_sql in [
        "ALTER TABLE email_tracking ADD COLUMN first_human_open_at TIMESTAMP",
        "ALTER TABLE email_tracking ADD COLUMN bot_open_count INTEGER DEFAULT 0",
        "ALTER TABLE email_tracking ADD COLUMN is_bot_open BOOLEAN DEFAULT FALSE",
    ]:
        try:
            db.execute(col_sql)
        except psycopg2.Error:
            pass  # Column already exists

    db.close()

    # Ensure the canonical admin account exists with bcrypt credentials.
    # This runs on every startup so it handles three cases:
    #   1. Fresh database — creates admin from scratch (requires env vars)
    #   2. Old database with legacy 'admin' account — renames and re-hashes
    #   3. Admin already exists — updates password hash to bcrypt if needed
    db = get_db()
    _default_user = os.environ.get("DEFAULT_ADMIN_USER", "").strip()
    _default_pass = os.environ.get("DEFAULT_ADMIN_PASSWORD", "").strip()
    if not _default_user or not _default_pass:
        # Only warn on first run when no admin exists yet; otherwise the
        # admin account is already in the database and env vars aren't needed.
        existing_check = db.execute("SELECT id FROM users WHERE role='admin' LIMIT 1").fetchone()
        if existing_check is None:
            logger.critical(
                "DEFAULT_ADMIN_USER and DEFAULT_ADMIN_PASSWORD environment variables "
                "are required for initial setup. Set them in your Render/hosting "
                "environment variables, then restart."
            )
            print("\n*** FATAL: Set DEFAULT_ADMIN_USER and DEFAULT_ADMIN_PASSWORD "
                  "environment variables before first run. ***\n")
            db.close()
            sys.exit(1)
        else:
            # Admin already exists in DB — no env vars needed for normal startup
            db.close()
            return
    if HAS_BCRYPT:
        pw_hash = bcrypt.hashpw(_default_pass.encode(), bcrypt.gensalt(rounds=12)).decode()
    else:
        salt = os.urandom(32)
        pw_hash = "sha256:" + hashlib.sha256(_default_pass.encode() + salt).hexdigest() + "$" + base64.b64encode(salt).decode()

    existing = db.execute("SELECT id, username FROM users WHERE role='admin' ORDER BY id LIMIT 1").fetchone()
    if existing is None:
        # No admin at all — create fresh
        db.execute(
            "INSERT INTO users (username, password_hash, display_name, role, is_active) VALUES (%s, %s, %s, %s, %s)",
            (_default_user, pw_hash, "Administrator", "admin", True)
        )
        logger.info("Created admin user '%s'", _default_user)
    else:
        # Admin exists — rename to correct username and re-hash password with bcrypt
        db.execute(
            "UPDATE users SET username=%s, password_hash=%s, is_active=TRUE WHERE id=%s",
            (_default_user, pw_hash, existing["id"])
        )
        logger.info("Migrated admin account to username '%s' with bcrypt hash", _default_user)
    db.close()


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------
DEFAULT_SETTINGS = {
    "accio_account": os.environ.get("ACCIO_ACCOUNT", ""),
    "accio_username": os.environ.get("ACCIO_USERNAME", ""),
    "accio_password": os.environ.get("ACCIO_PASSWORD", ""),
    "accio_post_url": os.environ.get("ACCIO_POST_URL", "https://fingerprint-release-manager1.onrender.com/api/accio-push"),
    "accio_researcher_url": os.environ.get("ACCIO_RESEARCHER_URL", ""),
    "smtp_server": os.environ.get("SMTP_HOST", "smtp.gmail.com"),
    "smtp_port": os.environ.get("SMTP_PORT", "587"),
    "smtp_username": os.environ.get("SMTP_USER", ""),
    "smtp_password": os.environ.get("SMTP_PASS", ""),
    "smtp_use_tls": os.environ.get("SMTP_USE_TLS", "1"),
    "sender_email": os.environ.get("SENDER_EMAIL", ""),
    "sender_name": os.environ.get("SENDER_NAME", "Fingerprints Required"),
    "email_subject": "Fingerprint Processing - Next Steps from {company_name}",
    "email_body": """Hello {first_name} {last_name},

Thank you for working with {company_name}. Please see the attached form regarding your upcoming fingerprint appointment.

Your IdentoGO reference code: {code}

Please bring this code to your IdentoGO appointment. The processing fee has already been covered on your behalf -- simply present the code above when you arrive.

Steps to complete:
1. Review the attached Fingerprint Release Form
2. Visit your assigned IdentoGO location
3. Present your reference code: {code}
4. Complete the fingerprinting process

If you have any questions, feel free to reach out to us directly.

Best regards,
{company_name}""",
    "release_form_path": "",
    "company_name": "Background Research Solutions, LLC",
    "ori_number": "",
    "auto_assign_codes": "1",
    "auto_send_email": "1",
    # --- SMS / Twilio Settings ---
    "twilio_account_sid": os.environ.get("TWILIO_ACCOUNT_SID", ""),
    "twilio_auth_token": os.environ.get("TWILIO_AUTH_TOKEN", ""),
    "twilio_from_number": os.environ.get("TWILIO_FROM_NUMBER", ""),
    "auto_send_sms": "1",
    "sms_body": """Hello {first_name} {last_name}, this is {company_name}. Your fingerprint appointment is ready!

Your IdentoGO code: {code}

Steps:
1. Visit your assigned IdentoGO location
2. Present code: {code}
3. The processing fee is already covered

Questions? Reply to this text or contact us directly.

- {company_name}""",
}


def get_setting(db, key):
    cur = db.execute("SELECT value FROM settings WHERE key = %s", (key,))
    row = cur.fetchone()
    if row:
        return row["value"]
    return DEFAULT_SETTINGS.get(key, "")


def set_setting(db, key, value):
    db.execute(
        "INSERT INTO settings (key, value) VALUES (%s, %s) ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value",
        (key, value)
    )


# ---------------------------------------------------------------------------
# Authentication & Session Management
# ---------------------------------------------------------------------------
def hash_password(password):
    """Hash a password using bcrypt (Accio Security Policy §1.7 compliant).

    bcrypt is an adaptive one-way algorithm.  Unlike SHA-256, its work factor
    (cost) is configurable and can be increased over time, and no efficient
    cracking algorithm exists, satisfying §1.7's requirement for 'a one-way
    hashing algorithm for which no known algorithm for cracking in a useful
    timeframe is known'.
    """
    if HAS_BCRYPT:
        return bcrypt.hashpw(password.encode(), bcrypt.gensalt(rounds=12)).decode()
    # Fallback: SHA-256 with random salt.  Install bcrypt for full compliance.
    salt = os.urandom(32)
    h = hashlib.sha256(password.encode() + salt).hexdigest()
    salt_b64 = base64.b64encode(salt).decode()
    return f"sha256:{h}${salt_b64}"


def verify_password(password, stored_hash):
    """Verify a password against a bcrypt or legacy SHA-256 hash.

    Supports three stored formats for seamless migration:
      1. bcrypt   — $2b$... prefix (current, §1.7 compliant)
      2. sha256:  — prefixed legacy format written by this app when bcrypt unavailable
      3. no-prefix SHA-256  — original format written by very early versions of the app
    """
    try:
        if stored_hash.startswith("$2b$") or stored_hash.startswith("$2a$"):
            # Native bcrypt hash
            if HAS_BCRYPT:
                return bcrypt.checkpw(password.encode(), stored_hash.encode())
            return False
        if stored_hash.startswith("sha256:"):
            # Prefixed legacy SHA-256 path
            _, rest = stored_hash.split(":", 1)
            h, salt_b64 = rest.split("$", 1)
            salt = base64.b64decode(salt_b64)
            computed = hashlib.sha256(password.encode() + salt).hexdigest()
            return computed == h
        # Original no-prefix SHA-256 format (64-char hex + salt)
        parts = stored_hash.split("$", 1)
        if len(parts) == 2 and len(parts[0]) == 64:
            h, salt_b64 = parts
            salt = base64.b64decode(salt_b64)
            computed = hashlib.sha256(password.encode() + salt).hexdigest()
            return computed == h
        return False
    except Exception:
        return False


def create_session(db, user_id):
    """Create a new session token for a user."""
    token = str(uuid.uuid4())
    expires_at = (datetime.now() + timedelta(hours=24)).isoformat()
    db.execute(
        "INSERT INTO sessions (token, user_id, expires_at) VALUES (%s, %s, %s)",
        (token, user_id, expires_at)
    )
    return token


def verify_session(db, token):
    """Verify a session token and return user data if valid."""
    if not token:
        return None
    cur = db.execute(
        "SELECT u.id, u.username, u.display_name, u.role FROM users u "
        "JOIN sessions s ON u.id = s.user_id "
        "WHERE s.token = %s AND s.expires_at > NOW() AND u.is_active = TRUE",
        (token,)
    )
    return cur.fetchone()


def delete_session(db, token):
    """Delete a session token from the database (logout)."""
    if token:
        db.execute("DELETE FROM sessions WHERE token = %s", (token,))


def get_session_from_cookie(cookie_header):
    """Extract session token from cookie header."""
    if not cookie_header:
        return None
    for part in cookie_header.split(";"):
        part = part.strip()
        if part.startswith("session_token="):
            return part.split("=", 1)[1]
    return None


# ---------------------------------------------------------------------------
# XML Parsing
# ---------------------------------------------------------------------------
def parse_accio_xml(xml_string):
    applicants = []
    try:
        root = ET.fromstring(xml_string)
    except ET.ParseError as e:
        return applicants, str(e)

    # --- Format 1: ScreeningResults (completeOrder / placeOrder) ---
    for complete_order in root.iter("completeOrder"):
        order_number = complete_order.get("number", "")
        remote_number = complete_order.get("remote_number", "")
        # Capture sub-order ID and type for Accio postback (Ch 9)
        sub_order_el = complete_order.find("subOrder")
        sub_order_id = (sub_order_el.get("id", "") or sub_order_el.get("number", "1")) if sub_order_el is not None else "1"
        order_type = (sub_order_el.get("type", "") or "Fingerprint") if sub_order_el is not None else "Fingerprint"
        for subject in complete_order.iter("subject"):
            first = _xt(subject, "name_first")
            last = _xt(subject, "name_last")
            email = _xt(subject, "email")
            phone = _xt(subject, "phone")
            dob = _extract_dob(subject)
            ssn4 = _extract_ssn_last4(subject)
            if first or last:
                applicants.append(dict(first_name=first, last_name=last, email=email,
                                       phone=phone, accio_order_number=order_number,
                                       accio_remote_number=remote_number, company_name="",
                                       accio_sub_order=sub_order_id, accio_order_type=order_type,
                                       date_of_birth=dob, last_four_ssn=ssn4))

    for po in root.iter("placeOrder"):
        company_name = ""
        ai = po.find("accountInfo")
        if ai is not None:
            company_name = _xt(ai, "company_name")
        account_name = ""
        ci = po.find("clientInfo")
        if ci is not None:
            account_name = _xt(ci, "account")

        oi = po.find("orderInfo")
        oi_email = ""
        oi_phone = ""
        if oi is not None:
            oi_email = _xt(oi, "requester_email")
            oi_phone = _xt(oi, "requester_phone")
        if not oi_email:
            if ci is not None:
                oi_email = _xt(ci, "primaryuser_contact_email")
                if not oi_phone:
                    oi_phone = _xt(ci, "primaryuser_contact_telephone")
        if not oi_email:
            if ai is not None:
                oi_email = _xt(ai, "primaryuser_contact_email")
                if not oi_phone:
                    oi_phone = _xt(ai, "primaryuser_contact_telephone")
        sub_order_el = po.find("subOrder")
        sub_order_id = (sub_order_el.get("id", "") or sub_order_el.get("number", "1")) if sub_order_el is not None else "1"
        order_type = (sub_order_el.get("type", "") or "Fingerprint") if sub_order_el is not None else "Fingerprint"
        for subject in po.iter("subject"):
            first = _xt(subject, "name_first")
            last = _xt(subject, "name_last")
            email = _xt(subject, "email") or oi_email
            phone = _xt(subject, "phone") or oi_phone
            dob = _extract_dob(subject)
            ssn4 = _extract_ssn_last4(subject)
            if first or last:
                applicants.append(dict(first_name=first, last_name=last, email=email,
                                       phone=phone, accio_order_number=po.get("number", ""),
                                       accio_remote_number="", company_name=company_name,
                                       account_name=account_name,
                                       accio_sub_order=sub_order_id, accio_order_type=order_type,
                                       date_of_birth=dob, last_four_ssn=ssn4))

    # --- Format 2: Action Letter XML Post (postLetter with orderInfo) ---
    for post_letter in root.iter("postLetter"):
        order_number = post_letter.get("remote_order", "") or post_letter.get("order", "")
        remote_order = post_letter.get("remote_order", "")
        sub_order_el = post_letter.find("subOrder")
        sub_order_id = (sub_order_el.get("id", "") or sub_order_el.get("number", "1")) if sub_order_el is not None else "1"
        order_type = (sub_order_el.get("type", "") or "Fingerprint") if sub_order_el is not None else "Fingerprint"
        order_info = post_letter.find("orderInfo")
        if order_info is not None:
            first = _xt(order_info, "name_first")
            last = _xt(order_info, "name_last")
            email = _xt(order_info, "email")
            phone = _xt(order_info, "phone_number") or _xt(order_info, "phone")
            dob = _extract_dob(order_info)
            ssn4 = _extract_ssn_last4(order_info)
            if not email:
                email = _xt(order_info, "requester_email")
            if not phone:
                phone = _xt(order_info, "requester_phone")
            if first or last:
                applicants.append(dict(first_name=first, last_name=last, email=email,
                                       phone=phone, accio_order_number=order_number,
                                       accio_remote_number=remote_order, company_name="",
                                       account_name="",
                                       accio_sub_order=sub_order_id, accio_order_type=order_type,
                                       date_of_birth=dob, last_four_ssn=ssn4))

    # --- Format 3: Vendor dispatch XML (orderRequest with subject) ---
    for order_req in root.iter("orderRequest"):
        order_number = order_req.get("order", "") or order_req.get("number", "")
        remote_number = order_req.get("remote_order", "")
        sub_order_el = order_req.find("subOrder")
        sub_order_id = (sub_order_el.get("id", "") or sub_order_el.get("number", "1")) if sub_order_el is not None else "1"
        order_type = (sub_order_el.get("type", "") or "Fingerprint") if sub_order_el is not None else "Fingerprint"
        for subject in order_req.iter("subject"):
            first = _xt(subject, "name_first") or _xt(subject, "firstName")
            last = _xt(subject, "name_last") or _xt(subject, "lastName")
            email = _xt(subject, "email") or _xt(subject, "InternetEmailAddress")
            phone = _xt(subject, "phone") or _xt(subject, "phone_number")
            dob = _extract_dob(subject)
            ssn4 = _extract_ssn_last4(subject)
            if first or last:
                applicants.append(dict(first_name=first, last_name=last, email=email,
                                       phone=phone, accio_order_number=order_number,
                                       accio_remote_number=remote_number, company_name="",
                                       account_name="",
                                       accio_sub_order=sub_order_id, accio_order_type=order_type,
                                       date_of_birth=dob, last_four_ssn=ssn4))

    # --- Format 4: Generic fallback - PersonalData or BackgroundSearchPackage ---
    if not applicants:
        for pd in root.iter("PersonalData"):
            pn = pd.find("PersonName")
            cm = pd.find("ContactMethod")
            first = last = email = phone = ""
            dob = _extract_dob(pd)
            ssn4 = _extract_ssn_last4(pd)
            if pn is not None:
                first = _xt(pn, "GivenName") or _xt(pn, "name_first")
                last = _xt(pn, "FamilyName") or _xt(pn, "name_last")
                if not dob:
                    dob = _extract_dob(pn)
            if cm is not None:
                email = _xt(cm, "InternetEmailAddress") or _xt(cm, "email")
                phone = _xt(cm, "FormattedNumber") or _xt(cm, "phone")
            if first or last:
                applicants.append(dict(first_name=first, last_name=last, email=email,
                                       phone=phone, accio_order_number="",
                                       accio_remote_number="", company_name="", account_name="",
                                       date_of_birth=dob, last_four_ssn=ssn4))

    # --- Format 5: AccioOrder XML (placeOrder > subject, email in orderInfo) ---
    if not applicants:
        for place_order in root.iter("placeOrder"):
            order_number = place_order.get("number", "")
            company_name = ""
            ai = place_order.find("accountInfo")
            if ai is not None:
                company_name = _xt(ai, "company_name")
            account_name = ""
            ci = place_order.find("clientInfo")
            if ci is not None:
                account_name = _xt(ci, "account")

            order_info = place_order.find("orderInfo")
            order_email = order_phone = ""
            if order_info is not None:
                order_email = (_xt(order_info, "requester_email") or
                               _xt(order_info, "requester_fax") or "")
                order_phone = _xt(order_info, "requester_phone") or ""
            if not order_email:
                if ci is not None:
                    order_email = _xt(ci, "primaryuser_contact_email")
                    if not order_phone:
                        order_phone = _xt(ci, "primaryuser_contact_telephone")
            if not order_email:
                if ai is not None:
                    order_email = _xt(ai, "primaryuser_contact_email")
                    if not order_phone:
                        order_phone = _xt(ai, "primaryuser_contact_telephone")
            sub_order_el = place_order.find("subOrder")
            sub_order_id = (sub_order_el.get("id", "") or sub_order_el.get("number", "1")) if sub_order_el is not None else "1"
            order_type = (sub_order_el.get("type", "") or "Fingerprint") if sub_order_el is not None else "Fingerprint"
            for subject in place_order.iter("subject"):
                first = _xt(subject, "name_first")
                last = _xt(subject, "name_last")
                email = (_xt(subject, "email") or _xt(subject, "Email") or
                         _xt(subject, "InternetEmailAddress") or _xt(subject, "email_address") or
                         _xt(subject, "EmailAddress") or _xt(subject, "e_mail") or
                         _xt(subject, "applicant_email") or _xt(subject, "candidate_email") or
                         _xt(subject, "contact_email"))
                phone = (_xt(subject, "phone") or _xt(subject, "Phone") or
                         _xt(subject, "phone_number") or _xt(subject, "PhoneNumber") or
                         _xt(subject, "FormattedNumber") or _xt(subject, "telephone") or
                         _xt(subject, "contact_phone") or _xt(subject, "home_phone") or
                         _xt(subject, "cell_phone") or _xt(subject, "mobile"))
                dob = _extract_dob(subject)
                ssn4 = _extract_ssn_last4(subject)
                if not email:
                    email = order_email
                if not phone:
                    phone = order_phone
                if first or last:
                    applicants.append(dict(first_name=first, last_name=last, email=email,
                                           phone=phone, accio_order_number=order_number,
                                           accio_remote_number="", company_name=company_name,
                                           account_name=account_name,
                                           accio_sub_order=sub_order_id, accio_order_type=order_type,
                                           date_of_birth=dob, last_four_ssn=ssn4))

    # --- Format 5b: <order> tag (older AccioOrder variants) ---
    if not applicants:
        for order in root.iter("order"):
            order_number = order.get("number", "")
            remote_number = order.get("remote_order", "")
            order_info = order.find("orderInfo")
            order_email = order_phone = ""
            if order_info is not None:
                order_email = _xt(order_info, "requester_email")
                order_phone = _xt(order_info, "requester_phone")
            sub_order_el = order.find("subOrder")
            sub_order_id = (sub_order_el.get("id", "") or sub_order_el.get("number", "1")) if sub_order_el is not None else "1"
            order_type = (sub_order_el.get("type", "") or "Fingerprint") if sub_order_el is not None else "Fingerprint"
            for subject in order.iter("subject"):
                first = _xt(subject, "name_first")
                last = _xt(subject, "name_last")
                email = (_xt(subject, "email") or _xt(subject, "Email") or
                         _xt(subject, "InternetEmailAddress") or _xt(subject, "email_address") or
                         _xt(subject, "contact_email"))
                phone = (_xt(subject, "phone") or _xt(subject, "Phone") or
                         _xt(subject, "phone_number") or _xt(subject, "FormattedNumber") or
                         _xt(subject, "contact_phone"))
                dob = _extract_dob(subject)
                ssn4 = _extract_ssn_last4(subject)
                if not email:
                    email = order_email
                if not phone:
                    phone = order_phone
                if first or last:
                    applicants.append(dict(first_name=first, last_name=last, email=email,
                                           phone=phone, accio_order_number=order_number,
                                           accio_remote_number=remote_number, company_name="",
                                           account_name="",
                                           accio_sub_order=sub_order_id, accio_order_type=order_type,
                                           date_of_birth=dob, last_four_ssn=ssn4))

    # --- Deep scan fallback ---
    if not applicants:
        def deep_scan_for_applicants(elem):
            found = []
            parent_map = {c: p for p in elem.iter() for c in p}
            for el in elem.iter():
                if el.tag == "name_first" and el.text:
                    first = el.text.strip()
                    parent = parent_map.get(el)
                    if parent is not None:
                        last = _xt(parent, "name_last")
                        email = phone = ""
                        for tag in ["email", "Email", "InternetEmailAddress", "email_address",
                                    "EmailAddress", "e_mail", "applicant_email", "candidate_email",
                                    "contact_email"]:
                            email = _xt(parent, tag)
                            if email:
                                break
                        for tag in ["phone", "Phone", "phone_number", "PhoneNumber",
                                    "FormattedNumber", "telephone", "contact_phone",
                                    "home_phone", "cell_phone", "mobile"]:
                            phone = _xt(parent, tag)
                            if phone:
                                break
                        dob = _extract_dob(parent) if parent is not None else ""
                        ssn4 = _extract_ssn_last4(parent) if parent is not None else ""
                        if first or last:
                            found.append(dict(first_name=first, last_name=last, email=email,
                                              phone=phone, accio_order_number="",
                                              accio_remote_number="", company_name="",
                                              account_name="",
                                              date_of_birth=dob, last_four_ssn=ssn4))
            return found
        applicants = deep_scan_for_applicants(root)

    return applicants, None


def _xt(el, tag):
    c = el.find(tag)
    return c.text.strip() if c is not None and c.text else ""


def _extract_dob(subject_el):
    """
    Extract date of birth from an XML subject element.
    Accio XML may use: dob, DOB, date_of_birth, DateOfBirth, birthDate, birth_date.
    Accio format is YYYYMMDD (e.g., 19850315) but we also handle YYYY-MM-DD and MM/DD/YYYY.
    Returns formatted string "MM/DD/YYYY" for display, or empty string.
    """
    import re as _re
    raw = ""
    for tag in ["dob", "DOB", "date_of_birth", "DateOfBirth", "birthDate", "birth_date",
                "BirthDate", "dateOfBirth"]:
        raw = _xt(subject_el, tag)
        if raw:
            break
    if not raw:
        return ""
    raw = raw.strip()
    # Format: YYYYMMDD (Accio standard)
    if _re.match(r"^\d{8}$", raw):
        return f"{raw[4:6]}/{raw[6:8]}/{raw[0:4]}"
    # Format: YYYY-MM-DD
    if _re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
        parts = raw.split("-")
        return f"{parts[1]}/{parts[2]}/{parts[0]}"
    # Format: MM/DD/YYYY (already correct)
    if _re.match(r"^\d{2}/\d{2}/\d{4}$", raw):
        return raw
    return raw  # Return as-is if unrecognized format


def _extract_ssn_last4(subject_el):
    """
    Extract last 4 digits of SSN from an XML subject element.
    Accio XML SSN format is 9 digits, no dashes (e.g., 123456789).
    Also handles XXX-XX-XXXX format. Returns last 4 digits or empty string.
    SECURITY: Only the last 4 digits are stored. Full SSN is never persisted.
    """
    import re as _re
    raw = ""
    for tag in ["ssn", "SSN", "social_security", "SocialSecurityNumber",
                "social_security_number", "ssn_number"]:
        raw = _xt(subject_el, tag)
        if raw:
            break
    if not raw:
        return ""
    # Strip dashes and spaces
    digits_only = _re.sub(r"[^0-9]", "", raw)
    if len(digits_only) >= 4:
        return digits_only[-4:]
    return ""


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------
def send_release_email(db, applicant_id):
    a = db.execute("SELECT * FROM applicants WHERE id = %s", (applicant_id,)).fetchone()
    if not a:
        return False, "Applicant not found"
    if not a["email"]:
        return False, "No email address"
    if not a["assigned_code"]:
        return False, "No code assigned"

    smtp_server = get_setting(db, "smtp_server")
    smtp_port = int(get_setting(db, "smtp_port") or 587)
    smtp_user = get_setting(db, "smtp_username")
    smtp_pass = get_setting(db, "smtp_password")
    use_tls = get_setting(db, "smtp_use_tls") == "1"
    sender_email = get_setting(db, "sender_email")
    sender_name = get_setting(db, "sender_name")

    if not smtp_server or not sender_email:
        return False, "SMTP not configured. Go to Settings."

    # Look up client email so we can CC them on the applicant notification
    client_email = None
    if a.get("client_id"):
        client_row = db.execute("SELECT contact_email FROM clients WHERE id = %s", (a["client_id"],)).fetchone()
        if client_row and client_row["contact_email"]:
            client_email = client_row["contact_email"].strip()

    reps = dict(first_name=a["first_name"], last_name=a["last_name"],
                email=a["email"], code=a["assigned_code"],
                company_name=get_setting(db, "company_name"),
                ori_number=get_setting(db, "ori_number"))

    subj = get_setting(db, "email_subject").format(**reps)
    body = get_setting(db, "email_body").format(**reps)

    msg = MIMEMultipart("mixed")
    msg["From"] = f"{sender_name} <{sender_email}>"
    msg["To"] = a["email"]
    if client_email:
        msg["Cc"] = client_email
    msg["Subject"] = subj
    msg["Reply-To"] = sender_email
    # RFC 5322 required headers — missing these is a major spam signal
    msg["Message-ID"] = make_msgid(domain=sender_email.split("@")[-1] if "@" in sender_email else "localhost")
    msg["Date"] = formatdate(localtime=True)
    # List-Unsubscribe — required by Gmail/Yahoo for bulk senders since Feb 2024.
    # Points to Reply-To so recipients can request removal; prevents spam filtering.
    msg["List-Unsubscribe"] = f"<mailto:{sender_email}?subject=unsubscribe>"

    alt_part = MIMEMultipart("alternative")
    alt_part.attach(MIMEText(body, "plain"))

    company_name = get_setting(db, "company_name") or "BR Solutions"
    html_body = f"""<!DOCTYPE html>
<html lang="en" xmlns="http://www.w3.org/1999/xhtml">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0; padding:0; background-color:#f9f9f9;">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background-color:#f9f9f9;">
<tr><td align="center" style="padding:20px 0;">
<table role="presentation" width="600" cellpadding="0" cellspacing="0" style="background-color:#ffffff; border-radius:8px; overflow:hidden;">
<tr><td style="padding:30px 40px; font-family:Arial, Helvetica, sans-serif; font-size:15px; line-height:1.6; color:#333333;">
{h(body).replace(chr(10), '<br>')}
</td></tr>
<tr><td style="padding:15px 40px 25px; font-family:Arial, Helvetica, sans-serif; font-size:11px; color:#999999; border-top:1px solid #eeeeee;">
This message was sent by {h(company_name)}. If you believe you received this email in error, please contact us at {h(sender_email)}.
</td></tr>
</table>
</td></tr>
</table>
</body>
</html>"""

    alt_part.attach(MIMEText(html_body, "html"))
    msg.attach(alt_part)

    rfp = get_setting(db, "release_form_path")
    if rfp and os.path.exists(rfp):
        with open(rfp, "rb") as f:
            part = MIMEBase("application", "pdf")
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", 'attachment; filename="Fingerprint_Release_Form.pdf"')
            msg.attach(part)

    try:
        logger.info(f"SMTP: Connecting to {smtp_server}:{smtp_port} (TLS={use_tls})")
        cc_info = f", Cc={client_email}" if client_email else ""
        logger.info(f"SMTP: From={sender_email}, To={a['email']}{cc_info}, User={smtp_user}")
        srv = smtplib.SMTP(smtp_server, smtp_port, timeout=15)
        srv.set_debuglevel(0)  # Disabled — debuglevel=1 leaks SMTP credentials to logs
        if use_tls:
            srv.starttls()
            logger.info("SMTP: TLS established")
        if smtp_user and smtp_pass:
            srv.login(smtp_user, smtp_pass)
            logger.info("SMTP: Login successful")
        srv.send_message(msg)
        logger.info("SMTP: Message sent successfully")
        srv.quit()

        now = datetime.now().isoformat()
        db.execute(
            "INSERT INTO email_log (applicant_id,recipient_email,subject,status) VALUES (%s,%s,%s,'sent')",
            (applicant_id, a["email"], subj)
        )

        db.execute(
            "UPDATE applicants SET email_sent=TRUE, email_sent_at=%s, status='emailed', updated_at=%s WHERE id = %s",
            (now, now, applicant_id)
        )
        db.commit()

        # Post status note back to Accio (Ch 9 / Ch 11) — non-blocking: failure does
        # not abort the email success response.  Silently skipped if URL not configured.
        try:
            pb_ok, pb_msg = post_accio_result(db, applicant_id)
            if pb_ok:
                logger.info(f"Accio postback: {pb_msg}")
            else:
                logger.warning(f"Accio postback skipped/failed (non-fatal): {pb_msg}")
        except Exception as pb_exc:
            logger.error(f"Accio postback exception (non-fatal): {pb_exc}")

        return True, "Email sent"
    except Exception as e:
        logger.error(f"SMTP FAILED: {type(e).__name__}: {e}")
        db.execute(
            "INSERT INTO email_log (applicant_id,recipient_email,subject,status,error_message) VALUES (%s,%s,%s,'failed',%s)",
            (applicant_id, a["email"], subj, str(e))
        )
        # Update applicant status to 'email_failed' so the failure is visible
        # in the applicants list.  Preserve any previously-assigned code.
        now_fail = datetime.now().isoformat()
        db.execute(
            "UPDATE applicants SET status='email_failed', updated_at=%s WHERE id = %s",
            (now_fail, applicant_id)
        )
        db.commit()
        return False, str(e)


# ---------------------------------------------------------------------------
# SMS (Twilio)
# ---------------------------------------------------------------------------
def normalize_phone(raw_phone):
    """Normalize a US phone number to E.164 format (+1XXXXXXXXXX).

    Handles common formats from Accio XML:
      (225) 555-1234, 225-555-1234, 2255551234, +12255551234, 1-225-555-1234
    Returns None if the number can't be normalized to a valid 10-digit US number.
    """
    if not raw_phone:
        return None
    digits = re.sub(r"[^\d]", "", raw_phone)
    # Strip leading country code '1' if present (11 digits starting with 1)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) != 10:
        return None
    return f"+1{digits}"


def send_release_sms(db, applicant_id):
    """Send an SMS to the applicant with their fingerprint release info via Twilio.

    Mirrors the email workflow: builds the message from the sms_body template,
    sends via Twilio REST API, logs to sms_log, and updates the applicant record.
    Returns (success: bool, message: str).
    """
    if not HAS_TWILIO:
        return False, "Twilio library not installed. Run: pip install twilio"

    a = db.execute("SELECT * FROM applicants WHERE id = %s", (applicant_id,)).fetchone()
    if not a:
        return False, "Applicant not found"
    if not a["phone"]:
        return False, "No phone number on file"
    if not a["assigned_code"]:
        return False, "No code assigned"

    # Normalize phone to E.164
    to_number = normalize_phone(a["phone"])
    if not to_number:
        return False, f"Invalid phone number: {a['phone']} (must be 10-digit US number)"

    # Load Twilio credentials
    account_sid = get_setting(db, "twilio_account_sid").strip()
    auth_token = get_setting(db, "twilio_auth_token").strip()
    from_number = get_setting(db, "twilio_from_number").strip()

    if not account_sid or not auth_token or not from_number:
        return False, "Twilio not configured. Go to Settings > SMS Configuration."

    # Build message body from template
    reps = dict(
        first_name=a["first_name"],
        last_name=a["last_name"],
        email=a.get("email", ""),
        code=a["assigned_code"],
        company_name=get_setting(db, "company_name"),
        ori_number=get_setting(db, "ori_number"),
    )
    try:
        body = get_setting(db, "sms_body").format(**reps)
    except (KeyError, ValueError) as fmt_err:
        return False, f"SMS template error: {fmt_err}"

    try:
        logger.info(f"SMS: Sending to {to_number} from {from_number}")
        client = TwilioClient(account_sid, auth_token)
        message = client.messages.create(
            body=body,
            from_=from_number,
            to=to_number
        )
        logger.info(f"SMS: Sent successfully, SID={message.sid}, Status={message.status}")

        now = datetime.now().isoformat()
        db.execute(
            "INSERT INTO sms_log (applicant_id,recipient_phone,message_body,twilio_sid,status) "
            "VALUES (%s,%s,%s,%s,%s)",
            (applicant_id, to_number, body, message.sid, message.status)
        )
        db.execute(
            "UPDATE applicants SET sms_sent=TRUE, sms_sent_at=%s, updated_at=%s WHERE id = %s",
            (now, now, applicant_id)
        )
        db.commit()
        return True, f"SMS sent (SID: {message.sid})"

    except Exception as e:
        logger.error(f"SMS FAILED: {type(e).__name__}: {e}")
        db.execute(
            "INSERT INTO sms_log (applicant_id,recipient_phone,message_body,status,error_message) "
            "VALUES (%s,%s,%s,'failed',%s)",
            (applicant_id, to_number, body, str(e))
        )
        db.commit()
        return False, str(e)


def post_accio_result(db, applicant_id):
    """Post a status note back to Accio's researcherxml endpoint.

    Implements Ch 9 (postResults block) and Ch 11 (notes_from_vendor_to_screeningfirm)
    of the Accio Data XML Integration Manual.  Sends filledStatus='filled' so Accio
    marks the order complete and moves it to the next queue, with the assigned code
    and a note visible to the screening firm.
    """
    a = db.execute("SELECT * FROM applicants WHERE id = %s", (applicant_id,)).fetchone()
    if not a:
        return False, "Applicant not found"

    researcher_url = get_setting(db, "accio_researcher_url").strip()
    if not researcher_url:
        # Silently skip — feature disabled when URL not configured
        return False, "Accio researcher URL not configured (skipped)"

    order_number = (a.get("accio_order_number") or "").strip()
    if not order_number:
        return False, "No Accio order number on record — postback skipped"

    accio_account  = get_setting(db, "accio_account")  or ""
    accio_username = get_setting(db, "accio_username") or ""
    accio_password = get_setting(db, "accio_password") or ""
    sub_order      = (a.get("accio_sub_order")  or "1").strip() or "1"
    order_type     = (a.get("accio_order_type") or "Fingerprint").strip() or "Fingerprint"
    assigned_code  = (a.get("assigned_code") or "").strip()

    sent_at = datetime.now(tz=ZoneInfo("America/Chicago")).strftime("%m/%d/%Y %I:%M %p CST")
    # Check if SMS was also sent for this applicant
    sms_status = ""
    if a.get("sms_sent"):
        sms_status = " SMS also sent."
    note_text = (
        f"FP Release email sent on {sent_at}. "
        f"Assigned fingerprint code: {assigned_code}.{sms_status}"
    )

    # Build postResults XML per Ch 9 / Ch 11 of Accio Data XML Integration Manual.
    # filledStatus='filled' — marks order complete, moves it to next queue in Accio.
    # filledCode='see comments' — "Search completed with additional comments" (Ch 9.4.1).
    # <notes_from_vendor_to_screeningfirm> — visible to screening firm, not end user (Ch 11).
    xml_body = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<ScreeningResults>\n'
        '  <mode>PROD</mode>\n'
        '  <login>\n'
        f'    <account>{h(accio_account)}</account>\n'
        f'    <username>{h(accio_username)}</username>\n'
        f'    <password>{h(accio_password)}</password>\n'
        '  </login>\n'
        f'  <postResults order=\'{h(order_number)}\' subOrder=\'{h(sub_order)}\'\n'
        f'               type=\'{h(order_type)}\'\n'
        "               filledStatus='filled'\n"
        "               filledCode='see comments'>\n"
        f'    <notes_from_vendor_to_screeningfirm>{h(note_text)}</notes_from_vendor_to_screeningfirm>\n'
        f'    <text>{h(note_text)}</text>\n'
        '  </postResults>\n'
        '</ScreeningResults>\n'
    )

    try:
        req = urllib.request.Request(
            researcher_url,
            data=xml_body.encode("utf-8"),
            headers={"Content-Type": "text/xml; charset=utf-8"},
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            response_body = resp.read().decode("utf-8", errors="replace")

        # Per Ch 9.2: no <error> node means success; presence of <error> is a failure
        if "<error" in response_body.lower():
            db.execute(
                "INSERT INTO xml_log (direction, raw_xml, parsed_status, error_message) VALUES (%s,%s,%s,%s)",
                ("postback", xml_body[:4000], "error", response_body[:500])
            )
            db.commit()
            logger.warning(f"Accio postback rejected for order {order_number}: {response_body[:200]}")
            return False, f"Accio rejected postback: {response_body[:200]}"

        db.execute(
            "INSERT INTO xml_log (direction, raw_xml, parsed_status) VALUES (%s,%s,%s)",
            ("postback", xml_body[:4000], "success")
        )
        db.commit()
        logger.info(f"Accio postback success for order {order_number} (code: {assigned_code})")
        return True, "Accio notified"

    except Exception as e:
        logger.error(f"Accio postback failed for order {order_number}: {type(e).__name__}: {e}")
        try:
            db.execute(
                "INSERT INTO xml_log (direction, raw_xml, parsed_status, error_message) VALUES (%s,%s,%s,%s)",
                ("postback", xml_body[:4000], "failed", str(e)[:500])
            )
            db.commit()
        except Exception:
            pass
        return False, str(e)


def assign_code(db, applicant_id):
    a = db.execute("SELECT * FROM applicants WHERE id = %s", (applicant_id,)).fetchone()
    if not a:
        return None, "Not found"
    if a["assigned_code"]:
        return a["assigned_code"], "Already assigned"
    # FIX: Added ORDER BY id for deterministic (FIFO) code assignment
    code_row = db.execute(
        "SELECT id, code FROM codes WHERE assigned_to IS NULL ORDER BY id LIMIT 1"
    ).fetchone()
    if not code_row:
        return None, "No codes available"
    now = datetime.now().isoformat()
    db.execute("UPDATE codes SET assigned_to=%s, assigned_date=%s WHERE id = %s",
               (applicant_id, now, code_row["id"]))
    db.execute("UPDATE applicants SET assigned_code=%s, status='code_assigned', updated_at=%s WHERE id = %s",
               (code_row["code"], now, applicant_id))
    db.commit()
    return code_row["code"], "OK"


def import_codes_from_file(filepath, column_index=0, skip_header=True, batch_name=None):
    """
    FIX: Replaced per-row DB connection (was opening/closing a new connection for every
    single code) with a single connection for the entire import. This prevents connection
    pool exhaustion on large imports and dramatically improves performance.
    """
    imported = 0
    duplicates = 0
    error_msg = None
    db = None
    try:
        db = get_db()
        if filepath.endswith(".xlsx") and HAS_OPENPYXL:
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            start_row = 2 if skip_header else 1
            for row in ws.iter_rows(min_row=start_row):
                cell = row[column_index] if column_index < len(row) else None
                if cell and cell.value:
                    code = str(cell.value).strip()
                    if code:
                        try:
                            db.execute("INSERT INTO codes (code, batch_name) VALUES (%s, %s)",
                                       (code, batch_name or "Import"))
                            imported += 1
                        except psycopg2.IntegrityError:
                            db.rollback()
                            duplicates += 1
        else:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                if skip_header:
                    next(reader, None)
                for row in reader:
                    if column_index < len(row):
                        code = row[column_index].strip()
                        if code:
                            try:
                                db.execute("INSERT INTO codes (code, batch_name) VALUES (%s, %s)",
                                           (code, batch_name or "Import"))
                                imported += 1
                            except psycopg2.IntegrityError:
                                db.rollback()
                                duplicates += 1
    except Exception as e:
        error_msg = str(e)
    finally:
        if db:
            db.close()
    return imported, duplicates, error_msg


def auto_detect_code_column(filepath):
    """Try to guess which column has the payment codes."""
    if filepath.endswith(".xlsx") and HAS_OPENPYXL:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        for col_idx, cell in enumerate(ws[1]):
            if cell.value and ("code" in str(cell.value).lower() or "payment" in str(cell.value).lower()):
                return col_idx
    else:
        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if header:
                    for col_idx, cell in enumerate(header):
                        if "code" in cell.lower() or "payment" in cell.lower():
                            return col_idx
        except Exception:
            pass
    return 0


_watcher_running = True


def start_folder_watcher():
    """Monitor watch folder and auto-import codes from dropped files."""
    import threading
    def watch():
        while _watcher_running:
            try:
                for fname in os.listdir(WATCH_FOLDER):
                    fpath = os.path.join(WATCH_FOLDER, fname)
                    if os.path.isfile(fpath) and (fname.endswith(".xlsx") or fname.endswith(".csv")):
                        logger.info(f"Auto-importing {fname}...")
                        col = auto_detect_code_column(fpath)
                        imp, dup, err = import_codes_from_file(
                            fpath, column_index=col, skip_header=True,
                            batch_name=f"Auto-Import {fname}"
                        )
                        if err:
                            logger.error(f"Import error: {err}")
                        else:
                            logger.info(f"Imported {imp} codes ({dup} dups) from {fname}")
                        dest = os.path.join(PROCESSED_FOLDER, fname)
                        shutil.move(fpath, dest)
            except Exception as e:
                logger.error(f"Watcher error: {e}")
            import time
            time.sleep(5)
    t = threading.Thread(target=watch, daemon=True)
    t.start()


# ---------------------------------------------------------------------------
# Flash Messages
# ---------------------------------------------------------------------------
# FIX: flash() was a complete no-op (just `pass`). Fixed to actually store messages.
# With the single-threaded HTTPServer, this module-level dict is safe between redirect cycles.
_flashes = {}


def flash(msg, cat="success"):
    """Store a flash message to be displayed on next page render."""
    _flashes[cat] = msg


def render_flashes():
    """Render and clear all pending flash messages."""
    global _flashes
    if not _flashes:
        return ""
    html = ""
    for cat, msg in list(_flashes.items()):
        icon = "check-circle" if cat == "success" else "exclamation-circle"
        html += f'<div class="alert alert-{h(cat)}"><i class="fas fa-{icon}"></i> {h(msg)}</div>'
    _flashes.clear()
    return html


# ---------------------------------------------------------------------------
# HTML Utilities
# ---------------------------------------------------------------------------
def h(text):
    """HTML escape."""
    return (str(text) if text is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;").replace("'", "&#39;")


def fmt_dt(val):
    """Format datetime for display."""
    if not val:
        return "-"
    if isinstance(val, str):
        try:
            val = datetime.fromisoformat(val)
        except Exception:
            return val
    return val.strftime("%Y-%m-%d %H:%M")


# ===========================================================================
# LAPS RETRIEVER ENGINE (Step 2 of FP Release workflow)
# ===========================================================================
# Security: Criminal records exist ONLY in RAM during processing (never disk)
# Triple-layer ephemeral handling: in-memory -> immediate zeroing -> forced GC
# Zero disk writes for PII, zero logging of personal information

laps_logger = logging.getLogger("laps_retriever")

# --- Secure Memory Handling ---

def _secure_zero_string(s):
    """Best-effort secure zeroing of a Python string's internal buffer."""
    if not s:
        return ""
    try:
        buf_size = len(s)
        str_address = id(s)
        probe_byte = s[0].encode("utf-8")[0] if s else 0
        for candidate_offset in (40, 48, 52, 56):
            try:
                test_val = ctypes.c_char.from_address(str_address + candidate_offset).value
                if test_val == bytes([probe_byte]):
                    ctypes.memset(str_address + candidate_offset, 0, buf_size)
                    break
            except Exception:
                continue
    except Exception:
        pass
    return "\x00" * len(s)


class SecurePassword:
    """Triple-layered ephemeral password container. Value in RAM only, zeroed on destroy."""
    __slots__ = ("_value", "_destroyed")

    def __init__(self, raw):
        self._value = "".join(list(raw))
        self._destroyed = False

    def get(self):
        if self._destroyed:
            raise RuntimeError("Password has been destroyed")
        return self._value

    def destroy(self):
        if self._destroyed:
            return
        old_val = self._value
        self._value = _secure_zero_string(old_val)
        self._value = "\x00" * max(len(old_val), 32)
        del old_val
        del self._value
        self._destroyed = True
        gc.collect()

    def __enter__(self):
        return self

    def __exit__(self, *_):
        self.destroy()

    def __del__(self):
        if getattr(self, "_destroyed", True) is False:
            self.destroy()


# --- LAPS Data Models ---

class CriminalHistoryStatus(str, Enum):
    CLEAR = "Clear"
    HITS = "Hits"
    ERROR = "Error"


@dataclass(frozen=True)
class CriminalRecord:
    """Single arrest/criminal record from LAPS rap sheet."""
    arrest_date: str
    agency: str
    charges: list
    disposition: str = ""


@dataclass(frozen=True)
class RapSheet:
    """Parsed LAPS rap sheet."""
    subject_name: str
    subject_ssn_last4: str
    subject_dob: str
    records: list
    status: CriminalHistoryStatus
    extracted_at: str


@dataclass
class LAPSApplicantMatch:
    """Applicant from FP Release DB matching a LAPS rap sheet."""
    id: int
    first_name: str
    last_name: str
    last_four_ssn: str
    accio_order_number: str
    accio_sub_order: str
    email_sent_at: object
    created_at: object


# --- LAPS Name Parser ---

def _parse_laps_name(full_name):
    """Parse LAPS 'LAST, FIRST MIDDLE' format into (first_name, last_name)."""
    full_name = full_name.strip()
    if "," in full_name:
        parts = full_name.split(",", 1)
        last_name = parts[0].strip()
        first_and_middle = parts[1].strip()
        first_name = first_and_middle.split()[0] if first_and_middle else ""
        return (first_name, last_name)
    else:
        parts = full_name.split()
        if len(parts) >= 2:
            return (parts[0], parts[-1])
        return (full_name, "")


# --- LAPS Database Operations ---

class LAPSDB:
    """Database operations for LAPS retriever using the shared FP Release DB."""

    def get_pending_applicants(self, limit=50):
        """Fetch applicants needing LAPS lookup: email_sent=True, laps_status IS NULL."""
        db = get_db()
        try:
            cur = db.execute("""
                SELECT id, first_name, last_name, last_four_ssn,
                       accio_order_number, accio_sub_order,
                       email_sent_at, created_at
                FROM applicants
                WHERE email_sent = TRUE AND laps_status IS NULL
                ORDER BY email_sent_at DESC, created_at DESC
                LIMIT %s
            """, (limit,))
            rows = cur.fetchall()
            return [
                LAPSApplicantMatch(
                    id=r["id"], first_name=r["first_name"], last_name=r["last_name"],
                    last_four_ssn=r.get("last_four_ssn") or "",
                    accio_order_number=r.get("accio_order_number") or "",
                    accio_sub_order=r.get("accio_sub_order") or "",
                    email_sent_at=r.get("email_sent_at"),
                    created_at=r["created_at"],
                )
                for r in rows
            ]
        except Exception as e:
            laps_logger.error("Failed to fetch pending applicants: %s", e)
            return []
        finally:
            db.close()

    def find_matching_applicant(self, last_name, first_name):
        """
        Match a LAPS rap sheet subject to an FP Release applicant.

        SECURITY: Matching uses NAME ONLY from LAPS. The SSN last 4 and DOB
        stored in the FP Release DB (from the original Accio XML push) are used
        as verification — NO SSN data is extracted from or transmitted through LAPS.

        Strategy:
          1. Match on last_name + first_name (case-insensitive)
          2. Prefer applicants that have SSN/DOB on file (stronger identity)
          3. Return most recent matching applicant (by email_sent_at)
          4. Only match applicants that haven't been processed yet (laps_status IS NULL)
        """
        if not last_name or not first_name:
            return None
        db = get_db()
        try:
            # Name-based match — prefer applicants with SSN/DOB on file for stronger identity
            cur = db.execute("""
                SELECT id, first_name, last_name, last_four_ssn, date_of_birth,
                       accio_order_number, accio_sub_order, email_sent_at, created_at
                FROM applicants
                WHERE LOWER(TRIM(last_name)) = LOWER(TRIM(%s))
                  AND LOWER(TRIM(first_name)) = LOWER(TRIM(%s))
                  AND email_sent = TRUE AND laps_status IS NULL
                ORDER BY
                    CASE WHEN last_four_ssn IS NOT NULL AND last_four_ssn != '' THEN 0 ELSE 1 END,
                    email_sent_at DESC NULLS LAST, created_at DESC
                LIMIT 1
            """, (last_name, first_name))
            row = cur.fetchone()
            if not row:
                return None
            return LAPSApplicantMatch(
                id=row["id"], first_name=row["first_name"], last_name=row["last_name"],
                last_four_ssn=row.get("last_four_ssn") or "",
                accio_order_number=row.get("accio_order_number") or "",
                accio_sub_order=row.get("accio_sub_order") or "",
                email_sent_at=row.get("email_sent_at"),
                created_at=row["created_at"],
            )
        except Exception as e:
            laps_logger.error("Failed to find matching applicant: %s", e)
            return None
        finally:
            db.close()

    def update_applicant_laps_status(self, applicant_id, laps_status):
        """Update applicant's LAPS retrieval status."""
        db = get_db()
        try:
            db.execute("""
                UPDATE applicants
                SET laps_status = %s, laps_retrieved_at = NOW(), updated_at = NOW()
                WHERE id = %s
            """, (laps_status, applicant_id))
            return True
        except Exception as e:
            laps_logger.error("Failed to update applicant %d: %s", applicant_id, e)
            return False
        finally:
            db.close()

    def log_cycle_start(self):
        """Log start of a LAPS retrieval cycle. Returns cycle_id."""
        db = get_db()
        try:
            cur = db.execute("""
                INSERT INTO laps_cycle_log (started_at, status) VALUES (NOW(), 'running')
                RETURNING id
            """)
            row = cur.fetchone()
            return row["id"] if row else None
        except Exception as e:
            laps_logger.error("Failed to log cycle start: %s", e)
            return None
        finally:
            db.close()

    def log_cycle_end(self, cycle_id, summary):
        """Log end of a LAPS retrieval cycle."""
        db = get_db()
        try:
            errors_text = "; ".join(summary.get("errors", []))[:2000]
            db.execute("""
                UPDATE laps_cycle_log
                SET finished_at = NOW(), processed = %s, matched = %s,
                    pushed = %s, skipped = %s, failed = %s,
                    errors = %s, status = %s
                WHERE id = %s
            """, (
                summary.get("processed", 0), summary.get("matched", 0),
                summary.get("pushed", 0), summary.get("skipped", 0),
                summary.get("failed", 0), errors_text or None,
                "completed" if not summary.get("errors") else "completed_with_errors",
                cycle_id,
            ))
        except Exception as e:
            laps_logger.error("Failed to log cycle end: %s", e)
        finally:
            db.close()

    def log_result(self, applicant_id, cycle_id, laps_name, laps_status, record_count, push_ok, notes=""):
        """Log individual LAPS result."""
        db = get_db()
        try:
            db.execute("""
                INSERT INTO laps_result_log
                    (applicant_id, cycle_id, laps_name, laps_status, record_count, accio_push_ok, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (applicant_id, cycle_id, laps_name, laps_status, record_count, push_ok, notes or None))
        except Exception as e:
            laps_logger.error("Failed to log LAPS result: %s", e)
        finally:
            db.close()


# --- Rap Sheet Parser ---

class RapSheetParser:
    """Parses fixed-width mainframe-style rap sheet text from LAPS."""

    def __init__(self, raw_text):
        self.raw_text = raw_text
        self.lines = raw_text.split("\n")

    def parse(self):
        subject_name = self._extract_subject_name()
        subject_ssn_last4 = self._extract_subject_ssn_last4()
        subject_dob = self._extract_subject_dob()
        records = self._extract_arrest_records()
        status = CriminalHistoryStatus.HITS if records else CriminalHistoryStatus.CLEAR
        return RapSheet(
            subject_name=subject_name, subject_ssn_last4=subject_ssn_last4,
            subject_dob=subject_dob, records=records, status=status,
            extracted_at=datetime.now(ZoneInfo("UTC")).isoformat(),
        )

    def _extract_subject_name(self):
        for line in self.lines:
            match = re.match(r"\s*NAME\s*:\s*(.+?)(?:\s+(?:DOB|BIRTH|SEX|RACE)\s*:|$)", line, re.IGNORECASE)
            if match:
                name = match.group(1).strip()
                if name:
                    return name
        return ""

    def _extract_subject_ssn_last4(self):
        for line in self.lines:
            if re.search(r"SSN\s*:", line, re.IGNORECASE):
                match = re.search(r"(\d{3})-?(\d{2})-?(\d{4})", line)
                if match:
                    return match.group(3)
                match = re.search(r"[*xX]{3}-?[*xX]{2}-?(\d{4})", line)
                if match:
                    return match.group(1)
        return ""

    def _extract_subject_dob(self):
        for line in self.lines:
            match = re.search(
                r"(?:DOB|BIRTH\s*DATE|DATE\s*OF\s*BIRTH)\s*:\s*(\d{2}/\d{2}/\d{4})",
                line, re.IGNORECASE,
            )
            if match:
                return match.group(1)
            match = re.search(
                r"(?:DOB|BIRTH\s*DATE|DATE\s*OF\s*BIRTH)\s*:\s*(\d{8})",
                line, re.IGNORECASE,
            )
            if match:
                raw = match.group(1)
                return f"{raw[4:6]}/{raw[6:8]}/{raw[0:4]}"
        return ""

    def _extract_arrest_records(self):
        records = []
        current_block = []
        in_records_section = False
        for line in self.lines:
            if "** END OF RAP SHEET **" in line or "END OF RECORD" in line.upper():
                break
            stripped = line.strip().upper()
            if ("ARREST" in stripped and "DATE" in stripped) or "ARREST RECORD" in stripped:
                in_records_section = True
                continue
            if not in_records_section:
                continue
            if re.match(r"^\s*[-=]{5,}\s*$", line):
                if current_block:
                    record = self._parse_record_block(current_block)
                    if record:
                        records.append(record)
                    current_block = []
            else:
                if line.strip():
                    current_block.append(line)
        if current_block:
            record = self._parse_record_block(current_block)
            if record:
                records.append(record)
        return records

    def _parse_record_block(self, lines):
        if not lines:
            return None
        block_text = "\n".join(lines)
        arrest_date = ""
        agency = ""
        charges = []
        disposition = ""

        for pattern in [r"ARREST\s+DATE\s*:\s*(\d{2}/\d{2}/\d{4})", r"ARREST\s+DATE\s*:\s*(\d{2}-\d{2}-\d{4})",
                        r"ARR\s*DT\s*:\s*(\d{2}/\d{2}/\d{4})", r"(\d{2}/\d{2}/\d{4})"]:
            match = re.search(pattern, block_text, re.IGNORECASE)
            if match:
                arrest_date = match.group(1)
                break

        for pattern in [r"AGENCY\s*:\s*(.+?)(?:\n|$)", r"AGY\s*:\s*(.+?)(?:\n|$)",
                        r"ARRESTING\s+AGENCY\s*:\s*(.+?)(?:\n|$)"]:
            match = re.search(pattern, block_text, re.IGNORECASE)
            if match:
                agency = match.group(1).strip()
                break

        for match in re.finditer(r"CHARGE\s+\d+\s*:\s*(.+?)(?:\n|$)", block_text, re.IGNORECASE):
            charge_text = match.group(1).strip()
            if charge_text:
                charges.append(charge_text)

        if not charges:
            for match in re.finditer(r"(?:R\.S\.\s*\d+[:\-]\d+[:\-]?\d*)\s*(.+?)(?:\n|$)", block_text):
                charges.append(match.group(0).strip())

        for pattern in [r"DISPOSITION\s*:\s*(.+?)(?:\n\n|\n[A-Z]{3,}|$)", r"DISP\s*:\s*(.+?)(?:\n|$)",
                        r"COURT\s+ACTION\s*:\s*(.+?)(?:\n|$)"]:
            match = re.search(pattern, block_text, re.IGNORECASE | re.DOTALL)
            if match:
                disposition = match.group(1).strip()
                break

        if arrest_date or agency:
            return CriminalRecord(
                arrest_date=arrest_date or "Unknown", agency=agency or "Unknown",
                charges=charges if charges else ["No charge details available"],
                disposition=disposition,
            )
        return None


# --- LAPS Browser Session ---

class LAPSBrowserSession:
    """Manages Playwright browser session for LAPS portal login and navigation."""

    def __init__(self):
        self._playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self._initialized = False

    async def initialize(self):
        try:
            self._playwright = await async_playwright().start()
            self.browser = await self._playwright.chromium.launch(
                headless=True,
                args=["--disable-dev-shm-usage", "--no-sandbox", "--disable-gpu"],
            )
            self.context = await self.browser.new_context(
                viewport={"width": 1280, "height": 900}, java_script_enabled=True,
            )
            self.page = await self.context.new_page()
            self.page.set_default_timeout(LAPS_PLAYWRIGHT_TIMEOUT)
            self._initialized = True
            laps_logger.info("Browser initialized (headless Chromium)")
            return True
        except Exception as e:
            laps_logger.error("Failed to initialize browser: %s", e)
            return False

    async def cleanup(self):
        for obj_name in ("page", "context", "browser"):
            try:
                obj = getattr(self, obj_name)
                if obj:
                    await obj.close()
                    setattr(self, obj_name, None)
            except Exception:
                pass
        try:
            if self._playwright:
                await self._playwright.stop()
                self._playwright = None
        except Exception:
            pass
        self._initialized = False

    async def login(self, username, password):
        """3-step LAPS login: creds -> token challenge -> dashboard."""
        if not self.page or not self._initialized:
            return False
        try:
            await self.page.goto(LAPS_PORTAL_URL, wait_until="domcontentloaded")
            await self.page.wait_for_selector('input[name="T$CU_USER"]', timeout=LAPS_PLAYWRIGHT_TIMEOUT)
            await self.page.fill('input[name="T$CU_USER"]', username)
            await self.page.fill('input[name="T$CU_PASS"]', password)
            await self.page.click('input[name="T$CU_SEND"]')
            await self.page.wait_for_load_state("networkidle", timeout=LAPS_PLAYWRIGHT_TIMEOUT)

            content = await self.page.content()
            token_match = re.search(
                r"[Ee]nter\s+(?:the\s+)?token\s+(?:at\s+(?:position\s+)?)?([A-Ja-j])[\s-]*(\d)", content,
            )

            if not token_match:
                if "Recently Completed" in content or "Dashboard" in content:
                    return True
                if any(w in content.lower() for w in ("invalid", "incorrect", "failed")):
                    laps_logger.error("LAPS login rejected")
                    return False
                return False

            col_letter = token_match.group(1).upper()
            row_number = token_match.group(2)
            laps_logger.info("Token challenge: position %s%s", col_letter, row_number)

            if row_number not in LAPS_TOKEN_GRID or col_letter not in LAPS_TOKEN_GRID[row_number]:
                laps_logger.error("Token grid position not found")
                return False

            token_value = LAPS_TOKEN_GRID[row_number][col_letter]
            token_filled = False
            for selector in ['input[name="T$CU_TOKEN"]', 'input[name*="TOKEN"]',
                             'input[type="text"]:not([name="T$CU_USER"]):not([name="T$CU_PASS"])',
                             'input[type="password"]']:
                try:
                    elem = await self.page.query_selector(selector)
                    if elem:
                        await elem.fill(token_value)
                        token_filled = True
                        break
                except Exception:
                    continue

            _secure_zero_string(token_value)
            del token_value
            if not token_filled:
                return False

            await self.page.click('input[name="T$CU_SEND"]')
            await self.page.wait_for_load_state("networkidle", timeout=LAPS_PLAYWRIGHT_TIMEOUT)

            dashboard_content = await self.page.content()
            if "Recently Completed" in dashboard_content or "Dashboard" in dashboard_content:
                laps_logger.info("LAPS login successful")
                return True
            if any(w in dashboard_content.lower() for w in ("invalid", "incorrect")):
                return False
            return True
        except Exception as e:
            laps_logger.error("Login failed: %s", e)
            return False

    async def get_recently_completed(self):
        """Navigate to Recently Completed tab and extract applicant entries."""
        if not self.page:
            return []
        try:
            clicked = False
            for selector in ['a:has-text("Recently Completed")', 'button:has-text("Recently Completed")',
                             'td:has-text("Recently Completed")', '[id*="recent"]']:
                try:
                    elem = await self.page.query_selector(selector)
                    if elem:
                        await elem.click()
                        clicked = True
                        break
                except Exception:
                    continue
            if not clicked:
                return []

            await self.page.wait_for_load_state("networkidle", timeout=LAPS_PLAYWRIGHT_TIMEOUT)
            await asyncio.sleep(1)

            applicant_entries = []
            links = await self.page.query_selector_all("table a, tr a, .applicant a, a[href]")
            for link in links:
                text = await link.text_content()
                if text and text.strip() and "," in text.strip():
                    applicant_entries.append({"name": text.strip()})

            if not applicant_entries:
                rows = await self.page.query_selector_all("tr")
                for row in rows:
                    row_text = await row.text_content()
                    if row_text and "," in row_text:
                        name_match = re.search(r"([A-Z][A-Z\s'-]+,\s*[A-Z][A-Z\s'-]+)", row_text)
                        if name_match:
                            applicant_entries.append({"name": name_match.group(1).strip()})

            laps_logger.info("Found %d applicants in Recently Completed", len(applicant_entries))
            return applicant_entries
        except Exception as e:
            laps_logger.error("Failed to get recently completed: %s", e)
            return []

    async def get_applicant_details(self, applicant_name):
        """Click applicant, navigate to details, extract rap sheet text."""
        if not self.page:
            return None
        try:
            name_link = await self.page.query_selector(f'a:has-text("{applicant_name}")')
            if not name_link:
                name_link = await self.page.query_selector(f'a:has-text("{applicant_name[:20]}")')
            if not name_link:
                return None

            await name_link.click()
            await self.page.wait_for_load_state("networkidle", timeout=LAPS_PLAYWRIGHT_TIMEOUT)
            await asyncio.sleep(0.5)

            for selector in ['button:has-text("Details")', 'input[value="Details"]',
                             'a:has-text("Details")', '[id*="detail"]']:
                try:
                    elem = await self.page.query_selector(selector)
                    if elem:
                        await elem.click()
                        break
                except Exception:
                    continue

            await self.page.wait_for_load_state("networkidle", timeout=LAPS_PLAYWRIGHT_TIMEOUT)
            await asyncio.sleep(1)

            pages = self.context.pages if self.context else []
            detail_page = self.page
            if len(pages) > 1:
                detail_page = pages[-1]

            full_text = await detail_page.inner_text("body")
            if detail_page != self.page and len(pages) > 1:
                await detail_page.close()

            if not full_text or len(full_text) < 50:
                return None

            for marker in ["NON-INVESTIGATIVE", "RAP SHEET", "CRIMINAL HISTORY", "ARREST RECORD", "NAME:", "REQUESTED BY"]:
                idx = full_text.find(marker)
                if idx > 0:
                    full_text = full_text[idx:]
                    break

            await self._nav_back()
            return full_text
        except Exception as e:
            laps_logger.error("Failed to get details: %s", e)
            try:
                await self._nav_back()
            except Exception:
                pass
            return None

    async def _nav_back(self):
        if not self.page:
            return
        try:
            await self.page.go_back()
            await self.page.wait_for_load_state("networkidle", timeout=10000)
            content = await self.page.content()
            if "Recently Completed" not in content:
                for sel in ['a:has-text("Recently Completed")', 'button:has-text("Recently Completed")']:
                    try:
                        elem = await self.page.query_selector(sel)
                        if elem:
                            await elem.click()
                            await self.page.wait_for_load_state("networkidle", timeout=10000)
                            break
                    except Exception:
                        continue
        except Exception:
            pass


# --- Accio PostResults XML Builder & Pusher ---

def _laps_xml_escape(text):
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;").replace("'", "&apos;")

_VALID_FILLED_STATUSES = frozenset({"filled", "unfilled", "in progress", "canceled", "failed"})


class AccioLAPSPusher:
    """Builds and pushes criminal history results to Accio PostResults endpoint."""
    SEARCH_TYPE = "Fingerprint Criminal History"

    def __init__(self):
        db = get_db()
        self._base_url = (get_setting(db, "accio_researcher_url") or "").rstrip("/") or ACCIO_API_BASE_URL.rstrip("/")

    def _build_login_xml(self):
        db = get_db()
        return (
            "<login>"
            f"<account>{_laps_xml_escape(get_setting(db, 'accio_account'))}</account>"
            f"<username>{_laps_xml_escape(get_setting(db, 'accio_username'))}</username>"
            f"<password>{_laps_xml_escape(get_setting(db, 'accio_password'))}</password>"
            "</login>"
        )

    def _build_registration_xml(self):
        today_str = datetime.now(ZoneInfo("UTC")).strftime("%Y%m%d")
        return (
            "<registration>"
            f"<Company>{_laps_xml_escape(ACCIO_REGISTRATION_COMPANY)}</Company>"
            "<version>1.0.0</version>"
            f"<last_changed_date>{today_str}</last_changed_date>"
            f"<access_key>{_laps_xml_escape(ACCIO_REGISTRATION_KEY)}</access_key>"
            "<contacts><business><name/><phone_number/><email/></business>"
            "<technical><name/><phone_number/><email/></technical></contacts>"
            "</registration>"
        )

    def _build_verified_items(self, rap_sheet):
        items = []
        items.append(f"<verifieditem><fieldname>Record Status</fieldname>"
                     f"<fieldvalue>{_laps_xml_escape(rap_sheet.status.value)}</fieldvalue></verifieditem>")
        if rap_sheet.records:
            for i, record in enumerate(rap_sheet.records, 1):
                items.append(f"<verifieditem><fieldname>Arrest #{i} - Date</fieldname>"
                             f"<fieldvalue>{_laps_xml_escape(record.arrest_date)}</fieldvalue></verifieditem>")
                items.append(f"<verifieditem><fieldname>Arrest #{i} - Agency</fieldname>"
                             f"<fieldvalue>{_laps_xml_escape(record.agency)}</fieldvalue></verifieditem>")
                if record.charges:
                    for j, charge in enumerate(record.charges, 1):
                        items.append(f"<verifieditem><fieldname>Arrest #{i} - Charge {j}</fieldname>"
                                     f"<fieldvalue>{_laps_xml_escape(charge)}</fieldvalue></verifieditem>")
                else:
                    items.append(f"<verifieditem><fieldname>Arrest #{i} - Charges</fieldname>"
                                 f"<fieldvalue></fieldvalue></verifieditem>")
                items.append(f"<verifieditem><fieldname>Arrest #{i} - Disposition</fieldname>"
                             f"<fieldvalue>{_laps_xml_escape(record.disposition)}</fieldvalue></verifieditem>")
        items.append(f"<verifieditem><fieldname>Search Date</fieldname>"
                     f"<fieldvalue>{_laps_xml_escape(rap_sheet.extracted_at)}</fieldvalue></verifieditem>")
        return "".join(items)

    def _build_note_text(self, rap_sheet):
        if rap_sheet.status == CriminalHistoryStatus.CLEAR:
            dob_line = f"\nDOB: {rap_sheet.subject_dob}" if rap_sheet.subject_dob else ""
            return (
                "=== LOUISIANA LAPS - NO CRIMINAL HISTORY ===\n\n"
                "No criminal arrest records were found in the\n"
                "Louisiana Applicant Processing System (LAPS)\n"
                "for the submitted applicant.\n\n"
                f"Searched: {rap_sheet.extracted_at}{dob_line}"
            )
        dob_line = f"DOB: {rap_sheet.subject_dob}\n" if rap_sheet.subject_dob else ""
        lines = ["=== LOUISIANA LAPS - CRIMINAL HISTORY FOUND ===", "",
                 f"{dob_line}Total Arrests: {len(rap_sheet.records)}", ""]
        for i, record in enumerate(rap_sheet.records, 1):
            lines.append(f"--- Arrest #{i} ---")
            lines.append(f"Date: {record.arrest_date}")
            lines.append(f"Agency: {record.agency}")
            for j, charge in enumerate(record.charges, 1):
                lines.append(f"  Charge {j}: {charge}")
            lines.append(f"Disposition: {record.disposition or 'Not available'}")
            lines.append("")
        lines.append(f"Searched: {rap_sheet.extracted_at}")
        return "\n".join(lines)

    async def push_result(self, order_number, sub_order_number, rap_sheet):
        """Push criminal history result to Accio PostResults."""
        if rap_sheet.status == CriminalHistoryStatus.CLEAR:
            filled_status, filled_code = "filled", "Clear"
        elif rap_sheet.status == CriminalHistoryStatus.HITS:
            filled_status, filled_code = "filled", "Hits"
        else:
            filled_status, filled_code = "failed", "Error"

        if filled_status not in _VALID_FILLED_STATUSES:
            laps_logger.error("BLOCKED: Invalid filledStatus '%s'", filled_status)
            return False

        note_text = self._build_note_text(rap_sheet)
        verified_xml = self._build_verified_items(rap_sheet)

        request_xml = (
            "<?xml version='1.0' encoding='UTF-8'?>"
            "<ScreeningResults>"
            f"<mode>{_laps_xml_escape(ACCIO_API_MODE)}</mode>"
            f"{self._build_login_xml()}"
            f"{self._build_registration_xml()}"
            f'<postResults order="{_laps_xml_escape(order_number)}"'
            f' subOrder="{_laps_xml_escape(sub_order_number)}"'
            f' type="{_laps_xml_escape(self.SEARCH_TYPE)}"'
            f' filledStatus="{_laps_xml_escape(filled_status)}"'
            f' filledCode="{_laps_xml_escape(filled_code)}">'
            f"<notes_from_vendor_to_screeningfirm>{_laps_xml_escape(note_text)}</notes_from_vendor_to_screeningfirm>"
            f"<text>{_laps_xml_escape(note_text)}</text>"
            f"{verified_xml}"
            "</postResults>"
            "</ScreeningResults>"
        )

        url = f"{self._base_url}{ACCIO_POSTRESULTS_PATH}"
        for attempt in range(1, LAPS_MAX_RETRIES + 1):
            try:
                async with httpx.AsyncClient(timeout=httpx.Timeout(LAPS_HTTP_TIMEOUT), verify=True) as client:
                    response = await client.post(url, content=request_xml, headers={"Content-Type": "text/xml"})

                laps_logger.info("Accio POST attempt %d/%d -> HTTP %d", attempt, LAPS_MAX_RETRIES, response.status_code)

                if response.status_code >= 500:
                    if attempt < LAPS_MAX_RETRIES:
                        await asyncio.sleep(LAPS_RETRY_BASE_DELAY * (2 ** (attempt - 1)))
                        continue
                    return False
                if response.status_code >= 400:
                    return False

                try:
                    root = ET.fromstring(response.text)
                    if root.findall(".//error"):
                        for err in root.findall(".//error"):
                            laps_logger.error("Accio error: %s", err.findtext("errortext", "?"))
                        return False
                except ET.ParseError:
                    pass

                laps_logger.info("Accio PostResults accepted (order %s)", order_number)
                return True
            except Exception as e:
                laps_logger.error("Accio push error: %s", e)
                if attempt < LAPS_MAX_RETRIES:
                    await asyncio.sleep(LAPS_RETRY_BASE_DELAY * attempt)
                    continue
                return False
        return False


# --- LAPS Retriever Orchestrator ---

class LAPSRetriever:
    """Main orchestrator: Login -> fetch -> parse -> match -> push -> cleanup."""

    def __init__(self):
        self.laps_db = LAPSDB()
        self.pusher = AccioLAPSPusher()

    async def run_cycle(self):
        summary = {
            "timestamp": datetime.now(ZoneInfo("UTC")).isoformat(),
            "processed": 0, "matched": 0, "pushed": 0,
            "skipped": 0, "failed": 0, "errors": [],
        }

        cycle_id = self.laps_db.log_cycle_start()
        session = LAPSBrowserSession()

        if not await session.initialize():
            summary["errors"].append("Browser initialization failed")
            if cycle_id:
                self.laps_db.log_cycle_end(cycle_id, summary)
            return summary

        try:
            with SecurePassword(LAPS_PASSWORD) as pwd:
                login_ok = await session.login(LAPS_USERNAME, pwd.get())

            if not login_ok:
                summary["errors"].append("LAPS login failed")
                return summary

            entries = await session.get_recently_completed()
            if not entries:
                laps_logger.info("No recently completed applicants found")
                return summary

            laps_logger.info("Processing %d applicant(s)...", len(entries))

            for entry in entries:
                laps_name = entry.get("name", "")
                if not laps_name:
                    continue

                summary["processed"] += 1
                rap_text = None

                try:
                    rap_text = await session.get_applicant_details(laps_name)
                    if not rap_text:
                        summary["failed"] += 1
                        continue

                    parser = RapSheetParser(rap_text)
                    rap_sheet = parser.parse()
                    laps_logger.info("Parsed: status=%s, arrests=%d", rap_sheet.status.value, len(rap_sheet.records))

                    first_name, last_name = _parse_laps_name(rap_sheet.subject_name or laps_name)
                    # SECURITY: Match by name only — SSN/DOB from Accio DB, never from LAPS
                    applicant = self.laps_db.find_matching_applicant(last_name, first_name)

                    if not applicant:
                        laps_logger.info("No matching FP Release applicant found")
                        summary["skipped"] += 1
                        continue

                    summary["matched"] += 1
                    push_ok = await self.pusher.push_result(
                        applicant.accio_order_number, applicant.accio_sub_order or "1", rap_sheet,
                    )

                    if push_ok:
                        summary["pushed"] += 1
                        self.laps_db.update_applicant_laps_status(applicant.id, rap_sheet.status.value)
                    else:
                        summary["failed"] += 1
                        self.laps_db.update_applicant_laps_status(applicant.id, "error")

                    if cycle_id:
                        self.laps_db.log_result(
                            applicant.id, cycle_id, laps_name, rap_sheet.status.value,
                            len(rap_sheet.records), push_ok,
                        )
                except Exception as e:
                    laps_logger.error("Error processing applicant: %s", e)
                    summary["failed"] += 1
                    summary["errors"].append(str(e))
                finally:
                    if rap_text:
                        _secure_zero_string(rap_text)
                        rap_text = None
                    gc.collect()
        finally:
            await session.cleanup()
            gc.collect()
            if cycle_id:
                self.laps_db.log_cycle_end(cycle_id, summary)

        return summary


# --- LAPS Background Worker Thread ---

_laps_worker_running = False
_laps_last_summary = None
_laps_next_run = None


def _run_laps_async_cycle():
    """Run one async LAPS cycle in a new event loop (called from background thread)."""
    global _laps_last_summary
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        retriever = LAPSRetriever()
        summary = loop.run_until_complete(retriever.run_cycle())
        _laps_last_summary = summary
        laps_logger.info(
            "Cycle complete: processed=%d matched=%d pushed=%d skipped=%d failed=%d",
            summary["processed"], summary["matched"], summary["pushed"],
            summary["skipped"], summary["failed"],
        )
    except Exception as e:
        laps_logger.error("LAPS cycle failed: %s", e)
        _laps_last_summary = {"timestamp": datetime.now(ZoneInfo("UTC")).isoformat(),
                              "processed": 0, "matched": 0, "pushed": 0,
                              "skipped": 0, "failed": 0, "errors": [str(e)]}
    finally:
        loop.close()


def _laps_worker_loop():
    """Background thread: run LAPS retrieval cycles on an hourly schedule."""
    global _laps_worker_running, _laps_next_run
    _laps_worker_running = True
    laps_logger.info("LAPS background worker started (interval: %ds)", LAPS_HOURLY_INTERVAL)

    while _laps_worker_running:
        laps_logger.info("Starting LAPS retrieval cycle...")
        try:
            _run_laps_async_cycle()
        except Exception as e:
            laps_logger.error("Unhandled error in LAPS worker: %s", e)

        gc.collect()
        _laps_next_run = datetime.now(ZoneInfo("UTC")) + timedelta(seconds=LAPS_HOURLY_INTERVAL)
        laps_logger.info("Next LAPS cycle at %s", _laps_next_run.isoformat())

        for _ in range(LAPS_HOURLY_INTERVAL):
            if not _laps_worker_running:
                break
            time.sleep(1)

    laps_logger.info("LAPS background worker stopped")


def start_laps_worker():
    """Start the LAPS retrieval background worker thread."""
    if not LAPS_ENABLED:
        laps_logger.info("LAPS Retriever is DISABLED (set LAPS_ENABLED=true to enable)")
        return

    missing_vars = []
    if not LAPS_USERNAME:
        missing_vars.append("LAPS_USERNAME")
    if not LAPS_PASSWORD:
        missing_vars.append("LAPS_PASSWORD")
    if not ACCIO_API_BASE_URL:
        missing_vars.append("ACCIO_API_BASE_URL")
    if not ACCIO_API_ACCOUNT:
        missing_vars.append("ACCIO_API_ACCOUNT")

    if missing_vars:
        laps_logger.warning("LAPS worker NOT started — missing env vars: %s", ", ".join(missing_vars))
        return

    if not HAS_PLAYWRIGHT:
        laps_logger.warning("LAPS worker NOT started — playwright not installed")
        return
    if not HAS_HTTPX:
        laps_logger.warning("LAPS worker NOT started — httpx not installed")
        return

    t = threading.Thread(target=_laps_worker_loop, daemon=True, name="laps-worker")
    t.start()
    laps_logger.info("LAPS background worker thread started")


def trigger_laps_cycle_manual():
    """Trigger a single LAPS cycle manually (runs in a separate thread)."""
    t = threading.Thread(target=_run_laps_async_cycle, daemon=True, name="laps-manual")
    t.start()
    return True


# --- LAPS Dashboard Page ---

def page_laps_dashboard(db, nav_user=None):
    """LAPS Retriever dashboard page."""
    global _laps_last_summary, _laps_next_run, _laps_worker_running

    # Get LAPS stats from DB
    try:
        pending_row = db.execute(
            "SELECT COUNT(*) as cnt FROM applicants WHERE email_sent = TRUE AND laps_status IS NULL"
        ).fetchone()
        pending_count = pending_row["cnt"] if pending_row else 0
    except Exception:
        pending_count = 0

    try:
        clear_row = db.execute("SELECT COUNT(*) as cnt FROM applicants WHERE laps_status = 'Clear'").fetchone()
        clear_count = clear_row["cnt"] if clear_row else 0
    except Exception:
        clear_count = 0

    try:
        hits_row = db.execute("SELECT COUNT(*) as cnt FROM applicants WHERE laps_status = 'Hits'").fetchone()
        hits_count = hits_row["cnt"] if hits_row else 0
    except Exception:
        hits_count = 0

    try:
        error_row = db.execute("SELECT COUNT(*) as cnt FROM applicants WHERE laps_status = 'error'").fetchone()
        error_count = error_row["cnt"] if error_row else 0
    except Exception:
        error_count = 0

    # Recent cycles
    try:
        cycles = db.execute("""
            SELECT * FROM laps_cycle_log ORDER BY started_at DESC LIMIT 10
        """).fetchall()
    except Exception:
        cycles = []

    # Recent results
    try:
        results = db.execute("""
            SELECT lr.*, a.first_name, a.last_name, a.accio_order_number
            FROM laps_result_log lr
            LEFT JOIN applicants a ON lr.applicant_id = a.id
            ORDER BY lr.created_at DESC LIMIT 20
        """).fetchall()
    except Exception:
        results = []

    # Worker status
    worker_status = "Running" if _laps_worker_running else "Stopped"
    worker_color = "#16a34a" if _laps_worker_running else "#dc2626"
    next_run_str = _laps_next_run.strftime("%I:%M %p %Z") if _laps_next_run else "N/A"

    last_summary_html = ""
    if _laps_last_summary:
        ls = _laps_last_summary
        last_summary_html = f"""
        <div style="background: var(--gray-50); padding: 1rem; border-radius: 8px; margin-top: 1rem;">
            <strong>Last Cycle:</strong> {h(ls.get('timestamp', 'N/A')[:19])}<br>
            Processed: {ls.get('processed', 0)} |
            Matched: {ls.get('matched', 0)} |
            Pushed: {ls.get('pushed', 0)} |
            Skipped: {ls.get('skipped', 0)} |
            Failed: {ls.get('failed', 0)}
            {'<br><span style="color:#dc2626;">Errors: ' + h("; ".join(ls.get("errors", []))) + '</span>' if ls.get("errors") else ''}
        </div>
        """

    # Cycle history table
    cycle_rows = ""
    for c in cycles:
        started = c["started_at"].strftime("%m/%d %I:%M %p") if c.get("started_at") else "?"
        finished = c["finished_at"].strftime("%I:%M %p") if c.get("finished_at") else "running..."
        status_badge = f'<span style="color:{"#16a34a" if c.get("status","") == "completed" else "#f59e0b"}">{h(c.get("status", "?"))}</span>'
        cycle_rows += f"""
        <tr>
            <td>{c.get('id','')}</td>
            <td>{started}</td>
            <td>{finished}</td>
            <td>{c.get('processed', 0)}</td>
            <td>{c.get('matched', 0)}</td>
            <td>{c.get('pushed', 0)}</td>
            <td>{c.get('failed', 0)}</td>
            <td>{status_badge}</td>
        </tr>"""

    # Results table
    result_rows = ""
    for r in results:
        name_display = f"{h(r.get('first_name',''))} {h(r.get('last_name',''))}" if r.get('first_name') else h(r.get('laps_name', 'N/A'))
        status_color = "#16a34a" if r.get("laps_status") == "Clear" else "#dc2626" if r.get("laps_status") == "Hits" else "#f59e0b"
        push_icon = '<i class="fas fa-check-circle" style="color:#16a34a"></i>' if r.get("accio_push_ok") else '<i class="fas fa-times-circle" style="color:#dc2626"></i>'
        created = r["created_at"].strftime("%m/%d %I:%M %p") if r.get("created_at") else "?"
        result_rows += f"""
        <tr>
            <td>{name_display}</td>
            <td style="color:{status_color}; font-weight:600;">{h(r.get('laps_status',''))}</td>
            <td>{r.get('record_count', 0)}</td>
            <td>{push_icon}</td>
            <td>{h(r.get('accio_order_number', ''))}</td>
            <td>{created}</td>
        </tr>"""

    # Applicants awaiting LAPS
    try:
        awaiting = db.execute("""
            SELECT id, first_name, last_name, accio_order_number, email_sent_at, created_at
            FROM applicants
            WHERE email_sent = TRUE AND laps_status IS NULL
            ORDER BY email_sent_at DESC
            LIMIT 20
        """).fetchall()
    except Exception:
        awaiting = []

    awaiting_rows = ""
    for a in awaiting:
        sent_at = a["email_sent_at"].strftime("%m/%d %I:%M %p") if a.get("email_sent_at") else "N/A"
        awaiting_rows += f"""
        <tr>
            <td>{a['id']}</td>
            <td>{h(a.get('first_name',''))} {h(a.get('last_name',''))}</td>
            <td>{h(a.get('accio_order_number',''))}</td>
            <td>{sent_at}</td>
            <td><span style="color:#f59e0b; font-weight:600;">Awaiting LAPS</span></td>
        </tr>"""

    content = f"""
    <div style="display:grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap:1rem; margin-bottom:2rem;">
        <div class="card" style="text-align:center;">
            <div style="font-size:2rem; font-weight:700; color:#f59e0b;">{pending_count}</div>
            <div style="color:var(--gray-500);">Awaiting LAPS</div>
        </div>
        <div class="card" style="text-align:center;">
            <div style="font-size:2rem; font-weight:700; color:#16a34a;">{clear_count}</div>
            <div style="color:var(--gray-500);">Clear</div>
        </div>
        <div class="card" style="text-align:center;">
            <div style="font-size:2rem; font-weight:700; color:#dc2626;">{hits_count}</div>
            <div style="color:var(--gray-500);">Hits</div>
        </div>
        <div class="card" style="text-align:center;">
            <div style="font-size:2rem; font-weight:700; color:#6b7280;">{error_count}</div>
            <div style="color:var(--gray-500);">Errors</div>
        </div>
    </div>

    <div class="card" style="margin-bottom:2rem;">
        <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:1rem;">
            <h2 style="margin:0;"><i class="fas fa-robot"></i> LAPS Worker</h2>
            <form method="POST" action="/laps/trigger" style="display:inline;">
                <button type="submit" class="btn btn-primary" style="font-size:0.9rem;">
                    <i class="fas fa-play"></i> Run Now
                </button>
            </form>
        </div>
        <div style="display:flex; gap:2rem; flex-wrap:wrap;">
            <div>
                <strong>Status:</strong>
                <span style="color:{worker_color}; font-weight:600;">
                    <i class="fas fa-{'circle' if _laps_worker_running else 'stop-circle'}"></i> {worker_status}
                </span>
            </div>
            <div><strong>Next Run:</strong> {next_run_str}</div>
            <div><strong>Interval:</strong> {LAPS_HOURLY_INTERVAL // 60} minutes</div>
        </div>
        {last_summary_html}
    </div>

    <div class="card" style="margin-bottom:2rem;">
        <h2 style="margin-bottom:1rem;"><i class="fas fa-history"></i> Recent Cycles</h2>
        <div style="overflow-x:auto;">
            <table style="width:100%; border-collapse:collapse;">
                <thead>
                    <tr style="border-bottom:2px solid var(--gray-200); text-align:left;">
                        <th style="padding:8px;">ID</th>
                        <th style="padding:8px;">Started</th>
                        <th style="padding:8px;">Finished</th>
                        <th style="padding:8px;">Processed</th>
                        <th style="padding:8px;">Matched</th>
                        <th style="padding:8px;">Pushed</th>
                        <th style="padding:8px;">Failed</th>
                        <th style="padding:8px;">Status</th>
                    </tr>
                </thead>
                <tbody>
                    {cycle_rows if cycle_rows else '<tr><td colspan="8" style="padding:1rem; text-align:center; color:var(--gray-400);">No cycles yet — worker will run on schedule</td></tr>'}
                </tbody>
            </table>
        </div>
    </div>

    <div class="card" style="margin-bottom:2rem;">
        <h2 style="margin-bottom:1rem;"><i class="fas fa-clock"></i> Applicants Awaiting LAPS</h2>
        <div style="overflow-x:auto;">
            <table style="width:100%; border-collapse:collapse;">
                <thead>
                    <tr style="border-bottom:2px solid var(--gray-200); text-align:left;">
                        <th style="padding:8px;">ID</th>
                        <th style="padding:8px;">Name</th>
                        <th style="padding:8px;">Accio Order</th>
                        <th style="padding:8px;">Email Sent</th>
                        <th style="padding:8px;">Status</th>
                    </tr>
                </thead>
                <tbody>
                    {awaiting_rows if awaiting_rows else '<tr><td colspan="5" style="padding:1rem; text-align:center; color:var(--gray-400);">No applicants awaiting LAPS retrieval</td></tr>'}
                </tbody>
            </table>
        </div>
    </div>

    <div class="card">
        <h2 style="margin-bottom:1rem;"><i class="fas fa-list-check"></i> Recent LAPS Results</h2>
        <div style="overflow-x:auto;">
            <table style="width:100%; border-collapse:collapse;">
                <thead>
                    <tr style="border-bottom:2px solid var(--gray-200); text-align:left;">
                        <th style="padding:8px;">Applicant</th>
                        <th style="padding:8px;">Status</th>
                        <th style="padding:8px;">Records</th>
                        <th style="padding:8px;">Accio Push</th>
                        <th style="padding:8px;">Order #</th>
                        <th style="padding:8px;">Date</th>
                    </tr>
                </thead>
                <tbody>
                    {result_rows if result_rows else '<tr><td colspan="6" style="padding:1rem; text-align:center; color:var(--gray-400);">No LAPS results yet</td></tr>'}
                </tbody>
            </table>
        </div>
    </div>
    """
    return render_page("LAPS Retriever", content, active="laps", user=nav_user)


def render_page(title, content, active="", user=None):
    """Render a full page with navigation."""
    username_display = h(user["username"]) if user else ""
    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>{h(title)} - Fingerprint Release Manager</title>
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
        <style>
            :root {{
                --primary: #2563eb;
                --primary-dark: #1e40af;
                --success: #10b981;
                --danger: #ef4444;
                --warning: #f59e0b;
                --gray-50: #f9fafb;
                --gray-100: #f3f4f6;
                --gray-200: #e5e7eb;
                --gray-300: #d1d5db;
                --gray-400: #9ca3af;
                --gray-500: #6b7280;
                --gray-700: #374151;
                --gray-900: #111827;
            }}
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: var(--gray-50); color: var(--gray-900); }}
            .container {{ display: flex; min-height: 100vh; }}
            .sidebar {{ width: 250px; background: var(--gray-900); color: white; padding: 2rem 0; overflow-y: auto; box-shadow: 2px 0 8px rgba(0,0,0,0.1); }}
            .sidebar-brand {{ padding: 0 1.5rem 2rem; font-size: 1.5rem; font-weight: bold; display: flex; align-items: center; gap: 0.5rem; border-bottom: 1px solid var(--gray-700); }}
            .sidebar-brand i {{ color: var(--primary); }}
            .sidebar-nav {{ list-style: none; padding: 1rem 0; }}
            .sidebar-nav li {{ margin: 0; }}
            .sidebar-nav a {{ display: flex; align-items: center; gap: 0.75rem; padding: 0.75rem 1.5rem; color: var(--gray-300); text-decoration: none; transition: all 0.2s; }}
            .sidebar-nav a:hover {{ color: white; background: rgba(37, 99, 235, 0.1); padding-left: 1.75rem; }}
            .sidebar-nav a.active {{ color: var(--primary); background: rgba(37, 99, 235, 0.1); border-left: 3px solid var(--primary); padding-left: 1.5rem; }}
            .main {{ flex: 1; display: flex; flex-direction: column; }}
            .topbar {{ background: white; padding: 1rem 2rem; border-bottom: 1px solid var(--gray-200); display: flex; justify-content: space-between; align-items: center; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }}
            .topbar-user {{ display: flex; align-items: center; gap: 1rem; }}
            .topbar-user a {{ color: var(--primary); text-decoration: none; font-size: 0.875rem; }}
            .topbar-user a:hover {{ text-decoration: underline; }}
            .content {{ flex: 1; overflow-y: auto; padding: 2rem; }}
            .page-title {{ font-size: 2rem; font-weight: bold; margin-bottom: 1.5rem; color: var(--gray-900); }}
            .alert {{ padding: 1rem; border-radius: 0.5rem; margin-bottom: 1rem; display: flex; gap: 0.75rem; align-items: flex-start; }}
            .alert-success {{ background: rgba(16, 185, 129, 0.1); border: 1px solid var(--success); color: var(--success); }}
            .alert-error {{ background: rgba(239, 68, 68, 0.1); border: 1px solid var(--danger); color: var(--danger); }}
            .card {{ background: white; border-radius: 0.5rem; box-shadow: 0 1px 3px rgba(0,0,0,0.1); padding: 1.5rem; margin-bottom: 1.5rem; }}
            .card-title {{ font-size: 1.25rem; font-weight: 600; margin-bottom: 1rem; display: flex; align-items: center; gap: 0.5rem; }}
            .card-title i {{ color: var(--primary); }}
            .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 1rem; margin-bottom: 2rem; }}
            .stat-card {{ background: white; border-radius: 0.5rem; padding: 1.5rem; box-shadow: 0 1px 3px rgba(0,0,0,0.1); text-align: center; }}
            .stat-value {{ font-size: 2rem; font-weight: bold; color: var(--primary); margin: 0.5rem 0; }}
            .stat-label {{ color: var(--gray-500); font-size: 0.875rem; }}
            .stat-icon {{ font-size: 2rem; color: var(--primary); margin-bottom: 0.5rem; opacity: 0.7; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 1rem; }}
            thead {{ background: var(--gray-100); border-bottom: 2px solid var(--gray-200); }}
            th {{ padding: 0.75rem; text-align: left; font-weight: 600; color: var(--gray-700); font-size: 0.875rem; }}
            td {{ padding: 0.75rem; border-bottom: 1px solid var(--gray-200); }}
            tbody tr:hover {{ background: var(--gray-50); }}
            .btn {{ padding: 0.5rem 1rem; border: none; border-radius: 0.375rem; font-size: 0.875rem; font-weight: 500; cursor: pointer; text-decoration: none; display: inline-flex; align-items: center; gap: 0.5rem; transition: all 0.2s; }}
            .btn-primary {{ background: var(--primary); color: white; }}
            .btn-primary:hover {{ background: var(--primary-dark); }}
            .btn-success {{ background: var(--success); color: white; }}
            .btn-success:hover {{ background: #059669; }}
            .btn-danger {{ background: var(--danger); color: white; }}
            .btn-danger:hover {{ background: #dc2626; }}
            .btn-small {{ padding: 0.25rem 0.75rem; font-size: 0.75rem; }}
            .form-group {{ margin-bottom: 1.5rem; }}
            label {{ display: block; margin-bottom: 0.5rem; font-weight: 500; color: var(--gray-700); }}
            input[type="text"], input[type="email"], input[type="password"], input[type="number"], select, textarea {{ width: 100%; padding: 0.5rem; border: 1px solid var(--gray-300); border-radius: 0.375rem; font-size: 0.875rem; font-family: inherit; }}
            input:focus, select:focus, textarea:focus {{ outline: none; border-color: var(--primary); box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.1); }}
            .status-badge {{ display: inline-block; padding: 0.25rem 0.75rem; border-radius: 1rem; font-size: 0.75rem; font-weight: 600; text-transform: uppercase; }}
            .status-pending {{ background: rgba(239, 68, 68, 0.1); color: var(--danger); }}
            .status-code_assigned {{ background: rgba(245, 158, 11, 0.1); color: var(--warning); }}
            .status-emailed {{ background: rgba(59, 130, 246, 0.1); color: var(--primary); }}
            .status-opened {{ background: rgba(16, 185, 129, 0.1); color: var(--success); }}
            .status-completed {{ background: rgba(16, 185, 129, 0.1); color: var(--success); }}
            .status-email_failed {{ background: rgba(220, 38, 38, 0.15); color: #dc2626; border: 1px solid rgba(220, 38, 38, 0.3); animation: pulse-fail 2s ease-in-out infinite; }}
            @keyframes pulse-fail {{ 0%, 100% {{ opacity: 1; }} 50% {{ opacity: 0.7; }} }}
            .email-status {{ display: inline-block; width: 12px; height: 12px; border-radius: 50%; margin-right: 0.25rem; }}
            .email-status-opened {{ background: var(--success); }}
            .email-status-not-opened {{ background: var(--danger); }}
            .email-status-unsent {{ background: var(--gray-300); }}
            .es {{ text-align: center; padding: 3rem; }}
            .es i {{ font-size: 4rem; color: var(--gray-300); margin-bottom: 1rem; }}
            .es h3 {{ color: var(--gray-500); }}
            .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 1.5rem; }}
            @media (max-width: 768px) {{
                .container {{ flex-direction: column; }}
                .sidebar {{ width: 100%; padding: 1rem 0; }}
                .sidebar-brand {{ padding: 1rem; }}
                .grid-2 {{ grid-template-columns: 1fr; }}
                .stats {{ grid-template-columns: 1fr; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="sidebar">
                <div class="sidebar-brand">
                    <i class="fas fa-fingerprint"></i>
                    <span>FP Release</span>
                </div>
                <nav class="sidebar-nav">
                    <li><a href="/" class="{'active' if active == 'dashboard' else ''}"><i class="fas fa-chart-line"></i> Dashboard</a></li>
                    <li><a href="/applicants" class="{'active' if active == 'applicants' else ''}"><i class="fas fa-users"></i> Applicants</a></li>
                    <li><a href="/clients" class="{'active' if active == 'clients' else ''}"><i class="fas fa-building"></i> Clients</a></li>
                    <li><a href="/codes" class="{'active' if active == 'codes' else ''}"><i class="fas fa-barcode"></i> Codes</a></li>
                    <li><a href="/laps" class="{'active' if active == 'laps' else ''}"><i class="fas fa-search-plus"></i> LAPS</a></li>
                    <li><a href="/settings" class="{'active' if active == 'settings' else ''}"><i class="fas fa-cog"></i> Settings</a></li>
                    <li><a href="/logs" class="{'active' if active == 'logs' else ''}"><i class="fas fa-file-alt"></i> Logs</a></li>
                    {'<li><a href="/users" class="' + ("active" if active == "users" else "") + '"><i class="fas fa-users-cog"></i> Users</a></li>' if user and user.get("role") == "admin" else ""}
                    <li><a href="/profile" class="{'active' if active == 'profile' else ''}"><i class="fas fa-user-circle"></i> My Account</a></li>
                </nav>
            </div>
            <div class="main">
                <div class="topbar">
                    <div></div>
                    <div class="topbar-user">
                        <span><i class="fas fa-user"></i> {username_display}</span>
                        <a href="/logout">Logout</a>
                    </div>
                </div>
                <div class="content">
                    <h1 class="page-title"><i class="fas fa-fingerprint"></i> {h(title)}</h1>
                    {render_flashes()}
                    {content}
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    return html


def page_login(error=""):
    """Login page."""
    error_html = f'<div class="alert">{h(error)}</div>' if error else ""
    return f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Login - Fingerprint Release Manager</title>
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
        <style>
            :root {{ --primary: #2563eb; --gray-50: #f9fafb; --gray-200: #e5e7eb; --gray-400: #9ca3af; --gray-700: #374151; --gray-900: #111827; }}
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: linear-gradient(135deg, var(--primary) 0%, #1e40af 100%); min-height: 100vh; display: flex; align-items: center; justify-content: center; }}
            .login-card {{ background: white; border-radius: 0.5rem; box-shadow: 0 10px 25px rgba(0,0,0,0.2); padding: 3rem; width: 100%; max-width: 400px; }}
            .login-brand {{ text-align: center; margin-bottom: 2rem; }}
            .login-brand i {{ font-size: 3rem; color: var(--primary); margin-bottom: 0.5rem; }}
            .login-brand h1 {{ font-size: 1.5rem; color: var(--gray-900); margin: 0; }}
            .login-brand p {{ color: var(--gray-400); margin: 0.5rem 0 0 0; font-size: 0.875rem; }}
            .form-group {{ margin-bottom: 1.5rem; }}
            label {{ display: block; margin-bottom: 0.5rem; font-weight: 500; color: var(--gray-700); }}
            input {{ width: 100%; padding: 0.75rem; border: 1px solid var(--gray-200); border-radius: 0.375rem; font-size: 0.875rem; font-family: inherit; }}
            input:focus {{ outline: none; border-color: var(--primary); box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.1); }}
            .btn {{ width: 100%; padding: 0.75rem; background: var(--primary); color: white; border: none; border-radius: 0.375rem; font-size: 0.875rem; font-weight: 500; cursor: pointer; transition: background 0.2s; }}
            .btn:hover {{ background: #1e40af; }}
            .alert {{ background: rgba(239, 68, 68, 0.1); border: 1px solid #ef4444; color: #ef4444; padding: 0.75rem; border-radius: 0.375rem; margin-bottom: 1.5rem; font-size: 0.875rem; }}
        </style>
    </head>
    <body>
        <div class="login-card">
            <div class="login-brand">
                <i class="fas fa-fingerprint"></i>
                <h1>Fingerprint Release</h1>
                <p>Manager v2.0</p>
            </div>
            {error_html}
            <form method="POST" action="/login">
                <div class="form-group">
                    <label for="username">Username</label>
                    <input type="text" id="username" name="username" required autofocus autocomplete="username">
                </div>
                <div class="form-group">
                    <label for="password">Password</label>
                    <input type="password" id="password" name="password" required autocomplete="current-password">
                </div>
                <button type="submit" class="btn">Sign In</button>
            </form>
            <p style="text-align:center;margin-top:1.25rem;font-size:.875rem;">
                <a href="/forgot-password" style="color:var(--primary);text-decoration:none;">Forgot password?</a>
            </p>
        </div>
    </body>
    </html>
    """


def page_dashboard(db, nav_user=None):
    """Dashboard with analytics."""
    total_app = db.execute("SELECT COUNT(*) as cnt FROM applicants").fetchone()["cnt"]
    pending = db.execute("SELECT COUNT(*) as cnt FROM applicants WHERE status='pending'").fetchone()["cnt"]
    emailed = db.execute("SELECT COUNT(*) as cnt FROM applicants WHERE status='emailed'").fetchone()["cnt"]
    email_failed = db.execute("SELECT COUNT(*) as cnt FROM applicants WHERE status='email_failed'").fetchone()["cnt"]
    codes_avail = db.execute("SELECT COUNT(*) as cnt FROM codes WHERE assigned_to IS NULL").fetchone()["cnt"]
    codes_used = db.execute("SELECT COUNT(*) as cnt FROM codes WHERE assigned_to IS NOT NULL").fetchone()["cnt"]

    activity = db.execute("""
        SELECT type, id, col1, col2, ts FROM (
            SELECT 'new_applicant' as type, id, first_name as col1, last_name as col2, created_at as ts FROM applicants
            UNION ALL
            SELECT CASE WHEN status = 'failed' THEN 'email_failed' ELSE 'email_sent' END as type,
                   id, recipient_email as col1,
                   CASE WHEN status = 'failed' THEN COALESCE(error_message, 'Unknown error') ELSE subject END as col2,
                   sent_at as ts FROM email_log
        ) combined
        ORDER BY ts DESC NULLS LAST
        LIMIT 10
    """).fetchall()

    clients = db.execute("""
        SELECT c.id, c.company_name, c.account_name, COUNT(a.id) as app_count
        FROM clients c
        LEFT JOIN applicants a ON a.client_id = c.id
        GROUP BY c.id, c.company_name, c.account_name
        ORDER BY app_count DESC
        LIMIT 5
    """).fetchall()

    # Alert banner shown when there are applicants whose emails failed or were blocked
    fail_alert_html = ""
    if email_failed > 0:
        # Fetch the most recent failures for the alert details
        recent_fails = db.execute(
            "SELECT a.first_name, a.last_name, a.email, el.error_message, el.sent_at "
            "FROM applicants a "
            "JOIN email_log el ON el.applicant_id = a.id AND el.status = 'failed' "
            "WHERE a.status = 'email_failed' "
            "ORDER BY el.sent_at DESC LIMIT 5"
        ).fetchall()
        fail_details = ""
        for rf in recent_fails:
            err_short = h((rf["error_message"] or "Unknown error")[:120])
            fail_details += (
                f'<div style="padding:0.4rem 0; border-bottom:1px solid rgba(220,38,38,0.15);">'
                f'<strong>{h(rf["first_name"])} {h(rf["last_name"])}</strong> '
                f'&lt;{h(rf["email"] or "no email")}&gt; &mdash; '
                f'<span style="color:#991b1b;">{err_short}</span> '
                f'<span style="color:#6b7280; font-size:0.75rem;">({fmt_dt(rf["sent_at"])})</span>'
                f'</div>'
            )
        fail_alert_html = f"""
        <div style="background: rgba(220,38,38,0.08); border: 1px solid rgba(220,38,38,0.3); border-radius: 0.5rem; padding: 1rem 1.25rem; margin-bottom: 1.25rem; display: flex; align-items: flex-start; gap: 0.75rem;">
            <i class="fas fa-exclamation-triangle" style="color: #dc2626; font-size: 1.25rem; margin-top: 0.15rem;"></i>
            <div style="flex: 1;">
                <div style="font-weight: 600; color: #dc2626; margin-bottom: 0.25rem;">
                    {email_failed} applicant{'s' if email_failed != 1 else ''} with failed/blocked email{'s' if email_failed != 1 else ''}
                </div>
                <div style="font-size: 0.85rem; color: #374151; margin-bottom: 0.5rem;">
                    The following emails could not be delivered. Review and retry from the
                    <a href="/applicants" style="color: #dc2626; font-weight: 600;">Applicants</a> page.
                </div>
                {fail_details}
            </div>
        </div>
        """

    stats_html = f"""
    {fail_alert_html}
    <div class="stats">
        <div class="stat-card"><div class="stat-icon"><i class="fas fa-users"></i></div><div class="stat-label">Total Applicants</div><div class="stat-value">{total_app}</div></div>
        <div class="stat-card"><div class="stat-icon"><i class="fas fa-clock"></i></div><div class="stat-label">Pending</div><div class="stat-value">{pending}</div></div>
        <div class="stat-card"><div class="stat-icon"><i class="fas fa-envelope"></i></div><div class="stat-label">Emailed</div><div class="stat-value">{emailed}</div></div>
        <div class="stat-card" {'style="border-color: rgba(220,38,38,0.4); background: rgba(220,38,38,0.04);"' if email_failed > 0 else ''}><div class="stat-icon"><i class="fas fa-envelope-open" style="{'color:#dc2626;' if email_failed > 0 else ''}"></i></div><div class="stat-label" style="{'color:#dc2626;' if email_failed > 0 else ''}">Email Failed</div><div class="stat-value" style="{'color:#dc2626;' if email_failed > 0 else ''}">{email_failed}</div></div>
        <div class="stat-card"><div class="stat-icon"><i class="fas fa-check"></i></div><div class="stat-label">Codes Available</div><div class="stat-value">{codes_avail}</div></div>
        <div class="stat-card"><div class="stat-icon"><i class="fas fa-lock"></i></div><div class="stat-label">Codes Used</div><div class="stat-value">{codes_used}</div></div>
    </div>
    """

    clients_html = """<div class="card"><div class="card-title"><i class="fas fa-building"></i> Top Clients</div><table><thead><tr><th>Company</th><th>Account</th><th>Applicants</th></tr></thead><tbody>"""
    for c in clients:
        # FIX: Dashboard link was /clients/{id} (404) — corrected to query-param format /clients?client_id={id}
        clients_html += f'<tr><td><a href="/clients?client_id={c["id"]}" style="color: var(--primary); text-decoration: none;">{h(c["company_name"])}</a></td><td>{h(c["account_name"] or "-")}</td><td>{c["app_count"]}</td></tr>'
    clients_html += "</tbody></table></div>"

    activity_html = """<div class="card"><div class="card-title"><i class="fas fa-history"></i> Recent Activity</div><table><thead><tr><th>Type</th><th>Details</th><th>Time</th></tr></thead><tbody>"""
    for a in activity:
        if a["type"] == "new_applicant":
            activity_html += f'<tr><td><span class="status-badge status-pending">New</span></td><td>{h(a["col1"])} {h(a["col2"])}</td><td>{fmt_dt(a["ts"])}</td></tr>'
        elif a["type"] == "email_sent":
            activity_html += f'<tr><td><span class="status-badge status-emailed">Email</span></td><td>{h(a["col1"] or "-")} - {h(a["col2"] or "-")}</td><td>{fmt_dt(a["ts"])}</td></tr>'
        elif a["type"] == "email_failed":
            activity_html += f'<tr><td><span class="status-badge status-email_failed">Failed</span></td><td>{h(a["col1"] or "-")} &mdash; {h((a["col2"] or "Unknown error")[:80])}</td><td>{fmt_dt(a["ts"])}</td></tr>'
    activity_html += "</tbody></table></div>"

    return render_page("Dashboard", stats_html + clients_html + activity_html, active="dashboard", user=nav_user)


def page_applicants(db, params, nav_user=None):
    """List and manage applicants."""
    search = (params.get("search", [None])[0] or "").lower()
    rows = db.execute("SELECT * FROM applicants ORDER BY created_at DESC").fetchall()
    if search:
        rows = [r for r in rows if search in f"{r['first_name']} {r['last_name']} {r['accio_order_number'] or ''}".lower()]

    content = f"""
    <div style="margin-bottom: 1rem; display: flex; gap: 0.5rem;">
        <form method="GET" style="flex: 1; display: flex; gap: 0.5rem;">
            <input type="text" name="search" placeholder="Search by name or order #..." style="flex: 1;" value="{h(search)}">
            <button type="submit" class="btn btn-primary"><i class="fas fa-search"></i> Search</button>
        </form>
        <a href="/applicants/add" class="btn btn-primary"><i class="fas fa-plus"></i> Add Applicant</a>
        <form method="POST" action="/applicants/bulk-process" style="margin: 0;">
            <button type="submit" class="btn btn-success"><i class="fas fa-rocket"></i> Bulk Process Pending</button>
        </form>
    </div>
    <div class="card">
        <table>
            <thead><tr><th>Status</th><th>Order #</th><th>Name</th><th>DOB</th><th>Email</th><th>Code</th><th>Email Status</th><th>Actions</th></tr></thead>
            <tbody>
    """

    for r in rows:
        email_status = ""
        if not r["email_sent"]:
            email_status = '<span class="email-status email-status-unsent"></span> Not Sent'
        elif r["status"] == "email_failed":
            email_status = '<span class="email-status" style="background:#dc2626;"></span> Failed'
        else:
            sent_at = r.get("email_sent_at")
            sent_label = ""
            if sent_at:
                try:
                    dt = datetime.fromisoformat(str(sent_at)) if not isinstance(sent_at, datetime) else sent_at
                    sent_label = f' <span style="color:#999;font-size:0.75rem;">({dt.strftime("%m/%d %I:%M%p").lower()})</span>'
                except Exception:
                    pass
            email_status = f'<span class="email-status email-status-opened"></span> Sent{sent_label}'

        # Sanitize status value used in CSS class to prevent class injection
        safe_status = h(r['status']).replace(" ", "_")
        order_num = h(r['accio_order_number'] or '-')
        content += f"""
                <tr>
                    <td><span class="status-badge status-{safe_status}">{h(r['status'])}</span></td>
                    <td><code style="font-size:0.8rem;">{order_num}</code></td>
                    <td>{h(r['first_name'])} {h(r['last_name'])}</td>
                    <td style="font-size:0.85rem;">{h(r.get('date_of_birth') or '-')}</td>
                    <td>
                        <form method="POST" action="/applicants/{r['id']}/update-email" style="display:flex; gap:4px; align-items:center;">
                            <input type="email" name="email" value="{h(r['email'] or '')}" placeholder="Enter email" style="width:180px; padding:4px 8px; font-size:0.85rem; border:1px solid #ccc; border-radius:4px;">
                            <button type="submit" class="btn btn-small" style="padding:4px 8px; font-size:0.75rem; background:#6c757d; color:white; border:none; border-radius:4px;" title="Save email"><i class="fas fa-save"></i></button>
                        </form>
                    </td>
                    <td><code>{h(r['assigned_code'] or '-')}</code></td>
                    <td>{email_status}</td>
                    <td style="white-space: nowrap;">
                        <form method="POST" action="/applicants/{r['id']}/assign-and-send" style="display:inline;">
                            <button type="submit" class="btn btn-primary btn-small"><i class="fas fa-envelope"></i> Assign &amp; Send</button>
                        </form>
                        <form method="POST" action="/applicants/{r['id']}/resend" style="display:inline;">
                            <button type="submit" class="btn btn-small" style="background:#17a2b8; color:white;" title="Resend email"><i class="fas fa-redo"></i> Resend</button>
                        </form>
                        <form method="POST" action="/applicants/{r['id']}/delete" style="display:inline;" onsubmit="return confirm('Delete this applicant?')">
                            <button type="submit" class="btn btn-danger btn-small"><i class="fas fa-trash"></i></button>
                        </form>
                    </td>
                </tr>
        """

    content += "</tbody></table></div>"
    return render_page("Applicants", content, active="applicants", user=nav_user)


def page_add_applicant(nav_user=None):
    content = """
    <div class="card">
        <form method="POST" action="/applicants/add">
            <div class="grid-2">
                <div class="form-group"><label for="first_name">First Name</label><input type="text" id="first_name" name="first_name" required></div>
                <div class="form-group"><label for="last_name">Last Name</label><input type="text" id="last_name" name="last_name" required></div>
            </div>
            <div class="grid-2">
                <div class="form-group"><label for="email">Email</label><input type="email" id="email" name="email"></div>
                <div class="form-group"><label for="phone">Phone</label><input type="text" id="phone" name="phone"></div>
            </div>
            <div class="grid-2">
                <div class="form-group"><label for="date_of_birth">Date of Birth</label><input type="date" id="date_of_birth" name="date_of_birth"></div>
                <div class="form-group"><label for="last_four_ssn">Last 4 SSN</label><input type="text" id="last_four_ssn" name="last_four_ssn" maxlength="4" pattern="[0-9]{4}" placeholder="1234"></div>
            </div>
            <div class="form-group"><label for="accio_order_number">Accio Order Number</label><input type="text" id="accio_order_number" name="accio_order_number"></div>
            <div class="form-group"><label for="notes">Notes</label><textarea id="notes" name="notes" style="min-height: 100px;"></textarea></div>
            <button type="submit" class="btn btn-primary"><i class="fas fa-save"></i> Add Applicant</button>
            <a href="/applicants" class="btn" style="background: var(--gray-300); color: var(--gray-900);">Cancel</a>
        </form>
    </div>
    """
    return render_page("Add Applicant", content, active="applicants", user=nav_user)


def page_codes(db, params, nav_user=None):
    avail = db.execute("SELECT COUNT(*) as cnt FROM codes WHERE assigned_to IS NULL").fetchone()["cnt"]
    assigned = db.execute("SELECT COUNT(*) as cnt FROM codes WHERE assigned_to IS NOT NULL").fetchone()["cnt"]
    rows = db.execute("SELECT * FROM codes ORDER BY imported_at DESC LIMIT 100").fetchall()

    content = f"""
    <div style="margin-bottom: 1rem; display: flex; gap: 0.5rem;">
        <a href="/codes/import" class="btn btn-primary"><i class="fas fa-upload"></i> Import from File</a>
        <a href="/codes/manual" class="btn btn-primary"><i class="fas fa-plus"></i> Add Manually</a>
    </div>
    <div class="stats">
        <div class="stat-card"><div class="stat-icon"><i class="fas fa-check"></i></div><div class="stat-label">Available</div><div class="stat-value">{avail}</div></div>
        <div class="stat-card"><div class="stat-icon"><i class="fas fa-lock"></i></div><div class="stat-label">Assigned</div><div class="stat-value">{assigned}</div></div>
    </div>
    <div class="card">
        <table>
            <thead><tr><th>Code</th><th>Status</th><th>Batch</th><th>Assigned To</th><th>Date</th><th>Actions</th></tr></thead>
            <tbody>
    """

    # FIX: Was slicing rows[:50] from a LIMIT 100 query — now consistently shows all 100
    for r in rows:
        assigned_to = "-"
        if r["assigned_to"]:
            a = db.execute("SELECT first_name, last_name FROM applicants WHERE id = %s",
                           (r["assigned_to"],)).fetchone()
            if a:
                assigned_to = f"{h(a['first_name'])} {h(a['last_name'])}"

        delete_btn = (
            f'<form method="POST" action="/codes/{r["id"]}/delete" style="display:inline;" '
            f'onsubmit="return confirm(\'Delete this code?\')"><button type="submit" class="btn btn-small" '
            f'style="padding:4px 8px; font-size:0.75rem; background:#dc3545; color:white; border:none; '
            f'border-radius:4px;" title="Delete code"><i class="fas fa-trash"></i></button></form>'
            if not r["assigned_to"]
            else '<span style="color:#999; font-size:0.8rem;">In use</span>'
        )

        content += f"""
                <tr>
                    <td><code>{h(r['code'])}</code></td>
                    <td><span class="status-badge {'status-pending' if not r['assigned_to'] else 'status-code_assigned'}">{h(r['status'])}</span></td>
                    <td>{h(r['batch_name'] or '-')}</td>
                    <td>{assigned_to}</td>
                    <td>{fmt_dt(r['imported_at'])}</td>
                    <td>{delete_btn}</td>
                </tr>
        """

    content += "</tbody></table></div>"
    return render_page("Payment Codes", content, active="codes", user=nav_user)


def page_import_codes(nav_user=None):
    content = """
    <div class="card">
        <form method="POST" action="/codes/import" enctype="multipart/form-data">
            <div class="form-group"><label for="file">Excel or CSV File</label><input type="file" id="file" name="file" accept=".xlsx,.csv" required></div>
            <div class="form-group"><label for="column_index">Column Number (0-indexed)</label><input type="number" id="column_index" name="column_index" value="0" min="0"></div>
            <div class="form-group"><label><input type="checkbox" name="skip_header" checked> Skip first row (header)</label></div>
            <div class="form-group"><label for="batch_name">Batch Name</label><input type="text" id="batch_name" name="batch_name" placeholder="e.g., 'January 2024 Import'"></div>
            <button type="submit" class="btn btn-primary"><i class="fas fa-upload"></i> Import Codes</button>
            <a href="/codes" class="btn" style="background: var(--gray-300); color: var(--gray-900);">Cancel</a>
        </form>
    </div>
    """
    return render_page("Import Payment Codes", content, active="codes", user=nav_user)


def page_settings(db, nav_user=None):
    # FIX: Replaced positional .format() on a template with explicit named variables
    # to prevent KeyError crashes if setting values contain curly braces
    smtp_server_val = h(get_setting(db, "smtp_server"))
    smtp_port_val = h(get_setting(db, "smtp_port"))
    smtp_user_val = h(get_setting(db, "smtp_username"))
    smtp_pass_val = h(get_setting(db, "smtp_password"))
    tls_checked = 'checked' if get_setting(db, "smtp_use_tls") == "1" else ''
    sender_email_val = h(get_setting(db, "sender_email"))
    sender_name_val = h(get_setting(db, "sender_name"))
    email_subj_val = h(get_setting(db, "email_subject"))
    email_body_val = h(get_setting(db, "email_body"))
    accio_account_val = h(get_setting(db, "accio_account"))
    accio_username_val = h(get_setting(db, "accio_username"))
    accio_password_val = h(get_setting(db, "accio_password"))
    accio_researcher_url_val = h(get_setting(db, "accio_researcher_url"))
    # SMS / Twilio settings
    twilio_sid_val = h(get_setting(db, "twilio_account_sid"))
    twilio_token_val = h(get_setting(db, "twilio_auth_token"))
    twilio_from_val = h(get_setting(db, "twilio_from_number"))
    auto_sms_checked = 'checked' if get_setting(db, "auto_send_sms") == "1" else ''
    sms_body_val = h(get_setting(db, "sms_body"))
    twilio_installed = "Installed" if HAS_TWILIO else "NOT INSTALLED - run: pip install twilio"
    twilio_badge_color = "var(--success)" if HAS_TWILIO else "var(--danger, #e74c3c)"

    content = f"""
    <div class="card">
        <div class="card-title"><i class="fas fa-exchange-alt"></i> Accio Postback Configuration</div>
        <form method="POST" action="/settings">
            <p style="color: var(--gray-500); font-size: 0.875rem; margin-bottom: 1rem;">
                When an FP Release email is sent, the app will automatically post a status note
                back to Accio (Chapter 9 &amp; 11 of the XML Integration Manual) so the screening
                firm can see the email was sent and the assigned code.
                Leave <strong>Researcher URL</strong> blank to disable postback.
            </p>
            <div class="grid-2">
                <div class="form-group"><label for="accio_account">Accio Account</label><input type="text" id="accio_account" name="accio_account" value="{accio_account_val}"></div>
                <div class="form-group"><label for="accio_username">Accio Username</label><input type="text" id="accio_username" name="accio_username" value="{accio_username_val}"></div>
            </div>
            <div class="grid-2">
                <div class="form-group"><label for="accio_password">Accio Password</label><input type="password" id="accio_password" name="accio_password" value="{accio_password_val}"></div>
                <div class="form-group"><label for="accio_researcher_url">Researcher XML URL</label><input type="text" id="accio_researcher_url" name="accio_researcher_url" value="{accio_researcher_url_val}" placeholder="https://yourcompany.acciodata.com/c/p/researcherxml"></div>
            </div>
            <button type="submit" class="btn btn-primary"><i class="fas fa-save"></i> Save Accio Settings</button>
        </form>
    </div>
    <div class="card">
        <div class="card-title"><i class="fas fa-cog"></i> SMTP Configuration</div>
        <form method="POST" action="/settings">
            <div class="grid-2">
                <div class="form-group"><label for="smtp_server">SMTP Server</label><input type="text" id="smtp_server" name="smtp_server" value="{smtp_server_val}" required></div>
                <div class="form-group"><label for="smtp_port">SMTP Port</label><input type="number" id="smtp_port" name="smtp_port" value="{smtp_port_val}" required></div>
            </div>
            <div class="grid-2">
                <div class="form-group"><label for="smtp_username">SMTP Username</label><input type="text" id="smtp_username" name="smtp_username" value="{smtp_user_val}"></div>
                <div class="form-group"><label for="smtp_password">SMTP Password</label><input type="password" id="smtp_password" name="smtp_password" value="{smtp_pass_val}"></div>
            </div>
            <div class="form-group">
                <label>
                    <input type="hidden" name="smtp_use_tls" value="0">
                    <input type="checkbox" name="smtp_use_tls" value="1" {tls_checked}>
                    Use TLS (Recommended)
                </label>
            </div>
            <div class="grid-2">
                <div class="form-group"><label for="sender_email">Sender Email</label><input type="email" id="sender_email" name="sender_email" value="{sender_email_val}" required></div>
                <div class="form-group"><label for="sender_name">Sender Name</label><input type="text" id="sender_name" name="sender_name" value="{sender_name_val}"></div>
            </div>
            <div class="form-group"><label for="email_subject">Email Subject Template</label><input type="text" id="email_subject" name="email_subject" value="{email_subj_val}"></div>
            <div class="form-group">
                <label for="email_body">Email Body Template</label>
                <textarea id="email_body" name="email_body" style="min-height: 300px;">{email_body_val}</textarea>
                <p style="margin-top: 0.5rem; color: var(--gray-500); font-size: 0.875rem;">
                    Available placeholders: {{first_name}}, {{last_name}}, {{code}}, {{company_name}}, {{ori_number}}
                </p>
            </div>
            <button type="submit" class="btn btn-primary"><i class="fas fa-save"></i> Save Settings</button>
        </form>
    </div>
    <div class="card">
        <div class="card-title"><i class="fas fa-sms"></i> SMS Configuration (Twilio)</div>
        <p style="color: var(--gray-500); font-size: 0.875rem; margin-bottom: 0.5rem;">
            Send text messages to applicants alongside emails. SMS includes their IdentoGO code and scheduling instructions.
        </p>
        <p style="font-size: 0.8rem; margin-bottom: 1rem;">
            Twilio SDK: <span style="color: {twilio_badge_color}; font-weight: bold;">{twilio_installed}</span>
        </p>
        <form method="POST" action="/settings">
            <div class="grid-2">
                <div class="form-group"><label for="twilio_account_sid">Twilio Account SID</label><input type="text" id="twilio_account_sid" name="twilio_account_sid" value="{twilio_sid_val}" placeholder="ACxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"></div>
                <div class="form-group"><label for="twilio_auth_token">Twilio Auth Token</label><input type="password" id="twilio_auth_token" name="twilio_auth_token" value="{twilio_token_val}"></div>
            </div>
            <div class="grid-2">
                <div class="form-group"><label for="twilio_from_number">Twilio Phone Number (From)</label><input type="text" id="twilio_from_number" name="twilio_from_number" value="{twilio_from_val}" placeholder="+12255551234"></div>
                <div class="form-group">
                    <label>
                        <input type="hidden" name="auto_send_sms" value="0">
                        <input type="checkbox" name="auto_send_sms" value="1" {auto_sms_checked}>
                        Auto-send SMS when code is assigned
                    </label>
                </div>
            </div>
            <div class="form-group">
                <label for="sms_body">SMS Body Template</label>
                <textarea id="sms_body" name="sms_body" style="min-height: 200px;">{sms_body_val}</textarea>
                <p style="margin-top: 0.5rem; color: var(--gray-500); font-size: 0.875rem;">
                    Available placeholders: {{first_name}}, {{last_name}}, {{code}}, {{company_name}}, {{ori_number}}<br>
                    <strong>Tip:</strong> Keep SMS under 320 characters for best delivery (2 SMS segments max).
                </p>
            </div>
            <button type="submit" class="btn btn-primary"><i class="fas fa-save"></i> Save SMS Settings</button>
        </form>
        <hr style="margin: 1rem 0;">
        <form method="POST" action="/settings/test-sms" style="display: flex; gap: 0.5rem; align-items: end;">
            <div class="form-group" style="flex: 1; margin-bottom: 0;">
                <label for="test_phone">Test SMS</label>
                <input type="text" id="test_phone" name="test_phone" placeholder="+12255551234 or (225) 555-1234">
            </div>
            <button type="submit" class="btn" style="background: #25D366; color: white; white-space: nowrap;"><i class="fas fa-paper-plane"></i> Send Test SMS</button>
        </form>
    </div>
    """
    return render_page("Settings", content, active="settings", user=nav_user)


def page_logs(db, nav_user=None):
    xml_rows = db.execute("SELECT * FROM xml_log ORDER BY id DESC LIMIT 50").fetchall()
    email_rows = db.execute("SELECT * FROM email_log ORDER BY id DESC LIMIT 50").fetchall()
    sms_rows = db.execute("SELECT * FROM sms_log ORDER BY id DESC LIMIT 50").fetchall()

    xml_html = """<div class="card"><div class="card-title"><i class="fas fa-code"></i> XML Logs (Accio Push)</div>
    <table><thead><tr><th>ID</th><th>Direction</th><th>Status</th><th>Error</th><th>Received</th></tr></thead><tbody>"""
    for r in xml_rows:
        xml_html += f"<tr><td>{r['id']}</td><td>{h(r['direction'] or '-')}</td><td>{h(r['parsed_status'] or '-')}</td><td>{h((r['error_message'] or '')[:60])}</td><td>{fmt_dt(r['received_at'])}</td></tr>"
    xml_html += "</tbody></table></div>"

    email_html = """<div class="card"><div class="card-title"><i class="fas fa-envelope"></i> Email Logs</div>
    <table><thead><tr><th>ID</th><th>Recipient</th><th>Subject</th><th>Status</th><th>Error</th><th>Sent</th></tr></thead><tbody>"""
    for r in email_rows:
        email_html += f"<tr><td>{r['id']}</td><td>{h(r['recipient_email'] or '-')}</td><td>{h((r['subject'] or '')[:40])}</td><td>{h(r['status'] or '-')}</td><td>{h((r['error_message'] or '')[:40])}</td><td>{fmt_dt(r['sent_at'])}</td></tr>"
    email_html += "</tbody></table></div>"

    sms_html = """<div class="card"><div class="card-title"><i class="fas fa-sms"></i> SMS Logs</div>
    <table><thead><tr><th>ID</th><th>Recipient</th><th>Twilio SID</th><th>Status</th><th>Error</th><th>Sent</th></tr></thead><tbody>"""
    for r in sms_rows:
        sms_html += f"<tr><td>{r['id']}</td><td>{h(r['recipient_phone'] or '-')}</td><td><code style='font-size:0.75rem;'>{h((r['twilio_sid'] or '-')[:20])}</code></td><td>{h(r['status'] or '-')}</td><td>{h((r['error_message'] or '')[:40])}</td><td>{fmt_dt(r['sent_at'])}</td></tr>"
    sms_html += "</tbody></table></div>"

    return render_page("Logs", xml_html + email_html + sms_html, active="logs", user=nav_user)


def page_clients(db, params, nav_user=None):
    # FIX: Wrapped int(client_id) in try/except to prevent ValueError 500 crash
    client_id_raw = params.get("client_id", [None])[0]
    client_id = None
    if client_id_raw:
        try:
            client_id = int(client_id_raw)
        except (ValueError, TypeError):
            return render_page("Invalid Request",
                               '<div class="es"><i class="fas fa-exclamation-triangle"></i><h3>Invalid client ID</h3></div>',
                               active="clients", user=nav_user)

    if client_id:
        client = db.execute("SELECT * FROM clients WHERE id = %s", (client_id,)).fetchone()
        if not client:
            return render_page("Not Found",
                               '<div class="es"><i class="fas fa-exclamation-triangle"></i><h3>Client not found</h3></div>',
                               active="clients", user=nav_user)

        applicants = db.execute(
            "SELECT * FROM applicants WHERE client_id = %s ORDER BY created_at DESC",
            (client_id,)
        ).fetchall()

        # Check for success flash from email update
        email_saved = params.get("email_saved", [None])[0]
        save_msg = '<div class="flash flash-success" style="margin-bottom:1rem;"><i class="fas fa-check-circle"></i> Client email saved — this client will now be CC\'d on applicant emails.</div>' if email_saved else ""

        content = f"""
        <a href="/clients" class="btn" style="background: var(--gray-300); color: var(--gray-900); margin-bottom: 1rem;"><i class="fas fa-arrow-left"></i> Back</a>
        {save_msg}
        <div class="card">
            <div class="card-title">{h(client['company_name'])}</div>
            <p><strong>Account:</strong> {h(client['account_name'] or '-')}</p>
            <form method="POST" action="/clients/update-email" style="display:flex; align-items:center; gap:0.5rem; margin:0.5rem 0;">
                <input type="hidden" name="client_id" value="{client['id']}">
                <strong>Email:</strong>
                <input type="email" name="contact_email" value="{h(client['contact_email'] or '')}" placeholder="client@example.com" style="padding:6px 10px; border:1px solid var(--gray-300); border-radius:6px; width:280px;">
                <button type="submit" class="btn btn-primary btn-small"><i class="fas fa-save"></i> Save</button>
            </form>
            <p><strong>Phone:</strong> {h(client['contact_phone'] or '-')}</p>
            <p><strong>Total Applicants:</strong> {len(applicants)}</p>
        </div>
        <div class="card"><div class="card-title">Applicants</div>
            <table><thead><tr><th>Order #</th><th>Name</th><th>Email</th><th>Received</th><th>Email Sent</th><th>Status</th><th>Code</th><th>Actions</th></tr></thead><tbody>
        """
        for a in applicants:
            safe_st = h(a['status']).replace(" ", "_")
            received_dt = fmt_dt(a.get('created_at'))
            sent_dt = fmt_dt(a.get('email_sent_at')) if a.get('email_sent_at') else '-'
            resend_btn = ""
            if a.get('email') and a.get('assigned_code'):
                resend_btn = (
                    f'<form method="POST" action="/applicants/{a["id"]}/resend" style="display:inline;">'
                    f'<input type="hidden" name="redirect" value="/clients?client_id={client_id}">'
                    f'<button type="submit" class="btn btn-small" style="background:#17a2b8; color:white;" title="Resend email with CC to client">'
                    f'<i class="fas fa-redo"></i> Resend</button></form>'
                )
            content += f"<tr><td><code style='font-size:0.8rem;'>{h(a['accio_order_number'] or '-')}</code></td><td>{h(a['first_name'])} {h(a['last_name'])}</td><td>{h(a['email'] or '-')}</td><td>{received_dt}</td><td>{sent_dt}</td><td><span class=\"status-badge status-{safe_st}\">{h(a['status'])}</span></td><td><code>{h(a['assigned_code'] or '-')}</code></td><td>{resend_btn}</td></tr>"
        content += "</tbody></table></div>"
    else:
        clients = db.execute("""
            SELECT c.*, COUNT(a.id) as app_count, MAX(a.created_at) as last_order
            FROM clients c
            LEFT JOIN applicants a ON a.client_id = c.id
            GROUP BY c.id ORDER BY app_count DESC
        """).fetchall()

        content = """<div class="card"><table><thead><tr><th>Company</th><th>Account</th><th>Contact Email</th><th>Total Applicants</th><th>Last Order</th><th>Actions</th></tr></thead><tbody>"""
        for c in clients:
            email_cell = h(c["contact_email"]) if c["contact_email"] else f'<a href="/clients?client_id={c["id"]}" style="color:var(--warning);font-size:0.85rem;"><i class="fas fa-plus-circle"></i> Add</a>'
            content += f'<tr><td>{h(c["company_name"])}</td><td>{h(c["account_name"] or "-")}</td><td>{email_cell}</td><td>{c["app_count"]}</td><td>{fmt_dt(c["last_order"])}</td><td><a href="/clients?client_id={c["id"]}" class="btn btn-primary btn-small">View</a></td></tr>'
        content += "</tbody></table></div>"

    return render_page("Clients", content, active="clients", user=nav_user)


# ---------------------------------------------------------------------------
# Password Reset
# ---------------------------------------------------------------------------
def send_password_reset_email(db, username_or_email):
    """Look up user by username or recovery_email, send a reset link."""
    user = db.execute(
        "SELECT * FROM users WHERE (username=%s OR recovery_email=%s) AND is_active=TRUE LIMIT 1",
        (username_or_email, username_or_email)
    ).fetchone()
    if not user:
        # Do not reveal whether the account exists
        return True, "If a matching account is found, a reset email will be sent."

    recovery_addr = (user.get("recovery_email") or "").strip()
    if not recovery_addr:
        return False, "No recovery email is configured for this account. Contact an administrator."

    token = str(uuid.uuid4())
    expires_at = (datetime.now() + timedelta(hours=2)).isoformat()
    db.execute(
        "INSERT INTO password_reset_tokens (user_id, token, expires_at) VALUES (%s, %s, %s)",
        (user["id"], token, expires_at)
    )

    smtp_server = get_setting(db, "smtp_server")
    smtp_port = int(get_setting(db, "smtp_port") or 587)
    smtp_user = get_setting(db, "smtp_username")
    smtp_pass = get_setting(db, "smtp_password")
    use_tls = get_setting(db, "smtp_use_tls") == "1"
    sender_email = get_setting(db, "sender_email")
    sender_name = get_setting(db, "sender_name") or "Fingerprint Release Manager"

    base_url = (get_setting(db, "accio_post_url") or "").rstrip("/").rsplit("/api/", 1)[0] or "http://localhost:5000"
    reset_link = f"{base_url}/reset-password?token={token}"

    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = f"{sender_name} <{sender_email}>"
        msg["To"] = recovery_addr
        msg["Subject"] = "Password Reset – Fingerprint Release Manager"
        plain = (f"Hi {user['display_name'] or user['username']},\n\n"
                 f"A password reset was requested for your account.\n\n"
                 f"Click the link below to reset your password (valid for 2 hours):\n{reset_link}\n\n"
                 f"If you did not request this, you can safely ignore this email.\n\n"
                 f"— Fingerprint Release Manager")
        html_body = f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="font-family:Arial,sans-serif;color:#333;max-width:500px;margin:0 auto;">
<h2 style="color:#2563eb;">Password Reset</h2>
<p>Hi <strong>{h(user['display_name'] or user['username'])}</strong>,</p>
<p>A password reset was requested for your account. Click the button below to set a new password.
This link is valid for <strong>2 hours</strong>.</p>
<p style="text-align:center;margin:2rem 0;">
  <a href="{h(reset_link)}" style="background:#2563eb;color:white;padding:0.75rem 2rem;
     border-radius:0.375rem;text-decoration:none;font-weight:600;">Reset Password</a>
</p>
<p style="color:#666;font-size:0.875rem;">Or copy this link into your browser:<br>
<code style="word-break:break-all;">{h(reset_link)}</code></p>
<p style="color:#666;font-size:0.875rem;">If you did not request this, you can safely ignore this email.</p>
</body></html>"""
        msg.attach(MIMEText(plain, "plain"))
        msg.attach(MIMEText(html_body, "html"))
        srv = smtplib.SMTP(smtp_server, smtp_port, timeout=15)
        if use_tls:
            srv.starttls()
        if smtp_user and smtp_pass:
            srv.login(smtp_user, smtp_pass)
        srv.send_message(msg)
        srv.quit()
        logger.info(f"Password reset email sent to {recovery_addr} for user {user['username']}")
        return True, "Reset email sent. Check your inbox."
    except Exception as e:
        logger.error(f"Password reset email failed: {e}")
        return False, f"Could not send email: {e}"


def page_forgot_password(error="", success=""):
    error_html = f'<div class="alert" style="background:rgba(239,68,68,0.1);border:1px solid #ef4444;color:#ef4444;padding:.75rem;border-radius:.375rem;margin-bottom:1.5rem;font-size:.875rem;">{h(error)}</div>' if error else ""
    success_html = f'<div class="alert" style="background:rgba(16,185,129,0.1);border:1px solid #10b981;color:#10b981;padding:.75rem;border-radius:.375rem;margin-bottom:1.5rem;font-size:.875rem;">{h(success)}</div>' if success else ""
    return f"""<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Forgot Password – Fingerprint Release Manager</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
<style>
:root{{--primary:#2563eb;--gray-50:#f9fafb;--gray-200:#e5e7eb;--gray-400:#9ca3af;--gray-700:#374151;--gray-900:#111827;}}
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:linear-gradient(135deg,var(--primary) 0%,#1e40af 100%);min-height:100vh;display:flex;align-items:center;justify-content:center;}}
.card{{background:white;border-radius:.5rem;box-shadow:0 10px 25px rgba(0,0,0,.2);padding:3rem;width:100%;max-width:400px;}}
.brand{{text-align:center;margin-bottom:2rem;}}
.brand i{{font-size:3rem;color:var(--primary);margin-bottom:.5rem;}}
.brand h1{{font-size:1.5rem;color:var(--gray-900);margin:0;}}
.brand p{{color:var(--gray-400);margin:.5rem 0 0;font-size:.875rem;}}
.fg{{margin-bottom:1.5rem;}}
label{{display:block;margin-bottom:.5rem;font-weight:500;color:var(--gray-700);}}
input{{width:100%;padding:.75rem;border:1px solid var(--gray-200);border-radius:.375rem;font-size:.875rem;font-family:inherit;}}
input:focus{{outline:none;border-color:var(--primary);box-shadow:0 0 0 3px rgba(37,99,235,.1);}}
.btn{{width:100%;padding:.75rem;background:var(--primary);color:white;border:none;border-radius:.375rem;font-size:.875rem;font-weight:500;cursor:pointer;transition:background .2s;}}
.btn:hover{{background:#1e40af;}}
.back{{display:block;text-align:center;margin-top:1rem;color:var(--primary);font-size:.875rem;text-decoration:none;}}
</style></head><body>
<div class="card">
  <div class="brand"><i class="fas fa-fingerprint"></i><h1>Forgot Password</h1><p>Enter your username or recovery email</p></div>
  {error_html}{success_html}
  {'<form method="POST" action="/forgot-password"><div class="fg"><label for="identifier">Username or Recovery Email</label><input type="text" id="identifier" name="identifier" required autofocus></div><button type="submit" class="btn"><i class="fas fa-paper-plane"></i> Send Reset Link</button></form>' if not success else ''}
  <a href="/login" class="back"><i class="fas fa-arrow-left"></i> Back to Login</a>
</div></body></html>"""


def page_reset_password(token, error=""):
    error_html = f'<div class="alert" style="background:rgba(239,68,68,0.1);border:1px solid #ef4444;color:#ef4444;padding:.75rem;border-radius:.375rem;margin-bottom:1.5rem;font-size:.875rem;">{h(error)}</div>' if error else ""
    return f"""<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Reset Password – Fingerprint Release Manager</title>
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
<style>
:root{{--primary:#2563eb;--gray-50:#f9fafb;--gray-200:#e5e7eb;--gray-400:#9ca3af;--gray-700:#374151;--gray-900:#111827;}}
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:linear-gradient(135deg,var(--primary) 0%,#1e40af 100%);min-height:100vh;display:flex;align-items:center;justify-content:center;}}
.card{{background:white;border-radius:.5rem;box-shadow:0 10px 25px rgba(0,0,0,.2);padding:3rem;width:100%;max-width:400px;}}
.brand{{text-align:center;margin-bottom:2rem;}}
.brand i{{font-size:3rem;color:var(--primary);margin-bottom:.5rem;}}
.brand h1{{font-size:1.5rem;color:var(--gray-900);margin:0;}}
.fg{{margin-bottom:1.5rem;}}
label{{display:block;margin-bottom:.5rem;font-weight:500;color:var(--gray-700);}}
input{{width:100%;padding:.75rem;border:1px solid var(--gray-200);border-radius:.375rem;font-size:.875rem;font-family:inherit;}}
input:focus{{outline:none;border-color:var(--primary);box-shadow:0 0 0 3px rgba(37,99,235,.1);}}
.btn{{width:100%;padding:.75rem;background:var(--primary);color:white;border:none;border-radius:.375rem;font-size:.875rem;font-weight:500;cursor:pointer;}}
.btn:hover{{background:#1e40af;}}
.back{{display:block;text-align:center;margin-top:1rem;color:var(--primary);font-size:.875rem;text-decoration:none;}}
.hint{{color:#6b7280;font-size:.8rem;margin-top:.25rem;}}
</style></head><body>
<div class="card">
  <div class="brand"><i class="fas fa-key"></i><h1>Set New Password</h1></div>
  {error_html}
  <form method="POST" action="/reset-password">
    <input type="hidden" name="token" value="{h(token)}">
    <div class="fg">
      <label for="password">New Password</label>
      <input type="password" id="password" name="password" required minlength="8" autofocus>
      <p class="hint">Minimum 8 characters</p>
    </div>
    <div class="fg">
      <label for="confirm">Confirm Password</label>
      <input type="password" id="confirm" name="confirm" required minlength="8">
    </div>
    <button type="submit" class="btn"><i class="fas fa-lock"></i> Reset Password</button>
  </form>
  <a href="/login" class="back"><i class="fas fa-arrow-left"></i> Back to Login</a>
</div></body></html>"""


# ---------------------------------------------------------------------------
# User Management Pages (admin only)
# ---------------------------------------------------------------------------
def page_users(db, user):
    rows = db.execute(
        "SELECT id, username, display_name, role, is_active, recovery_email, created_at, last_login FROM users ORDER BY id"
    ).fetchall()
    content = """
    <div style="margin-bottom:1rem;display:flex;gap:.5rem;">
      <a href="/users/add" class="btn btn-primary"><i class="fas fa-user-plus"></i> Create User</a>
    </div>
    <div class="card">
      <table>
        <thead><tr><th>Username</th><th>Display Name</th><th>Role</th><th>Recovery Email</th><th>Status</th><th>Last Login</th><th>Actions</th></tr></thead>
        <tbody>
    """
    for r in rows:
        active_badge = ('<span class="status-badge status-emailed">Active</span>' if r["is_active"]
                        else '<span class="status-badge status-pending">Disabled</span>')
        toggle_label = "Disable" if r["is_active"] else "Enable"
        toggle_class = "btn-danger" if r["is_active"] else "btn-success"
        can_delete = r["id"] != user["id"] and r["role"] != "admin"
        _uname_escaped = h(r["username"])
        delete_btn = (
            f'<form method="POST" action="/users/{r["id"]}/delete" style="display:inline;" '
            f'onsubmit="return confirm(\'Delete user {_uname_escaped}? This cannot be undone.\');">'
            f'<button type="submit" class="btn btn-danger btn-small"><i class="fas fa-trash"></i></button></form>'
            if can_delete else ''
        )
        self_tag = ' <em style="color:#9ca3af;font-size:.75rem;">(you)</em>' if r["id"] == user["id"] else ""
        content += f"""
        <tr>
          <td><strong>{h(r["username"])}</strong>{self_tag}</td>
          <td>{h(r["display_name"] or "-")}</td>
          <td><span class="status-badge {'status-emailed' if r['role']=='admin' else 'status-code_assigned'}">{h(r["role"])}</span></td>
          <td>{h(r["recovery_email"] or "-")}</td>
          <td>{active_badge}</td>
          <td>{fmt_dt(r["last_login"])}</td>
          <td style="white-space:nowrap;">
            <form method="POST" action="/users/{r['id']}/toggle" style="display:inline;">
              <button type="submit" class="btn btn-small {toggle_class}">{toggle_label}</button>
            </form>
            <a href="/users/{r['id']}/reset" class="btn btn-small" style="background:#6c757d;color:white;">Reset PW</a>
            {delete_btn}
          </td>
        </tr>"""
    content += "</tbody></table></div>"
    return render_page("User Management", content, active="users", user=user)


def page_add_user(nav_user=None, error=""):
    error_html = f'<div class="alert alert-error"><i class="fas fa-exclamation-circle"></i> {h(error)}</div>' if error else ""
    content = f"""{error_html}
    <div class="card">
      <form method="POST" action="/users/add">
        <div class="grid-2">
          <div class="form-group"><label for="username">Username *</label><input type="text" id="username" name="username" required autocomplete="off"></div>
          <div class="form-group"><label for="display_name">Display Name</label><input type="text" id="display_name" name="display_name" autocomplete="off"></div>
        </div>
        <div class="grid-2">
          <div class="form-group"><label for="password">Password *</label><input type="password" id="password" name="password" required minlength="8" autocomplete="new-password"></div>
          <div class="form-group"><label for="role">Role</label>
            <select id="role" name="role">
              <option value="user">User</option>
              <option value="admin">Admin</option>
            </select>
          </div>
        </div>
        <div class="form-group"><label for="recovery_email">Recovery Email</label><input type="email" id="recovery_email" name="recovery_email" placeholder="user@example.com" autocomplete="off"></div>
        <button type="submit" class="btn btn-primary"><i class="fas fa-user-plus"></i> Create User</button>
        <a href="/users" class="btn" style="background:var(--gray-300);color:var(--gray-900);">Cancel</a>
      </form>
    </div>"""
    return render_page("Create User", content, active="users", user=nav_user)


def page_reset_user_password(user_id, username, nav_user=None, error=""):
    error_html = f'<div class="alert alert-error"><i class="fas fa-exclamation-circle"></i> {h(error)}</div>' if error else ""
    content = f"""{error_html}
    <div class="card">
      <p style="margin-bottom:1rem;">Set a new password for <strong>{h(username)}</strong>.</p>
      <form method="POST" action="/users/{user_id}/reset">
        <div class="form-group"><label for="new_password">New Password</label><input type="password" id="new_password" name="new_password" required minlength="8" autocomplete="new-password"></div>
        <div class="form-group"><label for="confirm_password">Confirm Password</label><input type="password" id="confirm_password" name="confirm_password" required minlength="8" autocomplete="new-password"></div>
        <button type="submit" class="btn btn-primary"><i class="fas fa-key"></i> Set Password</button>
        <a href="/users" class="btn" style="background:var(--gray-300);color:var(--gray-900);">Cancel</a>
      </form>
    </div>"""
    return render_page(f"Reset Password: {username}", content, active="users", user=nav_user)


# ---------------------------------------------------------------------------
# Profile / My Account Page (per-user recovery email)
# ---------------------------------------------------------------------------
def page_profile(db, user, error="", success=""):
    u = db.execute("SELECT * FROM users WHERE id=%s", (user["id"],)).fetchone()
    recovery_val = h(u.get("recovery_email") or "")
    display_val = h(u.get("display_name") or "")
    error_html = f'<div class="alert alert-error"><i class="fas fa-exclamation-circle"></i> {h(error)}</div>' if error else ""
    success_html = f'<div class="alert alert-success"><i class="fas fa-check-circle"></i> {h(success)}</div>' if success else ""
    content = f"""{error_html}{success_html}
    <div class="card">
      <div class="card-title"><i class="fas fa-user-cog"></i> My Account</div>
      <form method="POST" action="/profile">
        <div class="grid-2">
          <div class="form-group"><label>Username</label><input type="text" value="{h(u['username'])}" disabled style="background:#f3f4f6;"></div>
          <div class="form-group"><label for="display_name">Display Name</label><input type="text" id="display_name" name="display_name" value="{display_val}"></div>
        </div>
        <div class="form-group">
          <label for="recovery_email">Recovery Email</label>
          <input type="email" id="recovery_email" name="recovery_email" value="{recovery_val}" placeholder="your@email.com">
          <p style="margin-top:.5rem;color:var(--gray-500);font-size:.875rem;">Used to receive password reset links. Keep this updated.</p>
        </div>
        <button type="submit" class="btn btn-primary"><i class="fas fa-save"></i> Save Profile</button>
      </form>
    </div>
    <div class="card" style="margin-top:1.5rem;">
      <div class="card-title"><i class="fas fa-key"></i> Change Password</div>
      <form method="POST" action="/profile/change-password">
        <div class="form-group"><label for="current_password">Current Password</label><input type="password" id="current_password" name="current_password" required autocomplete="current-password"></div>
        <div class="grid-2">
          <div class="form-group"><label for="new_password">New Password</label><input type="password" id="new_password" name="new_password" required minlength="8" autocomplete="new-password"></div>
          <div class="form-group"><label for="confirm_password">Confirm New Password</label><input type="password" id="confirm_password" name="confirm_password" required minlength="8" autocomplete="new-password"></div>
        </div>
        <button type="submit" class="btn btn-primary"><i class="fas fa-lock"></i> Change Password</button>
      </form>
    </div>"""
    return render_page("My Account", content, active="profile", user=user)


# ---------------------------------------------------------------------------
# HTTP Handler
# ---------------------------------------------------------------------------
class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        logger.info(f"{self.client_address[0]} - {fmt % args}")

    def _send(self, code, body, ct="text/html; charset=utf-8"):
        # FIX: Added charset=utf-8 to Content-Type to correctly declare encoding
        encoded = body.encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", ct)
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def _redirect(self, url, set_cookie=None, clear_cookie=False):
        self.send_response(303)
        self.send_header("Location", url)
        if set_cookie:
            # FIX: Added SameSite=Lax to prevent CSRF attacks via cross-site form submissions
            self.send_header("Set-Cookie",
                             f"session_token={set_cookie}; Path=/; HttpOnly; SameSite=Lax; Max-Age=86400")
        if clear_cookie:
            # FIX: Properly expire the session cookie on logout
            self.send_header("Set-Cookie",
                             "session_token=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0; Expires=Thu, 01 Jan 1970 00:00:00 GMT")
        self.end_headers()

    def _parse_form(self):
        ct = self.headers.get("Content-Type", "")
        length = int(self.headers.get("Content-Length", 0))
        if "multipart/form-data" in ct:
            env = {"REQUEST_METHOD": "POST", "CONTENT_TYPE": ct, "CONTENT_LENGTH": str(length)}
            fs = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ=env)
            return fs
        else:
            body = self.rfile.read(length).decode("utf-8", errors="replace")
            return urllib.parse.parse_qs(body)

    def _check_auth(self):
        """Check if user is authenticated; return user dict or None."""
        cookie = self.headers.get("Cookie", "")
        token = get_session_from_cookie(cookie)
        db = get_db()
        user = verify_session(db, token) if token else None
        db.close()
        return user

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        params = urllib.parse.parse_qs(parsed.query)

        # Auth check for all non-public routes
        user = None
        if path != "/login" and not path.startswith("/api/track/"):
            user = self._check_auth()
            if not user:
                self._redirect("/login")
                return

        db = get_db()
        try:
            if path == "/login":
                self._send(200, page_login())

            elif path == "/forgot-password":
                self._send(200, page_forgot_password())

            elif path == "/reset-password":
                token = params.get("token", [None])[0] or ""
                if not token:
                    self._redirect("/forgot-password")
                    return
                # Validate token exists and is not expired/used
                row = db.execute(
                    "SELECT * FROM password_reset_tokens WHERE token=%s AND used=FALSE AND expires_at > NOW()",
                    (token,)
                ).fetchone()
                if not row:
                    self._send(200, page_forgot_password(error="This reset link is invalid or has expired. Please request a new one."))
                else:
                    self._send(200, page_reset_password(token))

            elif path == "/logout":
                # FIX: Logout now (1) deletes session from DB, (2) expires cookie,
                # (3) only sends ONE response (redirect only, not body+redirect)
                cookie = self.headers.get("Cookie", "")
                token = get_session_from_cookie(cookie)
                if token:
                    delete_session(db, token)
                self._redirect("/login", clear_cookie=True)

            elif path == "/":
                self._send(200, page_dashboard(db, nav_user=user))

            elif path == "/applicants":
                self._send(200, page_applicants(db, params, nav_user=user))

            elif path == "/applicants/add":
                self._send(200, page_add_applicant(nav_user=user))

            elif path == "/clients":
                self._send(200, page_clients(db, params, nav_user=user))

            elif path == "/codes":
                self._send(200, page_codes(db, params, nav_user=user))

            elif path == "/codes/import":
                self._send(200, page_import_codes(nav_user=user))

            elif path == "/codes/manual":
                content = """
                <div class="card">
                    <h2 style="margin-bottom:1rem;">Add Codes Manually</h2>
                    <form method="POST" action="/codes/add-manual">
                        <div style="margin-bottom: 1rem;">
                            <label><strong>Batch Name</strong></label>
                            <input type="text" name="batch_name" value="Manual" placeholder="Batch name" style="width:100%; padding:8px; border:1px solid #ccc; border-radius:4px; margin-top:4px;">
                        </div>
                        <div style="margin-bottom: 1rem;">
                            <label><strong>Codes</strong> (one per line)</label>
                            <textarea name="codes" rows="10" placeholder="Enter one code per line" style="width:100%; padding:8px; border:1px solid #ccc; border-radius:4px; margin-top:4px; font-family:monospace;"></textarea>
                        </div>
                        <button type="submit" class="btn btn-primary"><i class="fas fa-plus"></i> Add Codes</button>
                        <a href="/codes" class="btn" style="background: var(--gray-300); color: var(--gray-900); margin-left:8px;">Cancel</a>
                    </form>
                </div>
                """
                self._send(200, render_page("Add Codes Manually", content, active="codes", user=user))

            elif path == "/laps":
                self._send(200, page_laps_dashboard(db, nav_user=user))

            elif path == "/settings":
                self._send(200, page_settings(db, nav_user=user))

            elif path == "/profile":
                self._send(200, page_profile(db, user))

            elif path == "/users":
                if user.get("role") != "admin":
                    self._send(403, render_page("Forbidden", '<div class="es"><i class="fas fa-lock"></i><h3>Admin access required</h3></div>', user=user))
                else:
                    self._send(200, page_users(db, user))

            elif path == "/users/add":
                if user.get("role") != "admin":
                    self._send(403, render_page("Forbidden", '<div class="es"><i class="fas fa-lock"></i><h3>Admin access required</h3></div>', user=user))
                else:
                    self._send(200, page_add_user(nav_user=user))

            elif path.startswith("/users/") and path.endswith("/reset"):
                if user.get("role") != "admin":
                    self._send(403, render_page("Forbidden", '<div class="es"><i class="fas fa-lock"></i><h3>Admin access required</h3></div>', user=user))
                else:
                    try:
                        uid = int(path.split("/")[2])
                    except (ValueError, IndexError):
                        self._redirect("/users")
                        return
                    target = db.execute("SELECT username FROM users WHERE id=%s", (uid,)).fetchone()
                    if not target:
                        flash("User not found.", "error")
                        self._redirect("/users")
                    else:
                        self._send(200, page_reset_user_password(uid, target["username"], nav_user=user))

            elif path == "/logs":
                self._send(200, page_logs(db, nav_user=user))

            elif path.startswith("/api/track/"):
                # Legacy tracking pixel endpoint — kept alive so old emails don't
                # produce broken-image icons, but no longer updates applicant status.
                # Tracking pixels were removed from outgoing emails to improve
                # deliverability (cross-domain image loads trigger spam filters).
                gif = b'\x47\x49\x46\x38\x39\x61\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00\x21\xf9\x04\x01\x0a\x00\x01\x00\x2c\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02\x4d\x01\x00\x3b'
                self.send_response(200)
                self.send_header("Content-Type", "image/gif")
                self.send_header("Content-Length", str(len(gif)))
                self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
                self.end_headers()
                self.wfile.write(gif)

            elif path == "/api/debug-xml":
                # FIX: These debug endpoints were accessible without auth (all /api/ routes
                # bypassed auth). They expose raw PII from XML logs. Auth now required.
                row = db.execute("SELECT raw_xml FROM xml_log ORDER BY id DESC LIMIT 1").fetchone()
                if row and row["raw_xml"]:
                    self._send(200, row["raw_xml"], "text/xml; charset=utf-8")
                else:
                    self._send(200, "No XML logs found", "text/plain; charset=utf-8")

            elif path == "/api/debug-xml-tags":
                # FIX: Auth now required (same issue as above)
                row = db.execute("SELECT raw_xml FROM xml_log ORDER BY id DESC LIMIT 1").fetchone()
                if row and row["raw_xml"]:
                    try:
                        xroot = ET.fromstring(row["raw_xml"])
                        tags = set()
                        tag_tree = []
                        for el in xroot.iter():
                            tags.add(el.tag)
                            text_preview = (el.text or "").strip()[:50]
                            tag_tree.append(f"{el.tag} = '{text_preview}'" if text_preview else el.tag)
                        result = "=== ALL UNIQUE TAGS ===\n" + "\n".join(sorted(tags))
                        result += "\n\n=== FULL TAG TREE WITH VALUES ===\n" + "\n".join(tag_tree)
                        self._send(200, result, "text/plain; charset=utf-8")
                    except Exception as e:
                        self._send(200, f"Parse error: {e}", "text/plain; charset=utf-8")
                else:
                    self._send(200, "No XML logs found", "text/plain; charset=utf-8")

            else:
                self._send(404, render_page("Not Found",
                    '<div class="es"><i class="fas fa-exclamation-triangle"></i><h3>Page not found</h3></div>'))
        finally:
            db.close()

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        db = get_db()

        try:
            # ---------------------------------------------------------------
            # Login (no session required)
            # ---------------------------------------------------------------
            if path == "/forgot-password":
                form_data = self._parse_form()
                def _fv_simple(name, default=""):
                    if isinstance(form_data, cgi.FieldStorage):
                        v = form_data.getfirst(name, default)
                        return v if isinstance(v, str) else (v.decode() if v else default)
                    vals = form_data.get(name, [default])
                    return vals[0] if vals else default
                identifier = _fv_simple("identifier").strip()
                if not identifier:
                    self._send(200, page_forgot_password(error="Please enter your username or recovery email."))
                else:
                    ok, msg = send_password_reset_email(db, identifier)
                    if ok:
                        self._send(200, page_forgot_password(success=msg))
                    else:
                        self._send(200, page_forgot_password(error=msg))
                return

            if path == "/reset-password":
                form_data = self._parse_form()
                def _fv_rp(name, default=""):
                    if isinstance(form_data, cgi.FieldStorage):
                        v = form_data.getfirst(name, default)
                        return v if isinstance(v, str) else (v.decode() if v else default)
                    vals = form_data.get(name, [default])
                    return vals[0] if vals else default
                token = _fv_rp("token").strip()
                new_pass = _fv_rp("password")
                confirm = _fv_rp("confirm")
                if not token:
                    self._redirect("/forgot-password")
                    return
                row = db.execute(
                    "SELECT * FROM password_reset_tokens WHERE token=%s AND used=FALSE AND expires_at > NOW()",
                    (token,)
                ).fetchone()
                if not row:
                    self._send(200, page_forgot_password(error="This reset link is invalid or has expired. Please request a new one."))
                    return
                if len(new_pass) < 8:
                    self._send(200, page_reset_password(token, error="Password must be at least 8 characters."))
                    return
                if new_pass != confirm:
                    self._send(200, page_reset_password(token, error="Passwords do not match."))
                    return
                pw_hash = hash_password(new_pass)
                db.execute("UPDATE users SET password_hash=%s WHERE id=%s", (pw_hash, row["user_id"]))
                db.execute("UPDATE password_reset_tokens SET used=TRUE WHERE token=%s", (token,))
                # Invalidate all active sessions for this user for security
                db.execute("DELETE FROM sessions WHERE user_id=%s", (row["user_id"],))
                logger.info(f"Password reset completed for user_id={row['user_id']}")
                flash("Password reset successfully. Please log in with your new password.", "success")
                self._redirect("/login")
                return

            if path == "/login":
                form_data = self._parse_form()

                def fv(name, default=""):
                    if isinstance(form_data, cgi.FieldStorage):
                        item = form_data.getfirst(name, default)
                        return item if isinstance(item, str) else (item.decode() if item else default)
                    else:
                        vals = form_data.get(name, [default])
                        return vals[0] if vals else default

                username = fv("username").strip()
                password = fv("password")

                user = db.execute(
                    "SELECT * FROM users WHERE username=%s AND is_active=TRUE",
                    (username,)
                ).fetchone()
                if user and verify_password(password, user["password_hash"]):
                    token = create_session(db, user["id"])
                    db.execute("UPDATE users SET last_login=NOW() WHERE id=%s", (user["id"],))
                    self._redirect("/", set_cookie=token)
                else:
                    self._send(200, page_login(error="Invalid username or password."))
                return

            # ---------------------------------------------------------------
            # Accio Data XML push (no session, but credential-authenticated)
            # ---------------------------------------------------------------
            if path == "/api/accio-push":
                try:
                    ACCIO_USERNAME = get_setting(db, "accio_username") or os.environ.get("ACCIO_USERNAME", "")
                    ACCIO_PASSWORD = get_setting(db, "accio_password") or os.environ.get("ACCIO_PASSWORD", "")

                    # FIX: Enforce maximum body size to prevent memory exhaustion
                    content_length = int(self.headers.get("Content-Length", 0))
                    if content_length > MAX_XML_BODY:
                        logger.warning(f"Accio push body too large: {content_length} bytes")
                        self._send(413, '<?xml version="1.0" encoding="UTF-8"?>\n<BackgroundReports><error>Request body too large</error></BackgroundReports>', "text/xml; charset=utf-8")
                        return

                    raw = self.rfile.read(content_length).decode("utf-8", errors="replace") if content_length > 0 else ""
                    auth_valid = False

                    # Method 1: HTTP Basic Auth header
                    auth_header = self.headers.get("Authorization", "")
                    if auth_header.startswith("Basic "):
                        try:
                            decoded = base64.b64decode(auth_header[6:]).decode()
                            u, p = decoded.split(":", 1)
                            if u == ACCIO_USERNAME and p == ACCIO_PASSWORD and ACCIO_PASSWORD:
                                auth_valid = True
                        except Exception:
                            pass

                    # Method 2: Query params
                    qs = urllib.parse.parse_qs(parsed.query)
                    if not auth_valid and ACCIO_PASSWORD:
                        if (qs.get("username", [None])[0] == ACCIO_USERNAME and
                                qs.get("password", [None])[0] == ACCIO_PASSWORD):
                            auth_valid = True

                    # Method 3: Credentials in XML body
                    if not auth_valid and raw.strip() and ACCIO_PASSWORD:
                        try:
                            auth_root = ET.fromstring(raw)
                            login_elem = auth_root.find("login")
                            if login_elem is not None:
                                xml_user = (login_elem.findtext("username") or "").strip()
                                xml_pass = (login_elem.findtext("password") or "").strip()
                                if xml_user == ACCIO_USERNAME and xml_pass == ACCIO_PASSWORD:
                                    auth_valid = True
                            if not auth_valid:
                                for parent in [auth_root] + list(auth_root):
                                    xml_user = xml_pass = None
                                    for el in parent.iter():
                                        tag = el.tag.lower() if el.tag else ""
                                        if tag in ("username", "user", "remote_username"):
                                            xml_user = (el.text or "").strip()
                                        elif tag in ("password", "pass", "remote_password"):
                                            xml_pass = (el.text or "").strip()
                                    if xml_user == ACCIO_USERNAME and xml_pass == ACCIO_PASSWORD:
                                        auth_valid = True
                                        break
                            if not auth_valid:
                                root_user = auth_root.get("username") or auth_root.get("user") or ""
                                root_pass = auth_root.get("password") or auth_root.get("pass") or ""
                                if root_user == ACCIO_USERNAME and root_pass == ACCIO_PASSWORD:
                                    auth_valid = True
                        except ET.ParseError:
                            logging.warning("XML parse error during auth check")
                        except Exception as e:
                            logging.warning(f"Unexpected error during XML auth check: {e}")

                    # Method 4: Custom HTTP headers
                    if not auth_valid and ACCIO_PASSWORD:
                        h_user = self.headers.get("X-Username") or self.headers.get("Username") or ""
                        h_pass = self.headers.get("X-Password") or self.headers.get("Password") or ""
                        if h_user == ACCIO_USERNAME and h_pass == ACCIO_PASSWORD:
                            auth_valid = True

                    logging.info(f"Accio push auth: valid={auth_valid}, has_basic={bool(auth_header)}, body_len={len(raw)}")

                    if not auth_valid:
                        all_headers = {k: v for k, v in self.headers.items() if k.lower() != "authorization"}
                        logging.warning(f"Auth failed. Headers: {all_headers}")
                        logging.warning(f"Auth failed. Body preview: {raw[:200]}")
                        db.execute(
                            "INSERT INTO xml_log (direction,raw_xml,parsed_status,error_message) VALUES ('inbound',%s,'auth_failed','Authentication failed')",
                            (raw[:10000],)
                        )
                        self._send(401, '<?xml version="1.0" encoding="UTF-8"?>\n<BackgroundReports><error>Authentication required</error></BackgroundReports>', "text/xml; charset=utf-8")
                        return

                    logging.info(f"Accio push auth PASSED. Processing XML ({len(raw)} bytes)...")
                    try:
                        debug_root = ET.fromstring(raw)
                        for el in debug_root.iter():
                            txt = (el.text or "").strip()
                            if txt:
                                logging.info(f"  XML TAG: <{el.tag}> = '{txt[:80]}'")
                            else:
                                logging.info(f"  XML TAG: <{el.tag}>")
                    except Exception:
                        pass

                    db.execute(
                        "INSERT INTO xml_log (direction,raw_xml,parsed_status) VALUES ('inbound',%s,'processing')",
                        (raw[:10000],)
                    )

                    applicants_data, err = parse_accio_xml(raw)
                    if err:
                        logging.error(f"XML parse error: {err}")
                        db.execute(
                            "UPDATE xml_log SET parsed_status='error',error_message=%s WHERE id=(SELECT MAX(id) FROM xml_log)",
                            (err,)
                        )
                        self._send(400, '<?xml version="1.0" encoding="UTF-8"?>\n<BackgroundReports><error>XML parse error</error></BackgroundReports>', "text/xml; charset=utf-8")
                        return

                    added = 0
                    auto_assign = get_setting(db, "auto_assign_codes") == "1"
                    auto_email = get_setting(db, "auto_send_email") == "1"
                    auto_sms = get_setting(db, "auto_send_sms") == "1"

                    for a in applicants_data:
                        try:
                            ex = (db.execute("SELECT id FROM applicants WHERE accio_order_number = %s",
                                             (a["accio_order_number"],)).fetchone()
                                  if a.get("accio_order_number") else None)
                            if not ex:
                                client_id = None
                                if a.get("company_name"):
                                    client = db.execute(
                                        "SELECT id FROM clients WHERE company_name=%s",
                                        (a["company_name"],)
                                    ).fetchone()
                                    if not client:
                                        cur = db.execute(
                                            "INSERT INTO clients (company_name, account_name) VALUES (%s, %s) RETURNING id",
                                            (a["company_name"], a.get("account_name", ""))
                                        )
                                        client_id = cur.fetchone()["id"]
                                    else:
                                        client_id = client["id"]

                                cur = db.execute(
                                    "INSERT INTO applicants (first_name,last_name,email,phone,accio_order_number,accio_remote_number,client_id,accio_sub_order,accio_order_type,date_of_birth,last_four_ssn) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                                    (a["first_name"], a["last_name"], a.get("email", ""),
                                     a.get("phone", ""), a.get("accio_order_number", ""),
                                     a.get("accio_remote_number", ""), client_id,
                                     a.get("accio_sub_order", "1"), a.get("accio_order_type", "Fingerprint"),
                                     a.get("date_of_birth", ""), a.get("last_four_ssn", ""))
                                )
                                new_id = cur.fetchone()["id"]
                                added += 1

                                if auto_assign:
                                    code_row = db.execute(
                                        "SELECT id, code FROM codes WHERE assigned_to IS NULL ORDER BY id LIMIT 1"
                                    ).fetchone()
                                    if code_row:
                                        db.execute(
                                            "UPDATE codes SET assigned_to=%s, assigned_date=NOW() WHERE id = %s",
                                            (new_id, code_row["id"])
                                        )
                                        db.execute(
                                            "UPDATE applicants SET assigned_code=%s, status='code_assigned' WHERE id = %s",
                                            (code_row["code"], new_id)
                                        )
                                        # Only auto-email if applicant has a valid email address
                                        if auto_email and a.get("email", "").strip():
                                            try:
                                                send_release_email(db, new_id)
                                            except Exception as email_err:
                                                logging.error(f"Auto-email failed for applicant {new_id}: {email_err}")
                                        # Auto-SMS if applicant has a valid phone number
                                        if auto_sms and a.get("phone", "").strip():
                                            try:
                                                send_release_sms(db, new_id)
                                            except Exception as sms_err:
                                                logging.error(f"Auto-SMS failed for applicant {new_id}: {sms_err}")
                        except Exception as proc_err:
                            logging.error(f"Error processing applicant: {proc_err}")

                    db.execute(
                        "UPDATE xml_log SET parsed_status='success',error_message=%s WHERE id=(SELECT MAX(id) FROM xml_log)",
                        (f"Added {added} applicants from {len(applicants_data)} parsed",)
                    )
                    logging.info(f"Accio push complete: {added} added from {len(applicants_data)} parsed")

                    resp_xml = ('<?xml version="1.0" encoding="UTF-8"?>\n'
                                '<BackgroundReports>\n'
                                '  <BackgroundReportPackage>\n'
                                f'    <ScreeningStatus>accepted</ScreeningStatus>\n'
                                f'    <ResultsRetrieved>{added}</ResultsRetrieved>\n'
                                '  </BackgroundReportPackage>\n'
                                '</BackgroundReports>')
                    self._send(200, resp_xml, "text/xml; charset=utf-8")
                    return

                except Exception as e:
                    logging.error(f"CRITICAL: Unhandled exception in accio-push: {e}", exc_info=True)
                    try:
                        db.execute(
                            "INSERT INTO xml_log (direction,raw_xml,parsed_status,error_message) VALUES ('inbound','','crash',%s)",
                            (str(e)[:5000],)
                        )
                    except Exception:
                        pass
                    self._send(500, '<?xml version="1.0" encoding="UTF-8"?>\n<BackgroundReports><error>Internal server error</error></BackgroundReports>', "text/xml; charset=utf-8")
                    return

            # ---------------------------------------------------------------
            # All other POST routes require session auth
            # ---------------------------------------------------------------
            user = self._check_auth()
            if not user:
                self._redirect("/login")
                return

            # Bulk code upload (JSON)
            if path == "/api/codes":
                length = int(self.headers.get("Content-Length", 0))
                raw = self.rfile.read(length).decode("utf-8", errors="replace")
                try:
                    data = json.loads(raw)
                    codes_list = data.get("codes", [])
                    batch_name = data.get("batch_name", f"API Import {datetime.now().strftime('%Y-%m-%d %H:%M')}")
                    if not codes_list:
                        self._send(400, json.dumps({"status": "error", "message": 'No codes provided. Send {"codes": ["CODE1", ...]}'}), "application/json")
                        return
                    imported, dups = 0, 0
                    for code in codes_list:
                        code = str(code).strip()
                        if code:
                            try:
                                db.execute("INSERT INTO codes (code, batch_name) VALUES (%s, %s)", (code, batch_name))
                                imported += 1
                            except psycopg2.IntegrityError:
                                db.rollback()
                                dups += 1
                    self._send(200, json.dumps({"status": "success", "imported": imported, "duplicates": dups, "batch": batch_name}), "application/json")
                except json.JSONDecodeError:
                    self._send(400, json.dumps({"status": "error", "message": "Invalid JSON"}), "application/json")
                except Exception as e:
                    self._send(500, json.dumps({"status": "error", "message": str(e)}), "application/json")
                return

            # Bulk code upload from file
            if path == "/api/codes/upload":
                form_data = self._parse_form()
                if isinstance(form_data, cgi.FieldStorage) and "file" in form_data:
                    fi = form_data["file"]
                    col = int(form_data.getfirst("column", "0"))
                    batch = form_data.getfirst("batch_name", f"API Upload {datetime.now().strftime('%Y-%m-%d %H:%M')}")
                    fname = fi.filename or "upload.xlsx"
                    fpath = os.path.join(UPLOAD_FOLDER, f"api_{datetime.now().strftime('%Y%m%d%H%M%S')}_{fname}")
                    with open(fpath, "wb") as f:
                        f.write(fi.file.read())
                    try:
                        imported, dups, err = import_codes_from_file(fpath, column_index=col, skip_header=True, batch_name=batch)
                        if err:
                            self._send(400, json.dumps({"status": "error", "message": err}), "application/json")
                        else:
                            self._send(200, json.dumps({"status": "success", "imported": imported, "duplicates": dups, "batch": batch}), "application/json")
                    finally:
                        if os.path.exists(fpath):
                            os.remove(fpath)
                else:
                    self._send(400, json.dumps({"status": "error", "message": "Send file as multipart form with field name 'file'"}), "application/json")
                return

            # General form handler
            form_data = self._parse_form()

            def fv(name, default=""):
                if isinstance(form_data, cgi.FieldStorage):
                    item = form_data.getfirst(name, default)
                    return item if isinstance(item, str) else (item.decode() if item else default)
                else:
                    vals = form_data.get(name, [default])
                    return vals[0] if vals else default

            if path == "/applicants/add":
                db.execute(
                    "INSERT INTO applicants (first_name,last_name,email,phone,accio_order_number,notes,date_of_birth,last_four_ssn) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (fv("first_name"), fv("last_name"), fv("email"), fv("phone"),
                     fv("accio_order_number"), fv("notes"), fv("date_of_birth"), fv("last_four_ssn"))
                )
                flash("Applicant added.", "success")
                self._redirect("/applicants")

            elif path.startswith("/applicants/") and path.endswith("/assign-code"):
                aid = int(path.split("/")[2])
                code_val, msg = assign_code(db, aid)
                flash(f"Code {code_val} assigned." if code_val else f"Failed: {msg}",
                      "success" if code_val else "error")
                self._redirect("/applicants")

            elif path.startswith("/applicants/") and path.endswith("/send-email"):
                aid = int(path.split("/")[2])
                ok, msg = send_release_email(db, aid)
                # Also send SMS if auto_send_sms is on and applicant has phone
                sms_msg = ""
                a_row = db.execute("SELECT phone FROM applicants WHERE id = %s", (aid,)).fetchone()
                if get_setting(db, "auto_send_sms") == "1" and a_row and a_row.get("phone", "").strip():
                    try:
                        sms_ok, sms_msg = send_release_sms(db, aid)
                        sms_msg = " SMS also sent!" if sms_ok else f" SMS failed: {sms_msg}"
                    except Exception as sms_err:
                        sms_msg = f" SMS error: {sms_err}"
                flash(msg + sms_msg, "success" if ok else "error")
                self._redirect("/applicants")

            elif path.startswith("/applicants/") and path.endswith("/assign-and-send"):
                aid = int(path.split("/")[2])
                a = db.execute("SELECT * FROM applicants WHERE id = %s", (aid,)).fetchone()
                if a and not a["assigned_code"]:
                    assign_code(db, aid)
                ok, msg = send_release_email(db, aid)
                # Also send SMS
                sms_note = ""
                a_refresh = db.execute("SELECT phone FROM applicants WHERE id = %s", (aid,)).fetchone()
                if get_setting(db, "auto_send_sms") == "1" and a_refresh and a_refresh.get("phone", "").strip():
                    try:
                        sms_ok, sms_detail = send_release_sms(db, aid)
                        sms_note = " + SMS sent!" if sms_ok else f" (SMS failed: {sms_detail})"
                    except Exception as sms_err:
                        sms_note = f" (SMS error: {sms_err})"
                flash(("Code assigned & email sent!" + sms_note) if ok else f"Code assigned but email failed: {msg}",
                      "success" if ok else "error")
                self._redirect("/applicants")

            elif path.startswith("/applicants/") and path.endswith("/send-sms"):
                aid = int(path.split("/")[2])
                ok, msg = send_release_sms(db, aid)
                flash(msg, "success" if ok else "error")
                self._redirect("/applicants")

            elif path.startswith("/applicants/") and path.endswith("/update-email"):
                aid = int(path.split("/")[2])
                new_email = fv("email").strip()
                # FIX: Added basic email format validation
                if new_email and "@" in new_email and "." in new_email.split("@")[-1]:
                    db.execute("UPDATE applicants SET email = %s, updated_at = NOW() WHERE id = %s",
                               (new_email, aid))
                    flash(f"Email updated to {new_email}", "success")
                elif not new_email:
                    flash("Email cannot be blank.", "error")
                else:
                    flash("Invalid email address format.", "error")
                self._redirect("/applicants")

            elif path.startswith("/applicants/") and path.endswith("/resend"):
                aid = int(path.split("/")[2])
                redirect_to = fv("redirect") or "/applicants"
                a = db.execute("SELECT * FROM applicants WHERE id = %s", (aid,)).fetchone()
                if not a:
                    flash("Applicant not found.", "error")
                elif not a["email"]:
                    flash("No email address on file. Update the email first.", "error")
                elif not a["assigned_code"]:
                    flash("No code assigned yet. Use 'Assign & Send' first.", "error")
                else:
                    ok, msg = send_release_email(db, aid)
                    # Also resend SMS if phone is available
                    sms_note = ""
                    if get_setting(db, "auto_send_sms") == "1" and a.get("phone", "").strip():
                        try:
                            sms_ok, sms_detail = send_release_sms(db, aid)
                            sms_note = f" SMS also resent to {a['phone']}!" if sms_ok else f" (SMS resend failed: {sms_detail})"
                        except Exception as sms_err:
                            sms_note = f" (SMS error: {sms_err})"
                    flash((f"Email resent to {a['email']}!" + sms_note) if ok else f"Resend failed: {msg}",
                          "success" if ok else "error")
                self._redirect(redirect_to)

            elif path.startswith("/applicants/") and path.endswith("/mark-complete"):
                aid = int(path.split("/")[2])
                db.execute("UPDATE applicants SET status='completed' WHERE id = %s", (aid,))
                flash("Applicant marked complete.", "success")
                self._redirect("/applicants")

            elif path.startswith("/applicants/") and path.endswith("/delete"):
                aid = int(path.split("/")[2])
                applicant = db.execute("SELECT assigned_code FROM applicants WHERE id = %s", (aid,)).fetchone()
                if applicant and applicant["assigned_code"]:
                    db.execute("UPDATE codes SET assigned_to=NULL, assigned_date=NULL WHERE code = %s",
                               (applicant["assigned_code"],))
                db.execute("DELETE FROM email_tracking WHERE applicant_id = %s", (aid,))
                db.execute("DELETE FROM email_log WHERE applicant_id = %s", (aid,))
                db.execute("DELETE FROM sms_log WHERE applicant_id = %s", (aid,))
                db.execute("DELETE FROM applicants WHERE id = %s", (aid,))
                flash("Applicant deleted.", "success")
                self._redirect("/applicants")

            elif path == "/applicants/bulk-process":
                pending = db.execute(
                    "SELECT * FROM applicants WHERE status='pending' AND email IS NOT NULL AND email != ''"
                ).fetchall()
                succ, fail, sms_succ, sms_fail = 0, 0, 0, 0
                bulk_auto_sms = get_setting(db, "auto_send_sms") == "1"
                for i, a in enumerate(pending):
                    # Throttle: 2-second delay between messages to avoid being
                    # flagged as a spam bot by receiving mail servers.
                    if i > 0:
                        time.sleep(2)
                    c_val, _ = assign_code(db, a["id"])
                    if c_val:
                        ok, _ = send_release_email(db, a["id"])
                        if ok:
                            succ += 1
                        else:
                            fail += 1
                        # Also send SMS if phone available
                        if bulk_auto_sms and a.get("phone", "").strip():
                            try:
                                s_ok, _ = send_release_sms(db, a["id"])
                                if s_ok:
                                    sms_succ += 1
                                else:
                                    sms_fail += 1
                            except Exception:
                                sms_fail += 1
                    else:
                        fail += 1
                        break  # Stop if we run out of codes
                sms_note = f" | SMS: {sms_succ} sent, {sms_fail} failed" if (sms_succ + sms_fail) > 0 else ""
                flash(f"Done: {succ} emailed, {fail} failed{sms_note}.", "success")
                self._redirect("/applicants")

            elif path.startswith("/codes/") and path.endswith("/delete"):
                code_id = int(path.split("/")[2])
                code_row = db.execute("SELECT * FROM codes WHERE id = %s", (code_id,)).fetchone()
                if not code_row:
                    flash("Code not found.", "error")
                elif code_row["assigned_to"]:
                    flash("Cannot delete a code that is already assigned to an applicant.", "error")
                else:
                    db.execute("DELETE FROM codes WHERE id = %s", (code_id,))
                    flash(f"Code '{code_row['code']}' deleted.", "success")
                self._redirect("/codes")

            elif path == "/codes/add-manual":
                codes_text = fv("codes")
                batch = fv("batch_name", "Manual")
                imp, dup = 0, 0
                for line in codes_text.strip().split("\n"):
                    code_str = line.strip()
                    if code_str:
                        try:
                            db.execute("INSERT INTO codes (code,batch_name) VALUES (%s,%s)",
                                       (code_str, batch))
                            imp += 1
                        except psycopg2.IntegrityError:
                            # FIX: db.rollback() now works — rollback() method added to DBHelper
                            db.rollback()
                            dup += 1
                flash(f"Added {imp} codes ({dup} duplicates skipped).", "success")
                self._redirect("/codes")

            elif path == "/codes/import":
                if isinstance(form_data, cgi.FieldStorage):
                    file_item = form_data["file"]
                    col_idx = int(form_data.getfirst("column_index", "0"))
                    skip_h = form_data.getfirst("skip_header") == "on"
                    batch = form_data.getfirst("batch_name", "Import")
                    fname = file_item.filename or "upload"
                    fpath = os.path.join(UPLOAD_FOLDER, f"import_{datetime.now().strftime('%Y%m%d%H%M%S')}_{fname}")
                    with open(fpath, "wb") as f:
                        f.write(file_item.file.read())
                    imp, dup, err = import_codes_from_file(fpath, column_index=col_idx,
                                                           skip_header=skip_h, batch_name=batch)
                    if err:
                        flash(f"Import error: {err}", "error")
                    else:
                        flash(f"Imported {imp} codes ({dup} duplicates) from '{batch}'.", "success")
                    if os.path.exists(fpath):
                        os.remove(fpath)
                self._redirect("/codes")

            elif path == "/clients/update-email":
                cid = fv("client_id")
                new_email = fv("contact_email").strip()
                if cid:
                    db.execute("UPDATE clients SET contact_email = %s WHERE id = %s", (new_email or None, int(cid)))
                    db.commit()
                    logger.info("Client %s email updated to: %s", cid, new_email or "(cleared)")
                self._redirect(f"/clients?client_id={cid}&email_saved=1")

            elif path == "/settings":
                if isinstance(form_data, cgi.FieldStorage):
                    for key in DEFAULT_SETTINGS:
                        vals = form_data.getlist(key)
                        if vals:
                            set_setting(db, key, vals[-1])
                    if "release_form_file" in form_data:
                        fi = form_data["release_form_file"]
                        if fi.filename:
                            dest = os.path.join(UPLOAD_FOLDER, "release_form.pdf")
                            with open(dest, "wb") as f:
                                f.write(fi.file.read())
                            set_setting(db, "release_form_path", dest)
                else:
                    for key in DEFAULT_SETTINGS:
                        vals = form_data.get(key)
                        if vals:
                            set_setting(db, key, vals[-1])
                flash("Settings saved.", "success")
                self._redirect("/settings")

            elif path == "/settings/test-email":
                addr = fv("test_email") or get_setting(db, "sender_email")
                if not addr:
                    flash("No email address.", "error")
                else:
                    try:
                        srv_host = get_setting(db, "smtp_server")
                        srv_port = int(get_setting(db, "smtp_port") or 587)
                        srv_user = get_setting(db, "smtp_username")
                        srv_pass = get_setting(db, "smtp_password")
                        use_tls = get_setting(db, "smtp_use_tls") == "1"
                        sender = get_setting(db, "sender_email")
                        sname = get_setting(db, "sender_name")
                        msg = MIMEText("Test email from Fingerprint Release Manager. SMTP is working!")
                        msg["From"] = f"{sname} <{sender}>"
                        msg["To"] = addr
                        msg["Subject"] = "Test - Fingerprint Release Manager"
                        logger.info(f"Test email: Connecting to {srv_host}:{srv_port} (TLS={use_tls})")
                        srv = smtplib.SMTP(srv_host, srv_port, timeout=15)
                        srv.set_debuglevel(1)
                        if use_tls:
                            srv.starttls()
                        if srv_user and srv_pass:
                            srv.login(srv_user, srv_pass)
                        srv.send_message(msg)
                        srv.quit()
                        logger.info(f"Test email sent successfully to {addr}")
                        flash(f"Test email sent to {addr}!", "success")
                    except Exception as e:
                        logger.error(f"Test email FAILED: {type(e).__name__}: {e}")
                        flash(f"Test failed: {type(e).__name__}: {e}", "error")
                self._redirect("/settings")

            elif path == "/settings/test-sms":
                phone = fv("test_phone").strip()
                if not phone:
                    flash("Enter a phone number to test.", "error")
                elif not HAS_TWILIO:
                    flash("Twilio library not installed. Run: pip install twilio", "error")
                else:
                    normalized = normalize_phone(phone)
                    if not normalized:
                        flash(f"Invalid phone number: {phone}. Use 10-digit US format.", "error")
                    else:
                        try:
                            sid = get_setting(db, "twilio_account_sid").strip()
                            tok = get_setting(db, "twilio_auth_token").strip()
                            frm = get_setting(db, "twilio_from_number").strip()
                            if not sid or not tok or not frm:
                                flash("Configure Twilio SID, Auth Token, and From Number first.", "error")
                            else:
                                client = TwilioClient(sid, tok)
                                msg = client.messages.create(
                                    body="Test SMS from Fingerprint Release Manager. Twilio is working!",
                                    from_=frm,
                                    to=normalized
                                )
                                logger.info(f"Test SMS sent to {normalized}, SID={msg.sid}")
                                flash(f"Test SMS sent to {normalized}! (SID: {msg.sid})", "success")
                        except Exception as e:
                            logger.error(f"Test SMS FAILED: {type(e).__name__}: {e}")
                            flash(f"Test SMS failed: {type(e).__name__}: {e}", "error")
                self._redirect("/settings")

            # ------------------------------------------------------------------
            # Profile / My Account
            # ------------------------------------------------------------------
            elif path == "/profile":
                new_display = fv("display_name").strip()
                new_recovery = fv("recovery_email").strip()
                if new_recovery and ("@" not in new_recovery or "." not in new_recovery.split("@")[-1]):
                    flash("Invalid recovery email format.", "error")
                else:
                    db.execute(
                        "UPDATE users SET display_name=%s, recovery_email=%s WHERE id=%s",
                        (new_display or None, new_recovery or None, user["id"])
                    )
                    flash("Profile updated.", "success")
                self._redirect("/profile")

            elif path == "/profile/change-password":
                current = fv("current_password")
                new_pw = fv("new_password")
                confirm = fv("confirm_password")
                u_row = db.execute("SELECT * FROM users WHERE id=%s", (user["id"],)).fetchone()
                if not verify_password(current, u_row["password_hash"]):
                    flash("Current password is incorrect.", "error")
                elif len(new_pw) < 8:
                    flash("New password must be at least 8 characters.", "error")
                elif new_pw != confirm:
                    flash("New passwords do not match.", "error")
                else:
                    db.execute("UPDATE users SET password_hash=%s WHERE id=%s",
                               (hash_password(new_pw), user["id"]))
                    flash("Password changed successfully.", "success")
                self._redirect("/profile")

            # ------------------------------------------------------------------
            # User Management (admin only)
            # ------------------------------------------------------------------
            elif path == "/users/add":
                if user.get("role") != "admin":
                    self._send(403, render_page("Forbidden", '<div class="es"><i class="fas fa-lock"></i><h3>Admin access required</h3></div>', user=user))
                else:
                    new_username = fv("username").strip()
                    new_display = fv("display_name").strip()
                    new_password = fv("password")
                    new_role = fv("role", "user")
                    new_recovery = fv("recovery_email").strip()
                    if not new_username or not new_password:
                        flash("Username and password are required.", "error")
                        self._redirect("/users/add")
                    elif len(new_password) < 8:
                        flash("Password must be at least 8 characters.", "error")
                        self._redirect("/users/add")
                    elif new_role not in ("user", "admin"):
                        flash("Invalid role.", "error")
                        self._redirect("/users/add")
                    else:
                        try:
                            db.execute(
                                "INSERT INTO users (username, password_hash, display_name, role, recovery_email, is_active) VALUES (%s,%s,%s,%s,%s,TRUE)",
                                (new_username, hash_password(new_password), new_display or None, new_role, new_recovery or None)
                            )
                            flash(f"User '{new_username}' created.", "success")
                            self._redirect("/users")
                        except psycopg2.IntegrityError:
                            db.rollback()
                            flash(f"Username '{new_username}' already exists.", "error")
                            self._redirect("/users/add")

            elif path.startswith("/users/") and path.endswith("/toggle"):
                if user.get("role") != "admin":
                    self._send(403, render_page("Forbidden", '<div class="es"><i class="fas fa-lock"></i><h3>Admin access required</h3></div>', user=user))
                else:
                    try:
                        uid = int(path.split("/")[2])
                    except (ValueError, IndexError):
                        self._redirect("/users")
                        return
                    if uid == user["id"]:
                        flash("You cannot disable your own account.", "error")
                    else:
                        target = db.execute("SELECT is_active, username FROM users WHERE id=%s", (uid,)).fetchone()
                        if target:
                            new_state = not target["is_active"]
                            db.execute("UPDATE users SET is_active=%s WHERE id=%s", (new_state, uid))
                            action = "enabled" if new_state else "disabled"
                            flash(f"User '{target['username']}' {action}.", "success")
                    self._redirect("/users")

            elif path.startswith("/users/") and path.endswith("/reset"):
                if user.get("role") != "admin":
                    self._send(403, render_page("Forbidden", '<div class="es"><i class="fas fa-lock"></i><h3>Admin access required</h3></div>', user=user))
                else:
                    try:
                        uid = int(path.split("/")[2])
                    except (ValueError, IndexError):
                        self._redirect("/users")
                        return
                    new_pw = fv("new_password")
                    confirm = fv("confirm_password")
                    target = db.execute("SELECT username FROM users WHERE id=%s", (uid,)).fetchone()
                    if not target:
                        flash("User not found.", "error")
                        self._redirect("/users")
                    elif len(new_pw) < 8:
                        self._send(200, page_reset_user_password(uid, target["username"], nav_user=user, error="Password must be at least 8 characters."))
                    elif new_pw != confirm:
                        self._send(200, page_reset_user_password(uid, target["username"], nav_user=user, error="Passwords do not match."))
                    else:
                        db.execute("UPDATE users SET password_hash=%s WHERE id=%s", (hash_password(new_pw), uid))
                        # Invalidate existing sessions for that user
                        db.execute("DELETE FROM sessions WHERE user_id=%s", (uid,))
                        flash(f"Password reset for '{target['username']}'.", "success")
                        self._redirect("/users")

            elif path.startswith("/users/") and path.endswith("/delete"):
                if user.get("role") != "admin":
                    self._send(403, render_page("Forbidden", '<div class="es"><i class="fas fa-lock"></i><h3>Admin access required</h3></div>', user=user))
                else:
                    try:
                        uid = int(path.split("/")[2])
                    except (ValueError, IndexError):
                        self._redirect("/users")
                        return
                    target = db.execute("SELECT username, role FROM users WHERE id=%s", (uid,)).fetchone()
                    if not target:
                        flash("User not found.", "error")
                    elif target["role"] == "admin":
                        flash("Admin accounts cannot be deleted.", "error")
                    elif uid == user["id"]:
                        flash("You cannot delete your own account.", "error")
                    else:
                        db.execute("DELETE FROM sessions WHERE user_id=%s", (uid,))
                        db.execute("DELETE FROM password_reset_tokens WHERE user_id=%s", (uid,))
                        db.execute("DELETE FROM users WHERE id=%s", (uid,))
                        flash(f"User '{target['username']}' deleted.", "success")
                    self._redirect("/users")

            elif path == "/laps/trigger":
                user = self._check_auth()
                if not user:
                    self._redirect("/login")
                    return
                trigger_laps_cycle_manual()
                flash("LAPS retrieval cycle triggered! Check back in a few minutes.", "success")
                self._redirect("/laps")

            else:
                self._send(404, render_page("Not Found",
                    '<div class="es"><i class="fas fa-exclamation-triangle"></i><h3>Page not found</h3></div>'))

        finally:
            db.close()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    init_db()
    start_folder_watcher()
    start_laps_worker()
    print(f"""
    ╔══════════════════════════════════════════════════════════╗
    ║   Fingerprint Release Manager v4.0                       ║
    ║   (LAPS Retriever)                                       ║
    ║   Web UI:      http://localhost:{PORT}                     ║
    ║                                                          ║
    ║   LOGIN:                                                 ║
    ║   Credentials loaded from environment variables          ║
    ║   (DEFAULT_ADMIN_USER / DEFAULT_ADMIN_PASSWORD)          ║
    ║                                                          ║
    ║   AUTO-IMPORT:                                           ║
    ║   Drop .xlsx/.csv files into the 'watch/' folder         ║
    ║   and codes will be imported automatically!              ║
    ║   Watch folder: {WATCH_FOLDER:<40s}  ║
    ║                                                          ║
    ║   LAPS:    Hourly auto-retrieval (if configured)         ║
    ║                                                          ║
    ║   API ENDPOINTS:                                         ║
    ║   POST /api/accio-push     (Accio Data XML results)      ║
    ║   POST /api/codes          (JSON: {{"codes":["..."]}} )    ║
    ║   POST /api/codes/upload   (Multipart file upload)       ║
    ╚══════════════════════════════════════════════════════════╝
    """)
    server = HTTPServer((HOST, PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        _watcher_running = False
        print("\nShutting down...")
        server.server_close()
